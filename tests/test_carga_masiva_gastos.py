"""La carga masiva de gastos: una planilla larga entra de una vez.

Es el mismo circuito que el de consorcios —bajar la plantilla, completarla,
subirla— pero un gasto no crea lo que referencia. Consorcio y unidad tienen
que existir de antes, y la fila que apunta a algo que no está se reporta en
vez de inventarlo.

Lo que estos tests cuidan, por orden de qué duele más si se rompe:
  · que no se dupliquen gastos (la liquidación del mes sale al doble),
  · que una fila mala no tire abajo el archivo entero,
  · que la fila no termine en el edificio de otro administrador.
"""

import io

import openpyxl
import pytest

XLSX = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'

# La copia a mano es a propósito: si alguien toca la plantilla, este test lo
# frena y obliga a decidir si el import también tiene que leer la columna
# nueva. Son las mismas que pide el formulario: proveedor, vencimiento, método
# de pago y comprobante se sacaron de los dos lados a la vez.
HEADERS = ['consorcio*', 'fecha*', 'descripcion*', 'monto*', 'categoria', 'unidad',
           'coeficiente', 'pagado', 'recurrente', 'frecuencia', 'dia_carga', 'notas']

# Los encabezados que traía la plantilla anterior. Una planilla vieja se sube
# igual: las columnas que ya no se piden no tienen alias y se ignoran.
HEADERS_VIEJOS = ['consorcio*', 'fecha*', 'descripcion*', 'monto*', 'categoria', 'proveedor',
                  'unidad', 'fecha_vencimiento', 'pagado', 'fecha_pago', 'metodo_pago',
                  'recurrente', 'frecuencia', 'dia_carga', 'coeficiente',
                  'comprobante_tipo', 'comprobante_numero', 'notas']


@pytest.fixture
def admin(base, app_modulo):
    app_modulo.app.config['TESTING'] = True
    c = app_modulo.app.test_client()
    with c.session_transaction() as s:
        s['user'] = {'sub': 'auth0|admin', 'email': 'admin@test',
                     'name': 'Admin', 'role': 'admin'}
    return c


@pytest.fixture
def datos(base):
    base['gastos'] = []
    return base


def planilla(filas, headers=HEADERS, hoja='Gastos'):
    """Un .xlsx en memoria con esas filas, como el que sube el administrador."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = hoja
    ws.append(headers)
    for fila in filas:
        ws.append(list(fila) + [None] * (len(headers) - len(fila)))
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def subir(client, buf, nombre='gastos.xlsx'):
    return client.post('/api/gastos/carga-masiva',
                       data={'file': (buf, nombre)},
                       content_type='multipart/form-data')


# ── La plantilla ──────────────────────────────────────────────────────────────

def test_la_plantilla_se_baja_con_sus_hojas(admin, datos):
    r = admin.get('/api/gastos/plantilla')
    assert r.status_code == 200
    assert r.mimetype == XLSX
    wb = openpyxl.load_workbook(io.BytesIO(r.data))
    assert wb.sheetnames == ['Instrucciones', 'Gastos', 'Consorcios existentes',
                             'Unidades existentes']


def test_la_plantilla_lista_las_referencias_del_administrador(admin, datos):
    """Las hojas de referencia son de él: el edificio ajeno no aparece.

    Sin el filtro, la plantilla es un listado de los consorcios de todos los
    administradores del sistema entregado en un Excel.
    """
    wb = openpyxl.load_workbook(io.BytesIO(admin.get('/api/gastos/plantilla').data))
    consorcios = [c[0] for c in wb['Consorcios existentes'].iter_rows(min_row=2, values_only=True)]
    assert consorcios == ['Mío']


def test_los_encabezados_de_la_plantilla_son_los_que_lee_el_import(admin, datos):
    """La plantilla que se baja y el parser que la lee no pueden divergir."""
    wb = openpyxl.load_workbook(io.BytesIO(admin.get('/api/gastos/plantilla').data))
    encabezados = [c.value for c in wb['Gastos'][1]]
    assert encabezados == HEADERS


def test_la_fila_de_ejemplo_de_la_plantilla_no_se_importa(admin, datos):
    """Va marcada "(borrar fila)": quien no la borre no se lleva un gasto trucho."""
    r = subir(admin, io.BytesIO(admin.get('/api/gastos/plantilla').data))
    assert r.get_json()['gastos_creados'] == 0
    assert datos['gastos'] == []


# ── El camino feliz ───────────────────────────────────────────────────────────

def test_importa_una_fila_con_solo_los_obligatorios(admin, datos):
    r = subir(admin, planilla([['Mío', '2026-06-05', 'Factura Edesur junio', 15430.50]]))
    assert r.status_code == 200
    d = r.get_json()
    assert (d['gastos_creados'], d['errores']) == (1, [])
    g = datos['gastos'][0]
    assert (g['consorcio_id'], g['fecha_gasto'], g['monto']) == ('cons-1', '2026-06-05', 15430.50)
    # Sin categoría escrita el gasto entra igual, en "otro".
    assert g['categoria'] == 'otro'
    # Los opcionales vacíos quedan en NULL y no en '': `unidad_id` en cadena
    # vacía no es un gasto general, es un UUID inválido que rompe el INSERT.
    assert g['unidad_id'] is None
    # Sin columna "pagado" el gasto entra pagado, igual que en el formulario.
    assert g['pagado'] is True


def test_importa_todos_los_campos_opcionales(admin, datos):
    r = subir(admin, planilla([[
        'Mío', '2026-06-05', 'Factura Edesur junio', 15430.50, 'electricidad',
        '1A', 'E', 'Sí', 'Sí', 'bimestral', 10, 'Factura B-0001-00012345']]))
    assert r.get_json()['errores'] == []
    g = datos['gastos'][0]
    assert g['categoria'] == 'electricidad'
    assert g['unidad_id'] == 'uf-1'
    assert g['pagado'] is True
    assert (g['recurrente'], g['frecuencia'], g['dia_carga']) == (True, 'bimestral', 10)
    assert g['notas'] == 'Factura B-0001-00012345'
    assert g['coeficiente'] == 'E'


def test_el_gasto_queda_a_nombre_del_administrador_que_lo_sube(admin, datos):
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100]]))
    assert datos['gastos'][0]['admin_id'] == 'admin-1'


def test_una_fila_mala_no_frena_a_las_demas(admin, datos):
    """El archivo de 300 filas con tres errores importa 297, no cero."""
    r = subir(admin, planilla([
        ['Mío', '2026-06-05', 'Buena 1', 100],
        ['Mío', '2026-06-05', '', 100],
        ['Mío', '2026-06-06', 'Buena 2', 200],
    ]))
    d = r.get_json()
    assert d['gastos_creados'] == 2
    assert [e['fila'] for e in d['errores']] == [3]
    assert 'descripción' in d['errores'][0]['mensaje'].lower()


def test_el_resumen_cuenta_lo_que_entro(admin, datos):
    d = subir(admin, planilla([
        ['Mío', '2026-06-05', 'Luz', 100],
        ['Mío', '2026-06-06', 'Agua', 250.50],
    ])).get_json()
    assert d['gastos_creados'] == 2
    assert d['consorcios_afectados'] == 1
    assert d['monto_total'] == 350.50


# ── Los obligatorios ──────────────────────────────────────────────────────────

@pytest.mark.parametrize('fila, esperado', [
    (['', '2026-06-05', 'Luz', 100], 'consorcio'),
    (['Mío', '', 'Luz', 100], 'fecha'),
    (['Mío', '2026-06-05', '', 100], 'descripción'),
    (['Mío', '2026-06-05', 'Luz', ''], 'monto'),
])
def test_sin_un_obligatorio_la_fila_se_reporta(admin, datos, fila, esperado):
    d = subir(admin, planilla([fila])).get_json()
    assert d['gastos_creados'] == 0
    assert esperado in d['errores'][0]['mensaje'].lower()
    assert datos['gastos'] == []


def test_el_monto_tiene_que_ser_mayor_a_cero(admin, datos):
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 0],
                               ['Mío', '2026-06-05', 'Gas', -50]])).get_json()
    assert d['gastos_creados'] == 0
    assert len(d['errores']) == 2


def test_una_fecha_ilegible_se_reporta_en_vez_de_inventarse(admin, datos):
    """Poner hoy por defecto haría que 300 gastos de meses distintos caigan
    todos en el período de la carga, y la liquidación los cobraría juntos."""
    d = subir(admin, planilla([['Mío', 'el martes', 'Luz', 100]])).get_json()
    assert d['gastos_creados'] == 0
    assert 'fecha' in d['errores'][0]['mensaje'].lower()


def test_sin_las_columnas_obligatorias_avisa_antes_de_procesar(admin, datos):
    r = subir(admin, planilla([['Mío', 100]], headers=['consorcio', 'monto']))
    assert r.status_code == 400
    msg = r.get_json()['error']
    assert 'fecha' in msg and 'descripcion' in msg


# ── Lo que no se crea de paso ─────────────────────────────────────────────────

def test_el_consorcio_de_otro_administrador_no_existe_para_esta_carga(admin, datos):
    """"Ajeno" es de admin-2. Que la fila entre sería cargarle un gasto a un
    edificio que este administrador no puede ni ver."""
    d = subir(admin, planilla([['Ajeno', '2026-06-05', 'Luz', 100]])).get_json()
    assert d['gastos_creados'] == 0
    assert 'no encontrado' in d['errores'][0]['mensaje'].lower()
    assert datos['gastos'] == []


def test_la_unidad_tiene_que_ser_de_ese_consorcio(admin, datos):
    """`uf-2` existe, pero en el edificio de al lado. El gasto se prorratearía
    en un consorcio y se cobraría en el otro."""
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, '', '2B']])).get_json()
    assert d['gastos_creados'] == 0
    assert 'unidad' in d['errores'][0]['mensaje'].lower()


# ── Deduplicación ─────────────────────────────────────────────────────────────

def test_no_duplica_contra_lo_que_ya_estaba(admin, datos):
    datos['gastos'].append({
        'id': 'gasto-1', 'admin_id': 'admin-1', 'consorcio_id': 'cons-1',
        'descripcion': 'Factura Edesur junio', 'monto': 15430.50,
        'fecha_gasto': '2026-06-05', 'categoria': 'electricidad'})
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Factura Edesur junio', 15430.50]])).get_json()
    assert (d['gastos_creados'], d['gastos_omitidos']) == (0, 1)
    assert len(datos['gastos']) == 1


def test_subir_dos_veces_el_mismo_archivo_no_carga_todo_dos_veces(admin, datos):
    """El reintento después de corregir errores es el caso normal, no el raro."""
    filas = [['Mío', '2026-06-05', 'Luz', 100], ['Mío', '2026-06-06', 'Agua', 200]]
    assert subir(admin, planilla(filas)).get_json()['gastos_creados'] == 2
    d = subir(admin, planilla(filas)).get_json()
    assert (d['gastos_creados'], d['gastos_omitidos']) == (0, 2)
    assert len(datos['gastos']) == 2


def test_la_fila_repetida_dentro_del_archivo_entra_una_sola_vez(admin, datos):
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100],
                               ['Mío', '2026-06-05', 'Luz', 100]])).get_json()
    assert (d['gastos_creados'], d['gastos_omitidos']) == (1, 1)


def test_dos_gastos_distintos_del_mismo_dia_entran_los_dos(admin, datos):
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100],
                               ['Mío', '2026-06-05', 'Gas', 100]])).get_json()
    assert d['gastos_creados'] == 2


# ── Cómo se escriben los datos en la vida real ────────────────────────────────

@pytest.mark.parametrize('escrito, esperado', [
    ('2026-06-05', '2026-06-05'),
    ('05/06/2026', '2026-06-05'),
    ('05-06-2026', '2026-06-05'),
])
def test_acepta_las_fechas_como_las_escribe_la_gente(admin, datos, escrito, esperado):
    subir(admin, planilla([['Mío', escrito, 'Luz', 100]]))
    assert datos['gastos'][0]['fecha_gasto'] == esperado


def test_acepta_la_fecha_que_excel_devuelve_como_datetime(admin, datos):
    """La celda con formato fecha no llega como texto: llega como datetime."""
    from datetime import datetime
    subir(admin, planilla([['Mío', datetime(2026, 6, 5), 'Luz', 100]]))
    assert datos['gastos'][0]['fecha_gasto'] == '2026-06-05'


@pytest.mark.parametrize('escrito, esperado', [
    (15430.50, 15430.50),
    ('15430.50', 15430.50),
    ('15.430,50', 15430.50),     # separador de miles argentino
    ('$ 15.430,50', 15430.50),   # copiado y pegado de la factura
    ('15,430.50', 15430.50),     # el mismo Excel en inglés
])
def test_acepta_el_monto_como_lo_pega_el_administrador(admin, datos, escrito, esperado):
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', escrito]]))
    assert datos['gastos'][0]['monto'] == esperado


@pytest.mark.parametrize('escrito, esperado', [
    ('Sí', True), ('si', True), ('SI', True), ('X', True), ('1', True),
    ('No', False),
    # La celda vacía se toma como pagado, igual que el tilde del formulario:
    # el gasto se carga cuando ya se pagó.
    ('', True), (None, True),
])
def test_el_pagado_se_escribe_de_muchas_formas(admin, datos, escrito, esperado):
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, '', '', '', escrito]]))
    assert datos['gastos'][0]['pagado'] is esperado


def test_un_si_o_no_ilegible_se_reporta(admin, datos):
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, '', '', '', 'tal vez']])).get_json()
    assert d['gastos_creados'] == 0
    assert 'pagado' in d['errores'][0]['mensaje'].lower()


def test_una_categoria_que_no_esta_en_la_lista_cae_en_otro(admin, datos):
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, 'iluminacion']]))
    assert datos['gastos'][0]['categoria'] == 'otro'


def test_la_frecuencia_y_el_dia_solo_valen_en_un_gasto_recurrente(admin, datos):
    """Un gasto común con día de carga es una plantilla que no genera nada y
    queda esperando a que alguien tilde "recurrente" sin querer."""
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, '', '', '', '', '', '', '',
                            'No', 'mensual', 10]]))
    g = datos['gastos'][0]
    assert (g['recurrente'], g['frecuencia'], g['dia_carga']) == (False, '', None)


def test_un_recurrente_sin_frecuencia_queda_mensual(admin, datos):
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, '', '', '', '', 'Sí']]))
    assert datos['gastos'][0]['frecuencia'] == 'mensual'


def test_un_dia_de_carga_fuera_del_mes_no_se_guarda(admin, datos):
    subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100, '', '', '', '',
                            'Sí', 'mensual', 45]]))
    assert datos['gastos'][0]['dia_carga'] is None


# ── El archivo exportado se vuelve a subir ────────────────────────────────────

def test_el_excel_que_baja_el_boton_exportar_se_puede_reimportar(admin, datos):
    """Los encabezados del export ("Fecha", "Categoría") normalizan a las
    mismas claves que la plantilla, así que exportar-editar-subir funciona."""
    headers = ['Fecha', 'Consorcio', 'Descripción', 'Categoría',
               'Monto', 'Pagado', 'Recurrente']
    r = subir(admin, planilla([['2026-06-05', 'Mío', 'Factura Edesur junio',
                                'electricidad', 15430.50, 'Sí', 'No']],
                              headers=headers, hoja='Gastos'))
    assert r.get_json()['errores'] == []
    g = datos['gastos'][0]
    assert (g['fecha_gasto'], g['descripcion'], g['monto']) == ('2026-06-05', 'Factura Edesur junio', 15430.50)
    assert (g['categoria'], g['pagado']) == ('electricidad', True)


def test_el_orden_de_las_columnas_no_importa(admin, datos):
    """Se lee por encabezado: correr una columna no desplaza todos los datos."""
    r = subir(admin, planilla([[15430.50, 'Factura Edesur junio', 'Mío', '2026-06-05']],
                              headers=['monto', 'descripcion', 'consorcio', 'fecha']))
    assert r.get_json()['gastos_creados'] == 1
    assert datos['gastos'][0]['monto'] == 15430.50


def test_si_no_hay_hoja_gastos_usa_la_primera(admin, datos):
    """El export trae una sola hoja y no se llama "Gastos"."""
    r = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100]], hoja='Sheet1'))
    assert r.get_json()['gastos_creados'] == 1


# ── Bordes ────────────────────────────────────────────────────────────────────

def test_sin_archivo_avisa(admin, datos):
    r = admin.post('/api/gastos/carga-masiva', data={}, content_type='multipart/form-data')
    assert r.status_code == 400


def test_un_archivo_que_no_es_excel_avisa_en_vez_de_romper(admin, datos):
    r = subir(admin, io.BytesIO(b'esto no es un xlsx'), 'gastos.xlsx')
    assert r.status_code == 400
    assert 'plantilla' in r.get_json()['error'].lower()


def test_las_filas_vacias_del_medio_se_saltean(admin, datos):
    r = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100],
                               [None, None, None, None],
                               ['Mío', '2026-06-06', 'Agua', 200]]))
    d = r.get_json()
    assert (d['gastos_creados'], d['errores']) == (2, [])


def test_el_vecino_no_puede_importar_gastos(client, datos):
    """`client` es la sesión de un vecino: la carga masiva es del administrador."""
    r = subir(client, planilla([['Mío', '2026-06-05', 'Luz', 100]]))
    assert r.status_code in (302, 403, 404)
    assert datos['gastos'] == []


def test_el_vecino_tampoco_baja_la_plantilla(client, datos):
    r = client.get('/api/gastos/plantilla')
    assert r.status_code in (302, 403, 404)


def test_la_validacion_por_rango_no_lleva_el_igual_adelante(admin, datos):
    """`<formula1>` guarda la referencia sin "=": con el signo, Excel abre el
    archivo diciendo que encontró contenido ilegible y lo "repara" borrando la
    validación. openpyxl lo escribe tal cual se lo pasan y no avisa nada.
    """
    import re
    import zipfile
    z = zipfile.ZipFile(io.BytesIO(admin.get('/api/gastos/plantilla').data))
    hoja = z.read('xl/worksheets/sheet2.xml').decode()
    formulas = re.findall(r'<formula1>(.*?)</formula1>', hoja, re.S)
    assert formulas, 'la hoja "Gastos" perdió sus desplegables'
    assert not [f for f in formulas if f.startswith('=')]
    assert "'Consorcios existentes'!$A$2:$A$2" in formulas


def test_la_deduplicacion_no_mira_fuera_del_rango_del_archivo(admin, datos):
    """Se consultan sólo las fechas que trae el archivo.

    Traerse el historial completo es lo que rompe con volumen: PostgREST corta
    en 1000 filas y la lista truncada dejaría entrar como nuevos justo los
    gastos viejos. El de enero no compite con una importación de junio.
    """
    datos['gastos'].append({
        'id': 'gasto-viejo', 'admin_id': 'admin-1', 'consorcio_id': 'cons-1',
        'descripcion': 'Luz', 'monto': 100, 'fecha_gasto': '2026-01-05'})
    d = subir(admin, planilla([['Mío', '2026-06-05', 'Luz', 100]])).get_json()
    assert (d['gastos_creados'], d['gastos_omitidos']) == (1, 0)


def test_la_plantilla_que_se_baja_se_completa_y_se_sube(admin, datos):
    """El circuito entero, sin planilla armada a mano.

    Los tests de arriba fabrican el .xlsx con los encabezados escritos acá: si
    la plantilla y el parser se separaran, seguirían pasando todos. Este es el
    único que agarra esa divergencia, porque el archivo sale del endpoint.
    """
    wb = openpyxl.load_workbook(io.BytesIO(admin.get('/api/gastos/plantilla').data))
    ws = wb['Gastos']
    ws.delete_rows(2)  # la fila de ejemplo, como haría el administrador
    ws.append(['Mío', '2026-06-05', 'Factura Edesur junio', 15430.50, 'electricidad',
               '1A', 'E', 'Sí', 'No', None, None, 'Factura B-0001-00012345'])
    ws.append(['Mío', '2026-06-01', 'Limpieza junio', 92000, 'limpieza', None,
               None, 'No', 'Sí', 'mensual', 5, None])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    d = subir(admin, buf).get_json()
    assert (d['gastos_creados'], d['errores']) == (2, [])
    assert d['monto_total'] == 107430.50
    luz, limpieza = datos['gastos']
    assert (luz['unidad_id'], luz['coeficiente'], luz['pagado']) == ('uf-1', 'E', True)
    assert (limpieza['recurrente'], limpieza['frecuencia'], limpieza['dia_carga']) == (True, 'mensual', 5)


# ── Coeficiente ──────────────────────────────────────────────────────────────
# La carga masiva y el prorrateo por coeficientes se hicieron en paralelo y no
# se conocían: trescientos gastos importados de una planilla entraban todos por
# el reparto A.

def _fila(**extra):
    """Una fila mínima, con las columnas que se quieran encima."""
    base = {'consorcio': 'Mío', 'fecha': '2026-08-01',
            'descripcion': 'Un gasto', 'monto': 1000}
    base.update(extra)
    return [base.get(h.rstrip('*'), '') for h in HEADERS]


@pytest.mark.parametrize('escrito,guardado', [
    ('E', 'E'),
    ('b', 'B'),          # se normaliza a mayúscula
    ('Z', 'A'),          # inválido: cae en A en vez de voltear la fila
    ('', 'A'),           # planilla vieja, sin la columna
])
def test_el_coeficiente_se_importa_y_se_normaliza(admin, datos, escrito, guardado):
    r = subir(admin, planilla([_fila(coeficiente=escrito)]))
    assert r.get_json()['errores'] == []
    assert datos['gastos'][0]['coeficiente'] == guardado


def test_una_letra_mal_tipeada_no_frena_el_resto_del_archivo(admin, datos):
    """Es la diferencia entre importar 299 gastos y no importar ninguno."""
    r = subir(admin, planilla([
        _fila(descripcion='Uno', coeficiente='Z'),
        _fila(descripcion='Dos', coeficiente='E'),
    ]))
    assert r.get_json()['errores'] == []
    assert [g['coeficiente'] for g in datos['gastos']] == ['A', 'E']


# ── La planilla vieja ────────────────────────────────────────────────────────

def test_la_planilla_anterior_se_sube_y_sus_columnas_de_mas_se_ignoran(admin, datos):
    """El administrador que tiene la plantilla vieja bajada no queda a pie.

    Proveedor, vencimiento, método de pago y comprobante perdieron su alias: la
    columna entra al archivo y el parser no la mira. Frenar la fila por un
    proveedor que ya no se puede dar de alta sería dejarlo sin salida.
    """
    fila = ['Mío', '2026-06-05', 'Factura Edesur junio', 15430.50, 'electricidad',
            'Edesur', '1A', '2026-06-20', 'Sí', '2026-06-18', 'transferencia',
            'No', None, None, 'E', 'LSP A', '0014-23144929', 'Una nota']
    r = subir(admin, planilla([fila], headers=HEADERS_VIEJOS))
    assert r.get_json()['errores'] == []
    g = datos['gastos'][0]
    assert (g['unidad_id'], g['coeficiente'], g['pagado']) == ('uf-1', 'E', True)
    assert g['notas'] == 'Una nota'
    assert 'proveedor_id' not in g and 'comprobante_numero' not in g
