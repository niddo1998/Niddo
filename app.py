import io
import os
import json
import base64
import re
from datetime import datetime, timezone, date, timedelta
from typing import Optional
from functools import wraps

from flask import (
    Flask, render_template, redirect, url_for,
    session, request, jsonify, send_file, Response, g, abort, make_response
)
from authlib.integrations.flask_client import OAuth
from markupsafe import escape
from dotenv import load_dotenv
from supabase import create_client, Client

from recurrentes import periodos_pendientes


# ── Config ─────────────────────────────────────────────────────────────────────
load_dotenv()

# Resolve paths relative to *this* file so the app works both locally
# and when imported from api/index.py on Vercel.
_HERE = os.path.dirname(os.path.abspath(__file__))

app = Flask(
    __name__,
    template_folder=os.path.join(_HERE, 'templates'),
    static_folder=os.path.join(_HERE, 'static'),
)
app.secret_key = os.environ['SECRET_KEY']
# Solo en desarrollo la cookie viaja por HTTP. En producción va marcada Secure:
# con el panel de superadmin la sesión de una de esas cuentas es la credencial
# más sensible del sistema y no puede salir en claro.
_ES_LOCAL = os.environ.get('FLASK_ENV') == 'development' or os.environ.get('NIDDO_LOCAL') == '1'
app.config['SESSION_COOKIE_SECURE'] = not _ES_LOCAL
app.config['SESSION_COOKIE_HTTPONLY'] = True
app.config['SESSION_COOKIE_SAMESITE'] = 'Lax'
app.config['MAX_CONTENT_LENGTH'] = 10 * 1024 * 1024  # 10 MB

# Fix for reverse proxy headers (Vercel) so url_for uses HTTPS
from werkzeug.middleware.proxy_fix import ProxyFix
app.wsgi_app = ProxyFix(app.wsgi_app, x_for=1, x_proto=1, x_host=1, x_prefix=1)

# ── Auth0 ──────────────────────────────────────────────────────────────────────
AUTH0_DOMAIN = os.environ['AUTH0_DOMAIN']

oauth = OAuth(app)
auth0 = oauth.register(
    'auth0',
    client_id=os.environ['AUTH0_CLIENT_ID'],
    client_secret=os.environ['AUTH0_CLIENT_SECRET'],
    client_kwargs={'scope': 'openid profile email'},
    server_metadata_url=f'https://{AUTH0_DOMAIN}/.well-known/openid-configuration',
)

# ── Supabase ───────────────────────────────────────────────────────────────────
supabase: Client = create_client(
    os.environ['SUPABASE_URL'],
    os.environ['SUPABASE_SERVICE_KEY'],
)

# ── Helpers ────────────────────────────────────────────────────────────────────
def now_iso() -> str:
    return datetime.now(timezone.utc).isoformat()


def upsert_user(role: str, auth0_id: str, email: str, nombre: str) -> Optional[dict]:
    """Crea o refresca la fila del usuario y la devuelve.

    El payload no incluye `estado` ni `es_superadmin` a propósito: en un
    ON CONFLICT solo se pisan las columnas que van en el payload, así que el
    estado de aprobación sobrevive a cada login. Si se mandaran, cada vez que un
    administrador aprobado volviera a entrar se le reiniciaría el estado.
    """
    table = 'administradores' if role == 'admin' else 'vecinos'
    res = supabase.table(table).upsert(
        {'auth0_id': auth0_id, 'email': email, 'nombre': nombre, 'last_login': now_iso()},
        on_conflict='auth0_id'
    ).execute()

    if role == 'vecino' and res.data:
        vecino_id = res.data[0]['id']
        current_consorcio = res.data[0].get('consorcio_id')
        # Auto-asociación por email: el administrador cargó esta unidad en el
        # padrón con este mail, así que ya dijo de quién es. Es el único camino
        # que entra sin pasar por la aprobación, y está bien que lo sea: la
        # verificación la hizo el admin cuando escribió el mail.
        if not current_consorcio:
            uf_res = supabase.table('unidades_funcionales').select('id, consorcio_id, numero').eq('vecino_email', email).is_('vecino_id', 'null').execute()
            if uf_res.data:
                uf = uf_res.data[0]
                supabase.table('vecinos').update({
                    'consorcio_id': uf['consorcio_id'],
                    'unidad': uf['numero'],
                    # Faltaba, y era lo que rompía este camino: todo lo que ve
                    # el vecino cuelga de `unidad_id`, no del número. Sin esto
                    # entraba al panel de su edificio y no veía una sola expensa.
                    'unidad_id': uf['id'],
                    # Queda 'aprobado' para que la fila no mienta: tiene
                    # consorcio_id, o sea que está adentro.
                    'estado_asociacion': 'aprobado',
                }).eq('id', vecino_id).execute()
                supabase.table('unidades_funcionales').update({
                    'vecino_id': vecino_id
                }).eq('id', uf['id']).execute()

    return res.data[0] if res.data else None


# ── Estado del administrador en sesión ────────────────────────────────────────
# Campos de `administradores` que necesita cualquier request de un admin. Se
# piden juntos para que el gate de aprobación no cueste una query extra: antes
# `get_admin_id()` ya hacía este mismo SELECT en cada ruta.
_CAMPOS_ADMIN = (
    'id, estado, es_superadmin, nombre, email, nombre_contacto, telefono, '
    'email_contacto, motivo_rechazo, solicitud_completada_at, ultima_actividad'
)

INACTIVO_DIAS = 30          # sin usar la plataforma en este plazo = inactivo
ACTIVIDAD_FRESCA_SEG = 3600  # cada cuánto se refresca ultima_actividad


def _fila_admin(campo: str, valor: str) -> Optional[dict]:
    res = supabase.table('administradores').select(_CAMPOS_ADMIN).eq(campo, valor).execute()
    return res.data[0] if res.data else None


def _tocar_actividad(fila: Optional[dict]) -> None:
    """Refresca `ultima_actividad`, como mucho una vez por hora.

    `last_login` solo se escribe cuando alguien pasa por Auth0, así que quien
    trabaja con la sesión abierta figuraba inactivo hace semanas. Esto arregla
    la señal sin pagar un UPDATE por carga de página: saber la hora exacta del
    último uso no vale más que saberla con una hora de margen.
    """
    if not fila or not fila.get('id'):
        return
    previo = fila.get('ultima_actividad')
    if previo:
        try:
            visto = datetime.fromisoformat(previo.replace('Z', '+00:00'))
            if (datetime.now(timezone.utc) - visto).total_seconds() < ACTIVIDAD_FRESCA_SEG:
                return
        except ValueError:
            pass  # valor ilegible: se reescribe y queda sano
    ahora = now_iso()
    try:
        supabase.table('administradores').update({'ultima_actividad': ahora}) \
            .eq('id', fila['id']).execute()
        fila['ultima_actividad'] = ahora
    except Exception:
        app.logger.exception('No se pudo refrescar ultima_actividad de %s', fila['id'])


def cargar_superadmin_real() -> Optional[dict]:
    """La fila de quien está realmente logueado, ignorando la impersonación.

    Es la que manda para los permisos: mientras se mira la plataforma como otro
    administrador, los privilegios siguen siendo los de la persona detrás.
    """
    if 'admin_real' in g:
        return g.admin_real
    g.admin_real = None
    user = session.get('user')
    if user and user.get('role') == 'admin':
        g.admin_real = _fila_admin('auth0_id', user['sub'])
        # Solo se registra la actividad de la persona real: si se registrara la
        # del impersonado, mirarlo lo haría figurar como activo.
        _tocar_actividad(g.admin_real)
    return g.admin_real


def cargar_admin_actual() -> Optional[dict]:
    """La fila cuyos datos se están viendo, cacheada por request.

    Con impersonación activa devuelve la del administrador impersonado, y con
    eso las ~40 rutas de datos que ya existen quedan apuntando a él sin tocar
    ninguna: todas pasan por `get_admin_id()`.

    El estado de aprobación se lee siempre de la base y nunca de la sesión: si
    se guardara en la cookie, revocarle el acceso a alguien no tendría efecto
    hasta que cerrara sesión, y a quien acabamos de aprobar habría que pedirle
    que se deslogueara para entrar.
    """
    if 'admin_row' in g:
        return g.admin_row

    real = cargar_superadmin_real()
    imp = session.get('impersonar')
    # El es_superadmin se revalida contra la base en cada request: si a alguien
    # le revocan el privilegio con una sesión de impersonación abierta, vuelve
    # a ver sus propios datos en el request siguiente.
    if imp and real and real.get('es_superadmin') and real.get('estado') == 'aprobado':
        g.admin_row = _fila_admin('id', imp['id'])
    else:
        g.admin_row = real
    return g.admin_row


def get_admin_id() -> Optional[str]:
    """Devuelve el UUID de la fila en `administradores` para el usuario en sesión."""
    row = cargar_admin_actual()
    return row['id'] if row else None


def get_vecino_id() -> Optional[str]:
    """Devuelve el UUID de la fila en `vecinos` para el vecino en sesión."""
    user = session.get('user')
    if not user:
        return None
    result = supabase.table('vecinos').select('id').eq('auth0_id', user['sub']).execute()
    return result.data[0]['id'] if result.data else None


def excel_response(wb, filename: str) -> Response:
    import openpyxl
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     download_name=filename, as_attachment=True)


def pdf_response(buf: io.BytesIO, filename: str) -> Response:
    buf.seek(0)
    return send_file(buf, mimetype='application/pdf', download_name=filename, as_attachment=True)


def make_excel(headers: list, rows: list, sheet_name: str):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 14)
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    return wb


TIPOS_UF_VALIDOS = ['departamento', 'local', 'cochera', 'baulera']
MARCA_FILA_EJEMPLO = '(borrar fila)'


def es_fila_ejemplo(texto: str) -> bool:
    return MARCA_FILA_EJEMPLO in (texto or '').lower()


def build_carga_masiva_template(consorcios_existentes: list):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    example_font = Font(italic=True, color="9CA3AF")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    wb = openpyxl.Workbook()

    ws_info = wb.active
    ws_info.title = 'Instrucciones'
    ws_info.column_dimensions['A'].width = 100
    info_lines = [
        ('Carga masiva de Consorcios y Unidades Funcionales', True),
        ('', False),
        ('1. Completá la hoja "Consorcios" para crear edificios nuevos. Dejala vacía si solo vas a cargar', False),
        ('   unidades de consorcios que ya existen.', False),
        ('2. Completá la hoja "Unidades" con las UF a cargar. En la columna "consorcio" escribí el nombre', False),
        ('   exacto del consorcio (nuevo, tal como lo escribiste en la hoja "Consorcios", o uno ya existente,', False),
        ('   tal como figura en la hoja "Consorcios existentes").', False),
        ('3. Guardá el archivo y subilo en el panel. No cambies los nombres de las hojas ni de las columnas.', False),
        ('', False),
        ('Campos obligatorios: nombre (Consorcios); consorcio y numero (Unidades). El resto es opcional.', False),
        (f'Valores válidos para "tipo": {", ".join(TIPOS_UF_VALIDOS)}.', False),
        ('Si un consorcio o una unidad ya existe, se reutiliza/omite automáticamente (no se duplica).', False),
    ]
    for i, (text, bold) in enumerate(info_lines, 1):
        cell = ws_info.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13)

    ws_c = wb.create_sheet('Consorcios')
    style_header(ws_c, ['nombre*', 'direccion', 'cuit', 'pisos', 'unidades_totales', 'encargado_nombre', 'encargado_tel'])
    example_c = ['Edificio Ejemplo 123 (borrar fila)', 'Av. Siempreviva 742', '30-12345678-9', 8, 24, 'Juan Pérez', '+54 9 11 1234-5678']
    for c, val in enumerate(example_c, 1):
        ws_c.cell(row=2, column=c, value=val).font = example_font

    ws_u = wb.create_sheet('Unidades')
    style_header(ws_u, ['consorcio*', 'numero*', 'piso', 'tipo', 'superficie_m2', 'vecino_nombre', 'vecino_email'])
    example_u = ['Edificio Ejemplo 123 (borrar fila)', '3B', '3', 'departamento', 65.5, 'Juan Pérez', 'juan@mail.com']
    for c, val in enumerate(example_u, 1):
        ws_u.cell(row=2, column=c, value=val).font = example_font
    tipo_dv = DataValidation(type='list', formula1=f'"{",".join(TIPOS_UF_VALIDOS)}"', allow_blank=True, showErrorMessage=False)
    ws_u.add_data_validation(tipo_dv)
    tipo_dv.add('D2:D1000')
    if consorcios_existentes:
        nombres = [c['nombre'] for c in consorcios_existentes]
        con_dv = DataValidation(type='list', formula1=f'"{",".join(nombres)[:255]}"', allow_blank=True, showErrorMessage=False)
        ws_u.add_data_validation(con_dv)
        con_dv.add('A2:A1000')

    ws_ref = wb.create_sheet('Consorcios existentes')
    style_header(ws_ref, ['nombre', 'direccion'])
    for r, c in enumerate(consorcios_existentes, 2):
        ws_ref.cell(row=r, column=1, value=c['nombre'])
        ws_ref.cell(row=r, column=2, value=c.get('direccion', ''))
    if not consorcios_existentes:
        ws_ref.cell(row=2, column=1, value='(todavía no tenés consorcios cargados)').font = example_font

    wb.active = 0
    return wb


# ── Carga masiva de gastos ────────────────────────────────────────────────────
# Mismo circuito que la de consorcios —bajás un .xlsx, lo completás, lo subís—
# pero con dos diferencias que salen de qué es un gasto:
#
#  1. El gasto no crea nada de lo que referencia. Consorcio, proveedor y unidad
#     tienen que existir de antes: el archivo los busca por nombre y la fila que
#     apunta a algo que no está se reporta en vez de inventarlo. Un proveedor
#     nuevo por un typo ensucia el maestro para siempre y nadie lo nota.
#  2. Las columnas se leen por encabezado y no por posición, con alias. Así el
#     Excel que baja "Exportar gastos" se puede volver a subir: sus títulos
#     ("Fecha", "Método Pago") normalizan a las mismas claves que la plantilla.

CATEGORIAS_GASTO = ('electricidad', 'gas', 'agua', 'limpieza', 'ascensor', 'seguro',
                    'honorarios', 'impuesto', 'mantenimiento', 'sueldos', 'otro')

# Los coeficientes de reparto. Cada gasto declara por cuál se prorratea y cada
# unidad tiene su porcentaje para cada uno: es lo que permite que el ascensor
# vaya por uno donde la cochera pesa 0 y el seguro por otro donde pesa.
COEFICIENTES = ('A', 'B', 'C', 'E')
METODOS_PAGO_GASTO = ('transferencia', 'cheque', 'efectivo', 'debito')
FRECUENCIAS_GASTO = ('mensual', 'bimestral', 'trimestral', 'anual')

# Encabezado normalizado -> clave canónica. Los alias no son cortesía: cubren
# los títulos del export ("fecha", "metodo_pago") y las formas en que la misma
# columna aparece escrita en las planillas que ya usan los administradores.
ALIAS_COLUMNAS_GASTO = {
    'consorcio': 'consorcio', 'edificio': 'consorcio',
    'fecha': 'fecha_gasto', 'fecha_gasto': 'fecha_gasto',
    'descripcion': 'descripcion', 'detalle': 'descripcion', 'concepto': 'descripcion',
    'monto': 'monto', 'importe': 'monto', 'total': 'monto',
    'categoria': 'categoria', 'rubro': 'categoria',
    'proveedor': 'proveedor',
    'unidad': 'unidad', 'uf': 'unidad', 'unidad_funcional': 'unidad',
    'fecha_vencimiento': 'fecha_vencimiento', 'vencimiento': 'fecha_vencimiento', 'vto': 'fecha_vencimiento',
    'pagado': 'pagado',
    'fecha_pago': 'fecha_pago',
    'metodo_pago': 'metodo_pago', 'metodo_de_pago': 'metodo_pago', 'forma_pago': 'metodo_pago',
    'recurrente': 'recurrente',
    'frecuencia': 'frecuencia',
    'dia_carga': 'dia_carga', 'dia_de_carga': 'dia_carga',
    'notas': 'notas', 'observaciones': 'notas',
    # Las tres que la liquidación necesita. Sin coeficiente, trescientos gastos
    # cargados por planilla entran todos por el reparto A; sin comprobante, el
    # PDF imprime "—" en la columna que la Ley 941 obliga a rendir.
    'coeficiente': 'coeficiente', 'coef': 'coeficiente',
    'comprobante_tipo': 'comprobante_tipo', 'tipo_comprobante': 'comprobante_tipo',
    'tipo_de_comprobante': 'comprobante_tipo',
    'comprobante_numero': 'comprobante_numero', 'numero_comprobante': 'comprobante_numero',
    'nro_comprobante': 'comprobante_numero', 'comprobante': 'comprobante_numero',
}

COLUMNAS_GASTO_OBLIGATORIAS = ('consorcio', 'fecha_gasto', 'descripcion', 'monto')

HEADERS_PLANTILLA_GASTOS = [
    'consorcio*', 'fecha*', 'descripcion*', 'monto*', 'categoria', 'proveedor',
    'unidad', 'fecha_vencimiento', 'pagado', 'fecha_pago', 'metodo_pago',
    'recurrente', 'frecuencia', 'dia_carga', 'coeficiente',
    'comprobante_tipo', 'comprobante_numero', 'notas',
]

FORMATOS_FECHA = ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y', '%d/%m/%y', '%Y/%m/%d')
_TEXTO_VERDADERO = {'si', 'sí', 's', 'x', 'true', 'verdadero', '1', 'yes', 'y'}
_TEXTO_FALSO = {'no', 'n', 'false', 'falso', '0', '-', ''}


def _norm_col(nombre) -> str:
    """El encabezado sin acentos, sin asterisco y con guiones bajos.

    "Método Pago", "metodo_pago" y "Método de pago*" tienen que caer todos en la
    misma clave: es lo que deja que el Excel exportado se vuelva a importar sin
    que nadie renombre columnas a mano.
    """
    import unicodedata
    txt = unicodedata.normalize('NFKD', str(nombre or '')).encode('ascii', 'ignore').decode()
    txt = txt.strip().lower().replace('*', '')
    for ch in (' ', '-', '.', '/'):
        txt = txt.replace(ch, '_')
    while '__' in txt:
        txt = txt.replace('__', '_')
    return txt.strip('_')


def leer_filas_por_encabezado(ws, alias: dict):
    """(columnas_encontradas, [(fila_excel, {clave: valor})]) de una hoja.

    Leer por encabezado y no por posición es lo que hace que agregar, correr o
    sacar una columna del archivo no desplace todos los datos una casilla.
    """
    primera = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), ()) or ()
    columnas = {}
    for idx, celda in enumerate(primera):
        canon = alias.get(_norm_col(celda))
        if canon and canon not in columnas:
            columnas[canon] = idx
    filas = []
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not row or all(v in (None, '') for v in row):
            continue
        filas.append((i, {k: (row[idx] if idx < len(row) else None) for k, idx in columnas.items()}))
    return columnas, filas


def _texto_celda(valor) -> str:
    """El contenido de la celda como texto limpio. Los enteros de Excel vienen
    como float, y una UF "3.0" no matchea con la unidad "3" que hay guardada."""
    if valor is None:
        return ''
    if isinstance(valor, float) and valor.is_integer():
        return str(int(valor))
    return str(valor).strip()


def _fecha_excel(valor):
    """Fecha de celda como 'YYYY-MM-DD', o None si está vacía.

    Excel devuelve datetime cuando la celda tiene formato fecha y texto cuando
    la escribieron a mano; las dos formas llegan del mismo archivo.
    """
    if valor in (None, ''):
        return None
    if isinstance(valor, datetime):
        return valor.date().isoformat()
    if isinstance(valor, date):
        return valor.isoformat()
    txt = str(valor).strip().split(' ')[0]
    if not txt:
        return None
    for formato in FORMATOS_FECHA:
        try:
            return datetime.strptime(txt, formato).date().isoformat()
        except ValueError:
            continue
    raise ValueError(f'Fecha inválida: "{valor}". Usá el formato AAAA-MM-DD.')


def _monto_excel(valor) -> float:
    """El monto como número, tolerando el separador de miles argentino.

    "$ 15.430,50" es lo que sale de copiar una factura, y `float()` lo lee como
    quince mil cuatrocientos treinta con cincuenta *centavos de más* o rompe.
    """
    if valor in (None, ''):
        raise ValueError('Falta el monto')
    if isinstance(valor, bool):
        raise ValueError(f'Monto inválido: "{valor}"')
    if isinstance(valor, (int, float)):
        monto = float(valor)
    else:
        txt = str(valor).strip().replace('$', '').replace(' ', '').replace('\xa0', '')
        if ',' in txt and '.' in txt:
            # El último separador es el decimal; el otro es de miles.
            txt = txt.replace('.', '').replace(',', '.') if txt.rfind(',') > txt.rfind('.') else txt.replace(',', '')
        elif ',' in txt:
            txt = txt.replace(',', '.')
        try:
            monto = float(txt)
        except ValueError:
            raise ValueError(f'Monto inválido: "{valor}"')
    if monto <= 0:
        raise ValueError(f'El monto tiene que ser mayor a cero (vino "{valor}")')
    return round(monto, 2)


def _bool_excel(valor, campo: str) -> bool:
    """Sí/No de una celda. Vacío es No: nadie escribe "No" 300 veces."""
    if valor in (None, ''):
        return False
    if isinstance(valor, bool):
        return valor
    if isinstance(valor, (int, float)):
        return bool(valor)
    txt = str(valor).strip().lower()
    if txt in _TEXTO_VERDADERO:
        return True
    if txt in _TEXTO_FALSO:
        return False
    raise ValueError(f'Valor inválido en "{campo}": "{valor}". Poné Sí o No.')


def build_gastos_template(consorcios: list, proveedores: list, unidades: list):
    """La plantilla .xlsx de carga masiva de gastos.

    Las hojas de referencia no son decorativas: la validación de la columna
    "consorcio" apunta a la lista de la hoja "Consorcios existentes", así que el
    administrador elige de un desplegable en vez de escribir un nombre que
    después no matchea.
    """
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.worksheet.datavalidation import DataValidation
    header_fill = PatternFill("solid", fgColor="7C3AED")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    example_font = Font(italic=True, color="9CA3AF")

    def style_header(ws, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[cell.column_letter].width = max(len(h) + 4, 16)

    def lista_dv(ws, opciones, rango):
        dv = DataValidation(type='list', formula1=f'"{",".join(opciones)}"',
                            allow_blank=True, showErrorMessage=False)
        ws.add_data_validation(dv)
        dv.add(rango)

    wb = openpyxl.Workbook()

    ws_info = wb.active
    ws_info.title = 'Instrucciones'
    ws_info.column_dimensions['A'].width = 105
    info_lines = [
        ('Carga masiva de Gastos', True),
        ('', False),
        ('1. Completá la hoja "Gastos": una fila por gasto. Borrá la fila de ejemplo en gris.', False),
        ('2. En "consorcio" elegí del desplegable o escribí el nombre exacto, tal como figura en la hoja', False),
        ('   "Consorcios existentes". Lo mismo con "proveedor" y "unidad": esta carga no crea consorcios,', False),
        ('   proveedores ni unidades nuevas, sólo los referencia. Si no existen, cargalos antes.', False),
        ('3. Guardá el archivo y subilo en el panel. No cambies el nombre de la hoja ni el de las columnas.', False),
        ('', False),
        ('CAMPOS OBLIGATORIOS (sin ellos la fila no se importa):', True),
        ('   • consorcio — nombre exacto de un consorcio tuyo.', False),
        ('   • fecha — fecha del gasto, formato AAAA-MM-DD (también se acepta DD/MM/AAAA).', False),
        ('   • descripcion — qué es el gasto. Ej: "Factura Edesur junio 2026".', False),
        ('   • monto — número mayor a cero. Podés escribirlo como 15430.50 o 15.430,50.', False),
        ('', False),
        ('CAMPOS OPCIONALES (si los dejás vacíos el gasto se carga igual):', True),
        (f'   • categoria — una de: {", ".join(CATEGORIAS_GASTO)}. Si va vacía o no coincide, queda "otro".', False),
        ('   • proveedor — nombre exacto de un proveedor tuyo (hoja "Proveedores existentes").', False),
        ('   • unidad — número de UF si el gasto es de una sola unidad y NO se prorratea entre todas.', False),
        ('     Dejala vacía para el caso normal: gasto general del consorcio.', False),
        ('   • fecha_vencimiento — cuándo vence el pago.', False),
        ('   • pagado — Sí / No. Vacío se toma como No.', False),
        ('   • fecha_pago — cuándo se pagó. Si la completás, el gasto queda marcado como pagado.', False),
        (f'   • metodo_pago — {", ".join(METODOS_PAGO_GASTO)}.', False),
        ('   • recurrente — Sí / No. Un gasto recurrente se vuelve a generar solo cada período.', False),
        (f'   • frecuencia — {", ".join(FRECUENCIAS_GASTO)}. Sólo aplica si "recurrente" es Sí (vacío = mensual).', False),
        ('   • dia_carga — día del mes (1 a 31) en que se genera el recurrente. Sólo si "recurrente" es Sí.', False),
        ('   • notas — número de factura, medidor, lo que quieras dejar anotado.', False),
        ('', False),
        ('Si un gasto ya existe con el mismo consorcio, fecha, descripción y monto, se omite y no se duplica.', False),
        ('Eso es lo que hace que volver a subir el mismo archivo por error no cargue todo dos veces.', False),
        ('', False),
        ('El Excel que baja el botón "Excel" de la pantalla de Gastos se puede volver a subir acá: sus', False),
        ('columnas se reconocen solas. Lo que no viaja en el archivo son los comprobantes adjuntos.', False),
    ]
    for i, (text, bold) in enumerate(info_lines, 1):
        cell = ws_info.cell(row=i, column=1, value=text)
        if bold:
            cell.font = Font(bold=True, size=13 if i == 1 else 11)

    ws_g = wb.create_sheet('Gastos')
    style_header(ws_g, HEADERS_PLANTILLA_GASTOS)
    nombre_ejemplo = consorcios[0]['nombre'] if consorcios else 'Edificio Ejemplo 123'
    ejemplo = [nombre_ejemplo, '2026-06-05', 'Factura Edesur junio 2026 (borrar fila)', 15430.50,
               'electricidad', '', '', '2026-06-20', 'No', '', 'transferencia', 'No', '', '', 'Factura B-0001-00012345']
    for c, val in enumerate(ejemplo, 1):
        ws_g.cell(row=2, column=c, value=val).font = example_font

    lista_dv(ws_g, CATEGORIAS_GASTO, 'E2:E2000')
    lista_dv(ws_g, ('Si', 'No'), 'I2:I2000')
    lista_dv(ws_g, METODOS_PAGO_GASTO, 'K2:K2000')
    lista_dv(ws_g, ('Si', 'No'), 'L2:L2000')
    lista_dv(ws_g, FRECUENCIAS_GASTO, 'M2:M2000')

    ws_c = wb.create_sheet('Consorcios existentes')
    style_header(ws_c, ['nombre', 'direccion'])
    for r, c in enumerate(consorcios, 2):
        ws_c.cell(row=r, column=1, value=c.get('nombre', ''))
        ws_c.cell(row=r, column=2, value=c.get('direccion', ''))
    if consorcios:
        # Rango y no lista literal: la validación por texto de openpyxl se corta
        # a 255 caracteres y con ocho edificios ya deja nombres partidos al medio.
        dv_con = DataValidation(
            type='list', allow_blank=True, showErrorMessage=False,
            formula1=f"'Consorcios existentes'!$A$2:$A${len(consorcios) + 1}")
        ws_g.add_data_validation(dv_con)
        dv_con.add('A2:A2000')
    else:
        ws_c.cell(row=2, column=1, value='(todavía no tenés consorcios cargados)').font = example_font

    ws_p = wb.create_sheet('Proveedores existentes')
    style_header(ws_p, ['nombre', 'rubro'])
    for r, p in enumerate(proveedores, 2):
        ws_p.cell(row=r, column=1, value=p.get('nombre', ''))
        ws_p.cell(row=r, column=2, value=p.get('rubro', ''))
    if not proveedores:
        ws_p.cell(row=2, column=1, value='(todavía no tenés proveedores cargados)').font = example_font

    ws_u = wb.create_sheet('Unidades existentes')
    style_header(ws_u, ['consorcio', 'unidad', 'piso'])
    for r, u in enumerate(unidades, 2):
        ws_u.cell(row=r, column=1, value=u.get('consorcio', ''))
        ws_u.cell(row=r, column=2, value=u.get('numero', ''))
        ws_u.cell(row=r, column=3, value=u.get('piso', ''))
    if not unidades:
        ws_u.cell(row=2, column=1, value='(todavía no tenés unidades cargadas)').font = example_font

    wb.active = 0
    return wb

def make_pdf(title: str, headers: list, rows: list) -> io.BytesIO:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=landscape(A4), leftMargin=1*cm, rightMargin=1*cm,
                            topMargin=1.5*cm, bottomMargin=1*cm)
    styles = getSampleStyleSheet()
    elements = [Paragraph(title, styles['Title']), Spacer(1, 0.4*cm)]
    data = [headers] + rows
    col_w = (landscape(A4)[0] - 2*cm) / max(len(headers), 1)
    t = Table(data, colWidths=[col_w] * len(headers), repeatRows=1)
    t.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')),
        ('TEXTCOLOR',  (0,0), (-1,0), colors.white),
        ('FONTNAME',   (0,0), (-1,0), 'Helvetica-Bold'),
        ('FONTSIZE',   (0,0), (-1,-1), 8),
        ('GRID',       (0,0), (-1,-1), 0.4, colors.HexColor('#cccccc')),
        ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f0ff')]),
        ('VALIGN',     (0,0), (-1,-1), 'MIDDLE'),
        ('ALIGN',      (0,0), (-1,-1), 'CENTER'),
    ]))
    elements.append(t)
    doc.build(elements)
    return buf


# ── Adjuntos ──────────────────────────────────────────────────────────────────
# Cuatro tipos y nada más, reconocidos por sus primeros bytes y no por lo que
# diga el cliente. Los adjuntos se sirven inline —el vecino hace click y los
# mira en una pestaña— así que un .html o un .svg subido con su propio
# `content_type` era una página ejecutándose en el dominio de Niddo, con la
# sesión de quien la abriera. SVG queda afuera a propósito: es XML con scripts.
FIRMAS_ADJUNTO = (
    (b'%PDF-',              'application/pdf'),
    (b'\xff\xd8\xff',       'image/jpeg'),
    (b'\x89PNG\r\n\x1a\n',  'image/png'),
)
MIMES_ADJUNTO = {'application/pdf', 'image/jpeg', 'image/png', 'image/webp'}
TAMANO_MAXIMO_ADJUNTO = 5 * 1024 * 1024


def _mime_real(datos: bytes) -> Optional[str]:
    """El tipo que declaran los primeros bytes, o None si no es ninguno de la lista."""
    for firma, mime in FIRMAS_ADJUNTO:
        if datos.startswith(firma):
            return mime
    # WEBP es un contenedor RIFF: el tipo está en los bytes 8..12, no al principio.
    if datos[:4] == b'RIFF' and datos[8:12] == b'WEBP':
        return 'image/webp'
    return None


def _nombre_limpio(nombre: str) -> str:
    """Saca separadores y caracteres de control, y deja el resto como estaba.

    No se usa `secure_filename` porque reescribe acentos y espacios, y el nombre
    se le muestra al usuario en la tabla: "Factura Edesur agosto.pdf" tiene que
    seguir leyéndose así. Lo peligroso es la ruta y el salto de línea, y eso sí
    se va.
    """
    nombre = (nombre or '').replace('\\', '/').split('/')[-1]
    nombre = ''.join(c for c in nombre if c.isprintable())
    return nombre[:200] or 'adjunto'


def validar_adjunto(archivo):
    """Devuelve (bytes, mime, nombre) del archivo subido, o levanta ValueError.

    El mime sale del contenido y nunca de `archivo.content_type`: ese lo elige
    quien sube y es exactamente el campo del que dependía el problema.
    """
    datos = archivo.read()
    if not datos:
        raise ValueError('El archivo está vacío')
    if len(datos) > TAMANO_MAXIMO_ADJUNTO:
        raise ValueError('El archivo supera el máximo de 5 MB')
    mime = _mime_real(datos)
    if not mime:
        raise ValueError('Formato no admitido: subí un PDF, JPG, PNG o WEBP')
    return datos, mime, _nombre_limpio(archivo.filename)


def enviar_adjunto(archivo_base64, nombre, mime_guardado, forzar_descarga=False):
    """Sirve un adjunto para verlo en el navegador, sin dejar que se ejecute.

    El tipo se vuelve a deducir del contenido en vez de confiar en el guardado:
    las filas cargadas antes de que existiera la validación pueden tener
    cualquier `mime_type`, y servir un `text/html` viejo sería el mismo agujero
    recién cerrado. Lo que no se reconoce baja como descarga y no se abre.

    `forzar_descarga` es para los archivos del consorcio, que siempre bajaron en
    vez de abrirse: el permiso es el que cambia, no la pantalla.
    """
    datos = base64.b64decode(archivo_base64)
    mime = _mime_real(datos)
    if mime is None and mime_guardado in MIMES_ADJUNTO:
        mime = mime_guardado
    res = send_file(
        io.BytesIO(datos),
        mimetype=mime or 'application/octet-stream',
        download_name=_nombre_limpio(nombre),
        as_attachment=forzar_descarga or mime is None,
    )
    # Sin esto el navegador puede ignorar el Content-Type y adivinar por el
    # contenido, que es la otra mitad del mismo problema.
    res.headers['X-Content-Type-Options'] = 'nosniff'
    return res


# ── Auth decorator ─────────────────────────────────────────────────────────────
def require_auth(allowed_roles=None, permitir_pendiente=False):
    """Exige sesión y, para administradores, que la cuenta esté aprobada.

    `permitir_pendiente=True` es solo para las rutas de la propia solicitud: si
    también exigieran aprobación, el pendiente no tendría dónde completar sus
    datos y quedaría rebotando entre el gate y el formulario.
    """
    def decorator(f):
        @wraps(f)
        def decorated(*args, **kwargs):
            user = session.get('user')
            if not user:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'error': 'No autenticado'}), 401
                return redirect(url_for('login'))
            if allowed_roles and user.get('role') not in allowed_roles:
                if request.is_json or request.path.startswith('/api/'):
                    return jsonify({'error': 'Sin permiso'}), 403
                return redirect(url_for('login'))

            # El gate es solo para administradores: los vecinos entran directo.
            if user.get('role') == 'admin' and not permitir_pendiente:
                if (cargar_admin_actual() or {}).get('estado') != 'aprobado':
                    if request.is_json or request.path.startswith('/api/'):
                        return jsonify({'error': 'Tu cuenta todavía no fue aprobada'}), 403
                    return redirect(url_for('solicitud'))
            return f(*args, **kwargs)
        return decorated
    return decorator


def require_superadmin(f):
    """Restringe una ruta a las cuentas dueñas del sitio.

    Responde 404 y no 403 a quien esté logueado sin el permiso: un 403 le
    confirmaría a un administrador común que la ruta existe. A quien no tiene
    sesión se lo manda al login como a cualquier otra página, porque el aviso
    por mail apunta acá y un 404 dejaría a los superadmins sin poder entrar.
    """
    @wraps(f)
    def decorated(*args, **kwargs):
        user = session.get('user')
        if not user:
            return redirect(url_for('login'))
        if user.get('role') != 'admin':
            abort(404)
        # La fila real, no la impersonada: si se usara cargar_admin_actual(),
        # entrar como un administrador común haría perder el acceso al portal.
        row = cargar_superadmin_real()
        if not row or not row.get('es_superadmin') or row.get('estado') != 'aprobado':
            abort(404)
        return f(*args, **kwargs)
    return decorated


# ── Pertenencia de los datos ───────────────────────────────────────────────────
# `require_auth` contesta una sola pregunta: quién sos. La otra —si este
# consorcio, esta liquidación o este cobro son tuyos— no la contestaba nadie.
# Sin ella, dos administradores aprobados eran indistinguibles frente a la base:
# alcanzaba con cambiar un UUID en la URL para editar el prorrateo de otro,
# bajarse su padrón o mandarle un mail a sus vecinos.
#
# La pertenencia se resuelve acá y no en cada endpoint a propósito. Repartida en
# cuarenta rutas, la garantía vale lo que valga la ruta que alguien olvide, y el
# olvido no se ve: la pantalla funciona igual.
#
# Toda la cadena termina en `consorcios.admin_id`, que es el único lugar donde
# está escrito de quién es un edificio. Las tablas que no tienen `admin_id`
# —cobros, reclamos, avisos de pago, prorrateos— cuelgan de ahí por
# `consorcio_id` o por `liquidacion_id`.

def _no_es_tuyo():
    """Corta la request como si el recurso no existiera.

    404 y no 403 por la misma razón que `require_superadmin`: un 403 confirma
    que el UUID probado existe, y eso es justo lo que le falta a quien está
    tanteando IDs ajenos. Va como JSON porque el 404 de Flask es HTML y le
    rompería el `res.json()` al front.
    """
    abort(make_response(jsonify({'error': 'No encontrado'}), 404))


def consorcio_propio(cid, campos='*'):
    """La fila del consorcio si es del administrador en sesión; si no, 404.

    Devuelve la fila porque varios endpoints ya necesitaban leer el consorcio
    (el nombre para el Excel, el CBU para el resumen): así el chequeo no agrega
    una query, reemplaza la que ya estaba.
    """
    if not cid:
        _no_es_tuyo()
    res = supabase.table('consorcios').select(campos) \
        .eq('id', cid).eq('admin_id', get_admin_id()).execute()
    if not res.data:
        _no_es_tuyo()
    return res.data[0]


def consorcios_propios_ids():
    """IDs de los consorcios del administrador, para los listados sin filtro.

    Una lista vacía es un administrador recién aprobado, sin edificios todavía.
    Quien la reciba tiene que cortar antes de consultar: un `.in_()` con lista
    vacía no filtra nada en algunos backends y devolvería toda la tabla, que es
    exactamente lo contrario de lo que se está pidiendo.
    """
    res = supabase.table('consorcios').select('id').eq('admin_id', get_admin_id()).execute()
    return [c['id'] for c in (res.data or [])]


def liquidacion_propia(lid, campos='*'):
    """La liquidación si es del administrador en sesión; si no, 404.

    `liquidaciones` sí tiene `admin_id`, así que no hace falta pasar por el
    consorcio.
    """
    if not lid:
        _no_es_tuyo()
    res = supabase.table('liquidaciones').select(campos) \
        .eq('id', lid).eq('admin_id', get_admin_id()).execute()
    if not res.data:
        _no_es_tuyo()
    return res.data[0]


def fila_de_consorcio_propio(tabla, rid):
    """Fila de una tabla que cuelga de `consorcio_id`, si el consorcio es propio.

    Son dos queries y no una porque PostgREST no filtra por una columna de la
    tabla padre sin un embed, y un embed acá se lee peor que el salto explícito.
    """
    if not rid:
        _no_es_tuyo()
    res = supabase.table(tabla).select('id, consorcio_id').eq('id', rid).execute()
    fila = res.data[0] if res.data else None
    if not fila:
        _no_es_tuyo()
    consorcio_propio(fila.get('consorcio_id'), 'id')
    return fila


def consorcio_visible(cid):
    """Para las rutas que comparten administradores y vecinos.

    El administrador ve los consorcios que administra; el vecino, el suyo. Sin
    esto, un vecino podía leer los datos de cualquier edificio de la plataforma
    cambiando el UUID.
    """
    if not cid:
        _no_es_tuyo()
    if (session.get('user') or {}).get('role') == 'vecino':
        vecino_id = get_vecino_id()
        if not vecino_id:
            _no_es_tuyo()
        res = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).execute()
        if not res.data or res.data[0].get('consorcio_id') != cid:
            _no_es_tuyo()
        return
    consorcio_propio(cid, 'id')


def amenity_visible(amenity_id):
    """El amenity, si es del consorcio de quien está en sesión.

    Las reservas no guardan el consorcio: cuelgan del amenity, y el amenity del
    consorcio. Ese salto es el único que dice de quién es un quincho, y lo usan
    las tres rutas de reservas.
    """
    if not amenity_id:
        _no_es_tuyo()
    res = supabase.table('amenities').select('id, consorcio_id').eq('id', amenity_id).execute()
    if not res.data:
        _no_es_tuyo()
    consorcio_visible(res.data[0].get('consorcio_id'))
    return res.data[0]


def unidades_del_vecino():
    """Los UUID de unidad que el vecino en sesión tiene derecho a mirar.

    Las expensas cuelgan de la unidad y no del vecino, así que las rutas de
    cobros reciben el `unidad_id` por query string y ninguna comprobaba que
    fuera suyo: cambiar el UUID mostraba la expensa del vecino de al lado.

    La unidad de `vecinos` va primera a propósito. Es la que esas rutas usaban
    cuando no venía `unidad_id`, y si el orden cambiara, un vecino con varias
    unidades abriría la pantalla en otra distinta a la de siempre.
    """
    if 'unidades_vecino' in g:
        return g.unidades_vecino
    g.unidades_vecino = []
    vecino_id = get_vecino_id()
    if not vecino_id:
        return g.unidades_vecino

    res = supabase.table('vecinos').select('unidad_id').eq('id', vecino_id).execute()
    principal = res.data[0].get('unidad_id') if res.data else None
    ids = [principal] if principal else []

    # `vecinos_unidades` es de v7. Si el entorno no la migró, el respaldo de
    # arriba alcanza y la ruta sigue andando en vez de tirar un 500.
    try:
        extra = supabase.table('vecinos_unidades').select('unidad_id') \
            .eq('vecino_id', vecino_id).eq('activo', True).execute().data or []
        for fila in extra:
            uid = fila.get('unidad_id')
            if uid and uid not in ids:
                ids.append(uid)
    except Exception:
        pass

    g.unidades_vecino = ids
    return ids


def unidad_propia(unidad_id):
    """Corta con 404 si la unidad no es del vecino en sesión."""
    if not unidad_id or unidad_id not in unidades_del_vecino():
        _no_es_tuyo()
    return unidad_id


def _vecino_sin_consorcio():
    """True si el vecino en sesión todavía no se asoció a ningún edificio.

    Es el único momento en que mirar consorcios ajenos tiene sentido: en la
    pantalla de alta hay que elegir uno de una lista. Después de asociarse no
    se vuelve a pasar por ahí, y dejar la puerta abierta era lo que permitía
    que cualquier vecino enumerara los edificios y las unidades de toda la
    plataforma.

    Con la solicitud esperando aprobación la puerta se cierra igual: el
    pendiente ya eligió, no tiene nada más que elegir, y si quedara abierta le
    alcanzaría con pedir el alta y no volver a tocarla para seguir enumerando
    la plataforma indefinidamente.
    """
    vecino_id = get_vecino_id()
    if not vecino_id:
        return False
    res = supabase.table('vecinos').select('consorcio_id, estado_asociacion') \
        .eq('id', vecino_id).execute()
    if not res.data:
        return False
    fila = res.data[0]
    if fila.get('estado_asociacion') == 'pendiente':
        return False
    return not fila.get('consorcio_id')


RUTA_SALIR_IMPERSONACION = '/superadmin/impersonar/salir'


@app.before_request
def _impersonacion_es_solo_lectura():
    """Un solo cerrojo para todo el modo impersonación.

    Bloquear por método en un único punto es lo que hace la promesa verificable:
    si dependiera de decorar cada endpoint, alcanzaría con olvidarse de uno para
    que salga un mail a 40 vecinos firmado por un administrador que nunca se
    enteró. Las rutas que escriben son todas POST/PUT/DELETE, así que dejar
    pasar solo lectura las cubre a todas, incluidas las que se agreguen después.
    """
    if not session.get('impersonar'):
        return None
    if request.method in ('GET', 'HEAD', 'OPTIONS'):
        return None
    if request.path == RUTA_SALIR_IMPERSONACION:
        return None
    if request.path.startswith('/api/'):
        return jsonify({'error': 'Estás viendo esta cuenta en modo lectura'}), 403
    return redirect(url_for('superadmin_panel'))


@app.context_processor
def _contexto_impersonacion():
    """Expone la impersonación a todos los templates, para la barra de aviso."""
    return {'impersonando': session.get('impersonar')}


@app.errorhandler(Exception)
def _api_errors_como_json(e):
    """Toda excepción bajo /api/ vuelve como JSON con su motivo.

    Sin esto Flask responde su página HTML de error, el cliente hace r.json(),
    falla el parseo y cae en el `catch(()=>({error:'Error'}))` de admin_dashboard:
    el usuario ve un toast "Error" sin causa y no queda nada con qué diagnosticar.
    """
    from werkzeug.exceptions import HTTPException
    if not request.path.startswith('/api/'):
        # Las páginas siguen con el manejo por defecto: un HTTPException ya es una
        # respuesta válida, cualquier otra cosa se re-lanza para no devolver el
        # objeto excepción como si fuera un body.
        if isinstance(e, HTTPException):
            return e
        raise e
    if isinstance(e, HTTPException):
        return jsonify({'error': e.description}), e.code
    app.logger.exception('Error no atrapado en %s', request.path)
    return jsonify({'error': f'{type(e).__name__}: {e}'}), 500


# ── Páginas públicas ───────────────────────────────────────────────────────────
@app.route('/')
def index():
    return render_template('index.html')


@app.route('/login')
def login():
    user = session.get('user')
    if user:
        return redirect(url_for('dashboard', role=user['role']))
    return render_template('login.html')


# ── Auth0 flow ─────────────────────────────────────────────────────────────────
@app.route('/auth/login')
def auth_login():
    role = request.args.get('role', 'vecino')
    if role not in ('admin', 'vecino'):
        role = 'vecino'
    session['pending_role'] = role
    callback_url = url_for('auth_callback', _external=True)
    return auth0.authorize_redirect(redirect_uri=callback_url)


@app.route('/auth/callback')
def auth_callback():
    token = auth0.authorize_access_token()
    userinfo = token.get('userinfo', {})
    auth0_id = userinfo.get('sub')
    email    = userinfo.get('email', '')
    nombre   = userinfo.get('name', email)
    role     = session.pop('pending_role', 'vecino')
    fila     = upsert_user(role, auth0_id, email, nombre)
    session['user'] = {'sub': auth0_id, 'email': email, 'name': nombre, 'role': role}

    # El administrador que todavía no fue aprobado no ve el dashboard: va a
    # completar sus datos de contacto o a esperar respuesta.
    if role == 'admin' and (fila or {}).get('estado') != 'aprobado':
        return redirect(url_for('solicitud'))
    return redirect(url_for('dashboard', role=role))


@app.route('/auth/logout')
def auth_logout():
    session.clear()
    return redirect(
        f'https://{AUTH0_DOMAIN}/v2/logout'
        f'?returnTo={url_for("index", _external=True)}'
        f'&client_id={os.environ["AUTH0_CLIENT_ID"]}'
    )


# ── Dashboards ────────────────────────────────────────────────────────────────
@app.route('/dashboard/<role>')
@require_auth()
def dashboard(role):
    user = session['user']
    if user['role'] != role:
        return redirect(url_for('dashboard', role=user['role']))
    if role == 'admin':
        return render_template('admin_dashboard.html', user=user,
                               es_superadmin=(cargar_admin_actual() or {}).get('es_superadmin', False))
    elif role == 'vecino':
        return render_template('vecino_dashboard.html', user=user)
    return redirect(url_for('login'))


# ══════════════════════════════════════════════════════════════════════════════
# ALTA DE ADMINISTRADORES — SOLICITUD Y APROBACIÓN
# ══════════════════════════════════════════════════════════════════════════════
LARGO_MAXIMO = {'nombre_contacto': 120, 'telefono': 40, 'email_contacto': 200}


def _campo_texto(datos: dict, campo: str, obligatorio: bool = True) -> str:
    """Valida y normaliza un campo del formulario de solicitud."""
    valor = (datos.get(campo) or '').strip()
    if obligatorio and not valor:
        raise ValueError(f'Falta completar {campo.replace("_", " ")}')
    if len(valor) > LARGO_MAXIMO[campo]:
        raise ValueError(f'El campo {campo.replace("_", " ")} es demasiado largo')
    return valor


def _superadmins() -> list:
    """Los dueños del sitio, para avisarles de cada solicitud nueva."""
    res = supabase.table('administradores').select('email, email_contacto, nombre') \
        .eq('es_superadmin', True).execute()
    return res.data or []


def _auditar(accion: str, target_id: str, motivo: Optional[str] = None) -> None:
    """Registra quién dio o quitó acceso a quién. Nunca interrumpe la acción.

    Si la auditoría fallara (tabla sin migrar, por ejemplo) preferimos que la
    aprobación igual ocurra: dejar al solicitante colgado es peor que perder una
    fila de log, y el error queda en el log de la aplicación.
    """
    try:
        real = cargar_superadmin_real()
        supabase.table('superadmin_auditoria').insert({
            # El actor es siempre la persona real. Con get_admin_id() una acción
            # hecha mientras se impersona quedaría firmada por el impersonado.
            'actor_id':   real['id'] if real else None,
            'target_id':  target_id,
            'accion':     accion,
            'motivo':     motivo,
            'ip':         request.remote_addr,
            'user_agent': (request.headers.get('User-Agent') or '')[:500],
        }).execute()
    except Exception:
        app.logger.exception('No se pudo auditar %s sobre %s', accion, target_id)


def _enviar_mail(destinatarios: list, asunto: str, html: str) -> None:
    """Manda un mail por Resend sin frenar la request si falla.

    Los avisos son notificaciones, no la fuente de verdad: la solicitud ya quedó
    guardada y el panel la muestra igual. Si Resend está caído, que se pierda el
    mail es preferible a devolverle un error al usuario por algo ya hecho.
    """
    destinatarios = [d for d in destinatarios if d]
    if not destinatarios:
        return
    try:
        import resend
        resend.api_key = os.environ.get('RESEND_API_KEY', '')
        if not resend.api_key:
            app.logger.warning('RESEND_API_KEY no configurada: no se envió "%s"', asunto)
            return
        resend.Emails.send({
            'from': os.environ.get('RESEND_FROM_EMAIL', 'Niddo <noreply@niddo.app>'),
            'to': destinatarios,
            'subject': asunto,
            'html': html,
        })
    except Exception:
        app.logger.exception('Falló el envío de "%s"', asunto)


def _mail_shell(titulo: str, cuerpo: str) -> str:
    """Envoltorio con la marca. Estilos en línea: Gmail descarta el <style>."""
    return f"""
    <div style="background:#F6EFE7;padding:32px 16px;font-family:'Nunito Sans',Helvetica,Arial,sans-serif">
      <div style="max-width:560px;margin:0 auto;background:#fff;border:1px solid #E2D6C8;border-radius:18px;overflow:hidden">
        <div style="background:#E8734A;padding:20px 28px">
          <h1 style="margin:0;color:#fff;font-size:19px;font-weight:600">{titulo}</h1>
        </div>
        <div style="padding:28px;color:#2A211C;font-size:15px;line-height:1.65">{cuerpo}</div>
        <div style="padding:16px 28px;border-top:1px solid #E2D6C8;color:#574C42;font-size:12px">
          Niddo — administración de consorcios
        </div>
      </div>
    </div>"""


def _mail_reserva_confirmada(reserva: dict, amenity_id: str) -> None:
    """Le manda al vecino el comprobante de su reserva.

    Hasta acá la única confirmación era un toast en pantalla: el vecino cerraba
    la pestaña y no le quedaba nada con la fecha y la hora que había elegido.

    Va después del insert y nunca antes: el mail confirma algo que ya pasó. Y
    como `_enviar_mail` se traga sus propios errores, que Resend esté caído no
    voltea una reserva que en la base ya está.
    """
    vecino_id = reserva.get('vecino_id')
    if not vecino_id:
        # Reserva cargada por el admin sin decir para quién: no hay a quién avisarle.
        return
    vecino = supabase.table('vecinos').select('nombre, email') \
        .eq('id', vecino_id).single().execute().data or {}
    email = vecino.get('email')
    if not email:
        return

    amenity = supabase.table('amenities') \
        .select('nombre, condiciones_uso, consorcios(nombre)') \
        .eq('id', amenity_id).single().execute().data or {}
    espacio = amenity.get('nombre') or 'el espacio común'
    edificio = (amenity.get('consorcios') or {}).get('nombre') or ''
    condiciones = (amenity.get('condiciones_uso') or '').strip()
    nombre = (vecino.get('nombre') or '').split(' ')[0]
    saludo = f'Hola {nombre},' if nombre else 'Hola,'

    cuerpo = f"""
      <p style="margin:0 0 16px">{saludo}</p>
      <p style="margin:0 0 22px">Tu reserva quedó confirmada.</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:22px">
        <tr><td style="padding:7px 0;color:#574C42;width:130px">Espacio</td><td style="padding:7px 0"><strong>{espacio}</strong></td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Edificio</td><td style="padding:7px 0">{edificio}</td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Día</td><td style="padding:7px 0"><strong>{reserva.get('fecha','')}</strong></td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Horario</td><td style="padding:7px 0"><strong>{reserva.get('hora_inicio','')} a {reserva.get('hora_fin','')}</strong></td></tr>
      </table>
      {f'<p style="margin:0 0 22px;color:#574C42;font-size:13px"><strong>Condiciones de uso:</strong><br>{condiciones}</p>' if condiciones else ''}
      <p style="margin:0;color:#574C42;font-size:13px">
        Si no vas a usarlo, cancelá la reserva desde el portal para que otro vecino pueda tomar el turno.
      </p>"""
    _enviar_mail([email],
                 f'Reserva confirmada — {espacio} el {reserva.get("fecha", "")}',
                 _mail_shell('Tu reserva quedó confirmada', cuerpo))


def _mail_solicitud_vecino(vecino_id: str, consorcio_id: str, unidad_id) -> None:
    """Le avisa al administrador que alguien pidió entrar a su edificio.

    Sin esto la bandeja se llena y nadie se entera: el vecino queda esperando
    a que el administrador entre al panel por casualidad.
    """
    try:
        vecino = supabase.table('vecinos').select('nombre, email, telefono') \
            .eq('id', vecino_id).single().execute().data or {}
        consorcio = supabase.table('consorcios').select('nombre, admin_id') \
            .eq('id', consorcio_id).single().execute().data or {}
        admin = supabase.table('administradores').select('email, email_contacto') \
            .eq('id', consorcio.get('admin_id')).single().execute().data or {}
    except Exception:
        app.logger.exception('No se pudo armar el aviso de solicitud de %s', vecino_id)
        return

    destino = admin.get('email_contacto') or admin.get('email')
    if not destino:
        return

    unidad = '—'
    if unidad_id:
        u = supabase.table('unidades_funcionales').select('numero') \
            .eq('id', unidad_id).execute().data
        unidad = u[0]['numero'] if u else '—'
    else:
        unidad = 'No supo cuál era la suya'

    cuerpo = f"""
      <p style="margin:0 0 18px"><strong>{vecino.get('nombre') or vecino.get('email')}</strong>
      pidió acceso a <strong>{consorcio.get('nombre', '')}</strong>.</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:22px">
        <tr><td style="padding:7px 0;color:#574C42;width:130px">Unidad que dice</td><td style="padding:7px 0"><strong>{unidad}</strong></td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Email</td><td style="padding:7px 0">{vecino.get('email', '')}</td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Teléfono</td><td style="padding:7px 0">{vecino.get('telefono') or '—'}</td></tr>
      </table>
      <a href="{url_for('dashboard', role='admin', _external=True)}"
         style="display:inline-block;background:#E8734A;color:#fff;text-decoration:none;
         padding:12px 24px;border-radius:12px;font-weight:700">Revisar en el panel</a>
      <p style="margin:20px 0 0;color:#574C42;font-size:13px">
        Hasta que lo apruebes no ve nada del edificio: ni expensas, ni gastos, ni comunicados.
        Confirmá que la unidad sea de verdad suya antes de dejarlo entrar.
      </p>"""
    _enviar_mail([destino],
                 f'Alguien pidió acceso a {consorcio.get("nombre", "tu edificio")}',
                 _mail_shell('Nueva solicitud de un vecino', cuerpo))


def _mail_alta_vecino_resuelta(vecino: dict, aprobado: bool, consorcio_id: str,
                               unidad=None, motivo: str = '') -> None:
    """Le dice al vecino cómo salió su solicitud, y si salió mal, por qué."""
    email = vecino.get('email')
    if not email:
        return
    consorcio = supabase.table('consorcios').select('nombre') \
        .eq('id', consorcio_id).execute().data
    nombre_con = consorcio[0]['nombre'] if consorcio else ''
    nombre = (vecino.get('nombre') or '').split(' ')[0]
    saludo = f'Hola {nombre},' if nombre else 'Hola,'

    if aprobado:
        cuerpo = f"""
          <p style="margin:0 0 16px">{saludo}</p>
          <p style="margin:0 0 22px">La administración confirmó tu unidad
          <strong>{unidad}</strong> en <strong>{nombre_con}</strong>. Ya podés entrar
          y ver tus expensas, reclamos, reservas y comunicados.</p>
          <a href="{url_for('dashboard', role='vecino', _external=True)}"
             style="display:inline-block;background:#E8734A;color:#fff;text-decoration:none;
             padding:12px 24px;border-radius:12px;font-weight:700">Entrar a Niddo</a>"""
        asunto = f'Ya podés entrar — {nombre_con}'
        titulo = 'Tu unidad quedó confirmada'
    else:
        explicacion = (f'<p style="margin:0 0 22px"><strong>Motivo:</strong> {motivo}</p>'
                       if motivo else
                       '<p style="margin:0 0 22px">No nos dejaron un motivo. Si creés que '
                       'es un error, hablá con la administración de tu edificio.</p>')
        cuerpo = f"""
          <p style="margin:0 0 16px">{saludo}</p>
          <p style="margin:0 0 16px">La administración de <strong>{nombre_con}</strong> no
          aprobó tu solicitud.</p>
          {explicacion}
          <p style="margin:0;color:#574C42;font-size:13px">
            Tu cuenta sigue activa: podés volver a pedir el alta eligiendo otra unidad.
          </p>"""
        asunto = f'Tu solicitud en {nombre_con} no fue aprobada'
        titulo = 'Sobre tu solicitud'

    _enviar_mail([email], asunto, _mail_shell(titulo, cuerpo))


def _mail_respuesta_reclamo(reclamo: dict, estado_nuevo: str, respuesta: str) -> None:
    """Le avisa al vecino que su reclamo se movió.

    Sin esto el vecino tiene que acordarse de entrar a mirar. Un reclamo es
    justamente lo que uno reporta y después espera: el aviso es la mitad del
    trámite, y era la mitad que faltaba.
    """
    vecino = supabase.table('vecinos').select('nombre, email') \
        .eq('id', reclamo.get('vecino_id')).single().execute().data or {}
    email = vecino.get('email')
    if not email:
        return

    nombre = (vecino.get('nombre') or '').split(' ')[0]
    saludo = f'Hola {nombre},' if nombre else 'Hola,'
    titulo = reclamo.get('titulo') or 'tu reclamo'
    legible = {'activo': 'Activo', 'en_proceso': 'En proceso',
               'resuelto': 'Resuelto', 'cerrado': 'Cerrado'}.get(estado_nuevo, estado_nuevo)

    cuerpo = f"""
      <p style="margin:0 0 16px">{saludo}</p>
      <p style="margin:0 0 22px">La administración actualizó tu reclamo
      <strong>{titulo}</strong>.</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:22px">
        <tr><td style="padding:7px 0;color:#574C42;width:130px">Estado</td><td style="padding:7px 0"><strong>{legible}</strong></td></tr>
      </table>
      {f'<p style="margin:0 0 8px;color:#574C42;font-size:13px"><strong>Respuesta:</strong></p><p style="margin:0 0 22px">{respuesta}</p>' if respuesta else ''}
      <a href="{url_for('dashboard', role='vecino', _external=True)}"
         style="display:inline-block;background:#E8734A;color:#fff;text-decoration:none;
         padding:12px 24px;border-radius:12px;font-weight:700">Ver en el portal</a>"""
    _enviar_mail([email], f'Tu reclamo se actualizó — {titulo}',
                 _mail_shell('Novedades de tu reclamo', cuerpo))


def _mail_aviso_solicitud(fila: dict) -> None:
    """Avisa a los superadmins, con todo lo necesario para levantar el teléfono."""
    panel = url_for('superadmin_panel', _external=True)
    nombre = fila.get('nombre_contacto') or fila.get('nombre') or fila.get('email')
    contacto = fila.get('email_contacto') or fila.get('email') or ''
    telefono = fila.get('telefono') or ''
    cuerpo = f"""
      <p style="margin:0 0 18px"><strong>{nombre}</strong> pidió acceso como administrador.</p>
      <table style="width:100%;border-collapse:collapse;margin-bottom:22px">
        <tr><td style="padding:7px 0;color:#574C42;width:130px">Nombre</td><td style="padding:7px 0"><strong>{nombre}</strong></td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Teléfono</td><td style="padding:7px 0"><a href="tel:{telefono}" style="color:#C4502B;font-weight:700">{telefono}</a></td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Email</td><td style="padding:7px 0"><a href="mailto:{contacto}" style="color:#C4502B;font-weight:700">{contacto}</a></td></tr>
        <tr><td style="padding:7px 0;color:#574C42">Cuenta</td><td style="padding:7px 0">{fila.get('email','')}</td></tr>
      </table>
      <a href="{panel}" style="display:inline-block;background:#E8734A;color:#fff;text-decoration:none;
         padding:12px 24px;border-radius:12px;font-weight:700">Ver en el panel</a>
      <p style="margin:20px 0 0;color:#574C42;font-size:13px">
        Aprobar o rechazar se hace desde el panel, con tu sesión iniciada. Este mail solo avisa.
      </p>"""
    _enviar_mail([s.get('email_contacto') or s.get('email') for s in _superadmins()],
                 f'Nueva solicitud de administrador — {nombre}',
                 _mail_shell('Alguien pidió acceso a Niddo', cuerpo))


def _mail_resultado(fila: dict, aprobado: bool, motivo: str = '') -> None:
    """Le avisa al solicitante cómo salió, y en el rechazo le dice por qué."""
    nombre = (fila.get('nombre_contacto') or fila.get('nombre') or '').split(' ')[0]
    saludo = f'Hola {nombre},' if nombre else 'Hola,'
    if aprobado:
        cuerpo = f"""
          <p style="margin:0 0 16px">{saludo}</p>
          <p style="margin:0 0 22px">Tu cuenta de administrador ya está habilitada. Podés entrar
          cuando quieras y empezar a cargar tus consorcios.</p>
          <a href="{url_for('dashboard', role='admin', _external=True)}"
             style="display:inline-block;background:#E8734A;color:#fff;text-decoration:none;
             padding:12px 24px;border-radius:12px;font-weight:700">Entrar a Niddo</a>
          <p style="margin:22px 0 0;color:#574C42;font-size:13px">
            Si necesitás una mano para arrancar, respondé este mail y te acompañamos.
          </p>"""
        asunto, titulo = 'Tu cuenta de Niddo ya está activa', '¡Bienvenido a Niddo!'
    else:
        cuerpo = f"""
          <p style="margin:0 0 16px">{saludo}</p>
          <p style="margin:0 0 18px">Por ahora no pudimos habilitar tu cuenta de administrador.</p>
          <div style="background:#F6EFE7;border-left:3px solid #E8734A;padding:14px 18px;border-radius:8px;margin-bottom:20px">
            {motivo}
          </div>
          <p style="margin:0 0 22px">Si creés que hubo un error o querés corregir los datos,
          entrá de nuevo y reenviá la solicitud.</p>
          <a href="{url_for('solicitud', _external=True)}"
             style="display:inline-block;background:#E8734A;color:#fff;text-decoration:none;
             padding:12px 24px;border-radius:12px;font-weight:700">Revisar mi solicitud</a>"""
        asunto, titulo = 'Sobre tu solicitud en Niddo', 'Tu solicitud de acceso'
    _enviar_mail([fila.get('email_contacto') or fila.get('email')], asunto,
                 _mail_shell(titulo, cuerpo))


@app.route('/solicitud')
@require_auth(allowed_roles=['admin'], permitir_pendiente=True)
def solicitud():
    """Formulario de contacto, sala de espera y aviso de rechazo, en una pantalla."""
    fila = cargar_admin_actual() or {}
    if fila.get('estado') == 'aprobado':
        return redirect(url_for('dashboard', role='admin'))
    return render_template('solicitud.html', user=session['user'], solicitud=fila)


@app.route('/api/solicitud', methods=['POST'])
@require_auth(allowed_roles=['admin'], permitir_pendiente=True)
def api_solicitud_enviar():
    fila = cargar_admin_actual()
    if not fila:
        return jsonify({'error': 'No encontramos tu cuenta'}), 404
    if fila.get('estado') == 'aprobado':
        return jsonify({'error': 'Tu cuenta ya está aprobada'}), 400

    # Un pendiente que ya mandó sus datos no puede reenviar: sin esto el
    # formulario sería un disparador de mails a los superadmins.
    if fila.get('estado') == 'pendiente' and fila.get('solicitud_completada_at'):
        return jsonify({'error': 'Ya recibimos tu solicitud, está en revisión'}), 409

    datos = request.json or {}
    try:
        nombre_contacto = _campo_texto(datos, 'nombre_contacto')
        telefono        = _campo_texto(datos, 'telefono')
        email_contacto  = _campo_texto(datos, 'email_contacto', obligatorio=False)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400

    supabase.table('administradores').update({
        'nombre_contacto':         nombre_contacto,
        'telefono':                telefono,
        'email_contacto':          email_contacto or fila.get('email'),
        'solicitud_completada_at': now_iso(),
        # Un rechazado que corrige sus datos vuelve a la cola, y el motivo viejo
        # se limpia para que no reaparezca en pantalla junto a la nueva espera.
        'estado':                  'pendiente',
        'motivo_rechazo':          None,
    }).eq('id', fila['id']).execute()

    _mail_aviso_solicitud({**fila, 'nombre_contacto': nombre_contacto,
                           'telefono': telefono,
                           'email_contacto': email_contacto or fila.get('email')})
    return jsonify({'ok': True})


@app.route('/superadmin')
@require_superadmin
def superadmin_panel():
    return render_template('superadmin.html', user=session['user'])


# ══════════════════════════════════════════════════════════════════════════════
# SUPERADMIN — IMPERSONACIÓN (solo lectura)
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/superadmin/impersonar/<aid>', methods=['POST'])
@require_superadmin
def api_superadmin_impersonar(aid):
    fila = _fila_admin('id', aid)
    if not fila:
        return jsonify({'error': 'Administrador no encontrado'}), 404
    if fila.get('es_superadmin'):
        # Entre dueños no hay nada que mirar que no se pueda pedir, y permitirlo
        # abriría la puerta a que uno actúe con la identidad del otro.
        return jsonify({'error': 'No se puede entrar como otro superadmin'}), 403
    if fila.get('estado') != 'aprobado':
        return jsonify({'error': 'Esa cuenta todavía no está aprobada'}), 400

    _auditar('impersonar_inicio', fila['id'])   # antes de tocar la sesión
    session['impersonar'] = {
        'id': fila['id'],
        'nombre': fila.get('nombre_contacto') or fila.get('nombre') or fila.get('email'),
    }
    return jsonify({'ok': True, 'destino': url_for('dashboard', role='admin')})


@app.route(RUTA_SALIR_IMPERSONACION)
@require_superadmin
def superadmin_impersonar_salir():
    imp = session.pop('impersonar', None)
    if imp:
        _auditar('impersonar_fin', imp['id'])
    return redirect(url_for('superadmin_panel'))


# ══════════════════════════════════════════════════════════════════════════════
# SUPERADMIN — MÉTRICAS DE LA PLATAFORMA
# ══════════════════════════════════════════════════════════════════════════════
def _mes(iso: Optional[str]) -> str:
    return (iso or '')[:7]


def _dias_desde(iso: Optional[str]) -> Optional[int]:
    if not iso:
        return None
    try:
        d = datetime.fromisoformat(iso.replace('Z', '+00:00'))
    except ValueError:
        return None
    return (datetime.now(timezone.utc) - d).days


def _ultimos_meses(n: int) -> list:
    """Los últimos n meses como 'YYYY-MM', del más viejo al más nuevo."""
    hoy = datetime.now(timezone.utc)
    meses, y, m = [], hoy.year, hoy.month
    for _ in range(n):
        meses.append(f'{y:04d}-{m:02d}')
        m -= 1
        if m == 0:
            y, m = y - 1, 12
    return list(reversed(meses))


def _cargar_universo() -> dict:
    """Trae de una vez lo que necesita el tablero.

    Se traen filas y se agrega en Python en lugar de pedirle GROUP BY a la base:
    PostgREST no expone agregaciones sin crear vistas o funciones, y a la escala
    actual —decenas de administradores, cientos de unidades— el costo es
    despreciable. Cuando estas tablas lleguen a decenas de miles de filas, esto
    hay que reemplazarlo por vistas materializadas; el resto del portal no se
    entera porque todo pasa por acá.
    """
    return {
        'admins': supabase.table('administradores')
            .select('id, estado, es_superadmin, created_at, ultima_actividad, last_login, '
                    'nombre, email, nombre_contacto, telefono, resuelto_at, '
                    'solicitud_completada_at').execute().data or [],
        'consorcios': supabase.table('consorcios')
            .select('id, admin_id, nombre, created_at, banco_cbu').execute().data or [],
        'ufs': supabase.table('unidades_funcionales')
            .select('id, consorcio_id, vecino_email, vecino_id').execute().data or [],
        'vecinos': supabase.table('vecinos')
            .select('id, created_at, consorcio_id, last_login').execute().data or [],
        'liqs': supabase.table('liquidaciones')
            .select('id, admin_id, periodo, estado, created_at').execute().data or [],
    }


def _indices(u: dict) -> dict:
    """Relaciones derivadas que usan casi todos los cálculos."""
    consorcios_por_admin, ufs_por_consorcio, liqs_por_admin = {}, {}, {}
    for c in u['consorcios']:
        consorcios_por_admin.setdefault(c.get('admin_id'), []).append(c)
    for uf in u['ufs']:
        ufs_por_consorcio.setdefault(uf.get('consorcio_id'), []).append(uf)
    for l in u['liqs']:
        liqs_por_admin.setdefault(l.get('admin_id'), []).append(l)
    return {
        'consorcios_por_admin': consorcios_por_admin,
        'ufs_por_consorcio': ufs_por_consorcio,
        'liqs_por_admin': liqs_por_admin,
    }


def _ufs_de_admin(aid: str, ix: dict) -> int:
    return sum(len(ix['ufs_por_consorcio'].get(c['id'], []))
               for c in ix['consorcios_por_admin'].get(aid, []))


@app.route('/api/superadmin/metricas', methods=['GET'])
@require_superadmin
def api_superadmin_metricas():
    u = _cargar_universo()
    ix = _indices(u)
    meses = _ultimos_meses(12)
    mes_actual, mes_previo = meses[-1], meses[-2]

    # Los superadmins quedan fuera de todas las métricas: son el equipo, no
    # clientes, y con una base chica dos cuentas propias distorsionan todo.
    clientes = [a for a in u['admins'] if not a.get('es_superadmin')]
    aprobados = [a for a in clientes if a.get('estado') == 'aprobado']
    activos = [a for a in aprobados
               if (_dias_desde(a.get('ultima_actividad')) or 9999) <= INACTIVO_DIAS]

    def altas_en(filas, mes):
        return sum(1 for f in filas if _mes(f.get('created_at')) == mes)

    liqs_mes = [l for l in u['liqs'] if l.get('periodo') == mes_actual]

    indicadores = {
        'activos':      {'valor': len(activos), 'total': len(aprobados)},
        'consorcios':   {'valor': len(u['consorcios']),
                         'nuevos': altas_en(u['consorcios'], mes_actual)},
        'unidades':     {'valor': len(u['ufs'])},
        'vecinos':      {'valor': len(u['vecinos']),
                         'nuevos': altas_en(u['vecinos'], mes_actual)},
        'liquidaciones': {'valor': len(liqs_mes),
                          'previo': sum(1 for l in u['liqs'] if l.get('periodo') == mes_previo)},
        'administradores': {'valor': len(aprobados), 'nuevos': altas_en(clientes, mes_actual)},
    }

    tendencia = [{
        'mes': m,
        'administradores': altas_en(clientes, m),
        'consorcios':      altas_en(u['consorcios'], m),
        'liquidaciones':   sum(1 for l in u['liqs'] if l.get('periodo') == m),
    } for m in meses]

    # Embudo: cada escalón es un subconjunto del anterior, así que la caída
    # entre dos escalones señala qué paso del producto cuesta más.
    con_consorcio = [a for a in aprobados if ix['consorcios_por_admin'].get(a['id'])]
    con_unidades  = [a for a in con_consorcio if _ufs_de_admin(a['id'], ix) > 0]
    con_liq       = [a for a in con_unidades if ix['liqs_por_admin'].get(a['id'])]
    embudo = [
        {'paso': 'Se registraron',        'valor': len(clientes)},
        {'paso': 'Aprobados',             'valor': len(aprobados)},
        {'paso': 'Cargaron un consorcio', 'valor': len(con_consorcio)},
        {'paso': 'Cargaron unidades',     'valor': len(con_unidades)},
        {'paso': 'Emitieron liquidación', 'valor': len(con_liq)},
    ]

    return jsonify({
        'indicadores': indicadores,
        'tendencia': tendencia,
        'embudo': embudo,
        'atencion': _lista_atencion(u, ix, aprobados),
    })


def _lista_atencion(u: dict, ix: dict, aprobados: list) -> dict:
    """Lo accionable: gente trabada y cosas rotas, cada una con su motivo."""
    def nombre(a):
        return a.get('nombre_contacto') or a.get('nombre') or a.get('email')

    def ficha(a, motivo):
        return {'id': a['id'], 'nombre': nombre(a), 'email': a.get('email'),
                'telefono': a.get('telefono'), 'motivo': motivo,
                'dias': _dias_desde(a.get('ultima_actividad'))}

    trabados, en_fuga = [], []
    for a in aprobados:
        cons = ix['consorcios_por_admin'].get(a['id'], [])
        liqs = ix['liqs_por_admin'].get(a['id'], [])
        desde_alta = _dias_desde(a.get('resuelto_at') or a.get('created_at'))
        inactivo = _dias_desde(a.get('ultima_actividad'))

        if not cons and (desde_alta or 0) >= 7:
            trabados.append(ficha(a, f'Aprobado hace {desde_alta} días y todavía no cargó ningún consorcio'))
        elif cons and _ufs_de_admin(a['id'], ix) == 0:
            trabados.append(ficha(a, 'Cargó consorcios pero ninguna unidad funcional'))
        elif cons and not liqs and (desde_alta or 0) >= 14:
            trabados.append(ficha(a, 'Tiene el consorcio armado pero nunca emitió una liquidación'))

        if liqs and (inactivo or 0) >= 45:
            en_fuga.append(ficha(a, f'Emitía liquidaciones y hace {inactivo} días que no entra'))

    # Salud: fallas concretas que el administrador todavía no reportó.
    fallidos = supabase.table('resumen_envios').select('id, email_destino, error_detalle, fecha_envio') \
        .eq('estado', 'fallido').order('fecha_envio', desc=True).limit(25).execute().data or []

    sin_cbu = [{'id': c['id'], 'nombre': c.get('nombre')}
               for c in u['consorcios'] if not (c.get('banco_cbu') or '').strip()]
    ufs_sin_email = sum(1 for uf in u['ufs'] if not (uf.get('vecino_email') or '').strip())

    return {
        'trabados': trabados,
        'en_fuga': en_fuga,
        'envios_fallidos': fallidos,
        'consorcios_sin_cbu': sin_cbu,
        'unidades_sin_email': ufs_sin_email,
    }


# ══════════════════════════════════════════════════════════════════════════════
# SUPERADMIN — ADMINISTRADORES, VECINOS Y AUDITORÍA
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/superadmin/administradores', methods=['GET'])
@require_superadmin
def api_superadmin_administradores():
    q = (request.args.get('q') or '').strip().lower()
    u = _cargar_universo()
    ix = _indices(u)

    filas = []
    for a in u['admins']:
        if a.get('es_superadmin'):
            continue
        nombre = a.get('nombre_contacto') or a.get('nombre') or a.get('email') or ''
        if q and q not in nombre.lower() and q not in (a.get('email') or '').lower():
            continue
        cons = ix['consorcios_por_admin'].get(a['id'], [])
        filas.append({
            'id': a['id'], 'nombre': nombre, 'email': a.get('email'),
            'telefono': a.get('telefono'), 'estado': a.get('estado'),
            'consorcios': len(cons), 'unidades': _ufs_de_admin(a['id'], ix),
            'liquidaciones': len(ix['liqs_por_admin'].get(a['id'], [])),
            'created_at': a.get('created_at'),
            'ultima_actividad': a.get('ultima_actividad'),
            'dias_inactivo': _dias_desde(a.get('ultima_actividad')),
        })
    filas.sort(key=lambda f: (f['dias_inactivo'] is None, f['dias_inactivo'] or 0))
    return jsonify({'administradores': filas})


@app.route('/api/superadmin/vecinos', methods=['GET'])
@require_superadmin
def api_superadmin_vecinos():
    """Consulta de vecinos. Sin impersonación: para destrabar a alguien que no
    puede entrar alcanza con ver si su cuenta quedó vinculada a una unidad."""
    q = (request.args.get('q') or '').strip()
    sel = supabase.table('vecinos').select(
        'id, nombre, email, unidad, consorcio_id, created_at, last_login, consorcios(nombre)')
    if q:
        sel = sel.or_(f'email.ilike.%{q}%,nombre.ilike.%{q}%')
    filas = sel.order('created_at', desc=True).limit(200).execute().data or []

    return jsonify({'vecinos': [{
        'id': v['id'], 'nombre': v.get('nombre'), 'email': v.get('email'),
        'unidad': v.get('unidad'),
        'consorcio': (v.get('consorcios') or {}).get('nombre'),
        # Sin consorcio asignado es el síntoma de que el email que cargó su
        # administrador no coincide con el que usó para registrarse.
        'huerfano': not v.get('consorcio_id'),
        'dias_inactivo': _dias_desde(v.get('last_login')),
    } for v in filas]})


@app.route('/api/superadmin/auditoria', methods=['GET'])
@require_superadmin
def api_superadmin_auditoria():
    filas = supabase.table('superadmin_auditoria') \
        .select('id, accion, motivo, ip, created_at, '
                'actor:actor_id(nombre, email), objetivo:target_id(nombre, email)') \
        .order('created_at', desc=True).limit(100).execute().data or []
    return jsonify({'auditoria': filas})


@app.route('/api/superadmin/solicitudes', methods=['GET'])
@require_superadmin
def api_superadmin_solicitudes():
    """Cola de pendientes y las últimas resueltas, para tener contexto."""
    campos = ('id, email, nombre, nombre_contacto, telefono, email_contacto, estado, '
              'created_at, last_login, solicitud_completada_at, resuelto_at, motivo_rechazo')

    pendientes = supabase.table('administradores').select(campos) \
        .eq('estado', 'pendiente').order('solicitud_completada_at', desc=False).execute().data or []
    resueltas = supabase.table('administradores').select(campos) \
        .in_('estado', ['aprobado', 'rechazado']).not_.is_('resuelto_at', 'null') \
        .order('resuelto_at', desc=True).limit(20).execute().data or []

    # Sin datos de contacto es alguien que se registró y abandonó a mitad de
    # camino: se muestra aparte para no ensuciar la cola de los que esperan.
    return jsonify({
        'esperando':   [s for s in pendientes if s.get('solicitud_completada_at')],
        'incompletas': [s for s in pendientes if not s.get('solicitud_completada_at')],
        'resueltas':   resueltas,
    })


@app.route('/api/superadmin/solicitudes/<sid>/aprobar', methods=['POST'])
@require_superadmin
def api_superadmin_aprobar(sid):
    # email_contacto va en el SELECT porque _mail_resultado lo prefiere sobre el
    # email de Auth0: si la persona dio otro para que la contactemos, el aviso
    # tiene que ir ahí y no a la cuenta con la que se registró.
    fila = supabase.table('administradores') \
        .select('id, email, email_contacto, nombre, nombre_contacto, estado') \
        .eq('id', sid).execute().data
    if not fila:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    fila = fila[0]
    if fila['estado'] == 'aprobado':
        return jsonify({'error': 'Esa cuenta ya estaba aprobada'}), 409

    supabase.table('administradores').update({
        'estado':       'aprobado',
        'aprobado_por': get_admin_id(),
        'resuelto_at':  now_iso(),
        'motivo_rechazo': None,
    }).eq('id', sid).execute()

    _auditar('aprobar', sid)
    _mail_resultado(fila, aprobado=True)
    return jsonify({'ok': True})


@app.route('/api/superadmin/solicitudes/<sid>/rechazar', methods=['POST'])
@require_superadmin
def api_superadmin_rechazar(sid):
    motivo = ((request.json or {}).get('motivo') or '').strip()
    if not motivo:
        return jsonify({'error': 'Escribí un motivo: se lo mandamos por mail'}), 400
    if len(motivo) > 1000:
        return jsonify({'error': 'El motivo es demasiado largo'}), 400

    # email_contacto va en el SELECT porque _mail_resultado lo prefiere sobre el
    # email de Auth0: si la persona dio otro para que la contactemos, el aviso
    # tiene que ir ahí y no a la cuenta con la que se registró.
    fila = supabase.table('administradores') \
        .select('id, email, email_contacto, nombre, nombre_contacto, estado') \
        .eq('id', sid).execute().data
    if not fila:
        return jsonify({'error': 'Solicitud no encontrada'}), 404
    fila = fila[0]

    supabase.table('administradores').update({
        'estado':         'rechazado',
        'aprobado_por':   get_admin_id(),
        'resuelto_at':    now_iso(),
        'motivo_rechazo': motivo,
    }).eq('id', sid).execute()

    _auditar('rechazar', sid, motivo)
    _mail_resultado(fila, aprobado=False, motivo=motivo)
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# SUPERADMIN — MAPA CONCEPTUAL
# ══════════════════════════════════════════════════════════════════════════════
# El shell abre las otras tres vistas en <iframe src="...">, no con srcdoc: al
# ser same-origin el navegador manda la cookie de sesión solo, así que las tres
# pasan por @require_superadmin igual que el shell.
@app.route('/superadmin/mapa')
@require_superadmin
def superadmin_mapa():
    """Shell con las 3 pestañas del mapa conceptual."""
    return render_template('superadmin_mapa.html')


@app.route('/superadmin/mapa/categorias')
@require_superadmin
def superadmin_mapa_categorias():
    """Categorías → funcionalidades → componentes, con el checklist persistido."""
    return render_template('superadmin_mapa_categorias.html')


@app.route('/superadmin/mapa/admin-vecino')
@require_superadmin
def superadmin_mapa_admin_vecino():
    """Qué funcionalidades conectan entre Admin, Vecino y Superadmin."""
    return render_template('superadmin_mapa_admin_vecino.html')


@app.route('/superadmin/mapa/arquitectura')
@require_superadmin
def superadmin_mapa_arquitectura():
    """Flujo Landing → Auth0 → Admin/Vecino y stack técnico por categoría."""
    return render_template('superadmin_mapa_arquitectura.html')


@app.route('/superadmin/mapa/lanzamiento')
@require_superadmin
def superadmin_mapa_lanzamiento():
    """Qué falta para el primer administrador real, y qué viene después.

    Las otras tres pestañas responden "¿qué hay construido?". Ésta responde
    otra cosa y por eso mezcla código con lo que no lo es —textos legales, un
    dominio verificado, una cuenta de monitoreo—, que es donde suelen morir
    los lanzamientos.
    """
    return render_template('superadmin_mapa_lanzamiento.html')


@app.route('/api/superadmin/roadmap', methods=['GET'])
@require_superadmin
def api_superadmin_roadmap():
    """Devuelve {"exp::0::0": "planeada", ...}.

    Solo vienen los items tocados a mano: 'sin_definir' es la ausencia de fila,
    así que el mapa arranca de su propio seed y pisa con lo que haya acá.
    """
    filas = supabase.table('roadmap_estado').select('id, estado').execute().data or []
    return jsonify({f['id']: f['estado'] for f in filas})


@app.route('/api/superadmin/roadmap/<path:item_id>', methods=['PUT'])
@require_superadmin
def api_superadmin_roadmap_update(item_id):
    """Guarda el estado de un item: {"estado": "planeada" | "construida" | null}.

    `null` es volver a 'sin_definir', que se representa borrando la fila.
    """
    estado = (request.json or {}).get('estado')
    if estado not in ('planeada', 'construida', None):
        return jsonify({'error': 'Estado inválido: "planeada", "construida" o null'}), 400

    if estado is None:
        supabase.table('roadmap_estado').delete().eq('id', item_id).execute()
    else:
        supabase.table('roadmap_estado').upsert({
            'id': item_id,
            'estado': estado,
            'actualizado_por': (cargar_superadmin_real() or {}).get('id'),
            'actualizado_at': now_iso(),
        }).execute()
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# API — CONSORCIOS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/consorcios', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_list():
    admin_id = get_admin_id()
    res = supabase.table('consorcios').select('*').eq('admin_id', admin_id).order('nombre').execute()
    return jsonify(res.data)


@app.route('/api/consorcios', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_create():
    admin_id = get_admin_id()
    d = request.json
    payload = {
        'nombre': d.get('nombre', '').strip(),
        'direccion': d.get('direccion', ''),
        'cuit': d.get('cuit', ''),
        'pisos': d.get('pisos'),
        'unidades_totales': d.get('unidades_totales'),
        'encargado_nombre': d.get('encargado_nombre', ''),
        'encargado_tel': d.get('encargado_tel', ''),
        'admin_id': admin_id,
    }
    res = supabase.table('consorcios').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/consorcios/<cid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_update(cid):
    admin_id = get_admin_id()
    d = request.json
    payload = {k: v for k, v in {
        'nombre': d.get('nombre'),
        'direccion': d.get('direccion'),
        'cuit': d.get('cuit'),
        'pisos': d.get('pisos'),
        'unidades_totales': d.get('unidades_totales'),
        'encargado_nombre': d.get('encargado_nombre'),
        'encargado_tel': d.get('encargado_tel'),
        # Parámetros de liquidación. Antes la tasa de mora estaba escrita a
        # mano en el HTML de Configuración y el botón no tenía handler.
        'tasa_interes_mora': d.get('tasa_interes_mora'),
        'dias_gracia_mora': d.get('dias_gracia_mora'),
        'recargo_segundo_vto': d.get('recargo_segundo_vto'),
        'clave_suterh': d.get('clave_suterh'),
        'banco_titular': d.get('banco_titular'),
        'banco_nombre': d.get('banco_nombre'),
        'banco_cbu': d.get('banco_cbu'),
        'banco_alias': d.get('banco_alias'),
    }.items() if v is not None}
    res = supabase.table('consorcios').update(payload).eq('id', cid).eq('admin_id', admin_id).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/consorcios/<cid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_consorcios_delete(cid):
    admin_id = get_admin_id()
    supabase.table('consorcios').delete().eq('id', cid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/consorcios/<cid>/unidades', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_ufs_list(cid):
    consorcio_propio(cid, 'id')
    res = supabase.table('unidades_funcionales').select('*').eq('consorcio_id', cid).order('numero').execute()
    ufs = res.data or []

    # Obtener todos los vecinos vinculados a este consorcio
    vecinos_res = supabase.table('vecinos').select('id, nombre, email, rol, unidad, unidad_id').eq('consorcio_id', cid).execute()
    vecinos = vecinos_res.data or []

    # Agrupar vecinos por unidad_id o por número de unidad como fallback
    vecinos_por_uf = {}
    for v in vecinos:
        key = v.get('unidad_id') or v.get('unidad')
        if key:
            if key not in vecinos_por_uf:
                vecinos_por_uf[key] = []
            vecinos_por_uf[key].append(v)

    # Asociar los vecinos correspondientes a cada UF
    for uf in ufs:
        uf_key_id = uf['id']
        uf_key_num = uf['numero']
        associated = vecinos_por_uf.get(uf_key_id, [])
        if not associated:
            associated = vecinos_por_uf.get(uf_key_num, [])
        uf['vecinos_vinculados'] = associated

    return jsonify(ufs)


@app.route('/api/consorcios/<cid>/unidades', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_ufs_create(cid):
    consorcio_propio(cid, 'id')
    d = request.json
    payload = {
        'consorcio_id': cid,
        'numero': d.get('numero', '').strip(),
        'piso': d.get('piso', ''),
        'tipo': d.get('tipo', 'departamento'),
        'superficie_m2': d.get('superficie_m2'),
        'vecino_nombre': d.get('vecino_nombre', ''),
        'vecino_email': d.get('vecino_email', ''),
        # El porcentaje de copropiedad por coeficiente. Era el bloqueo de todo:
        # las columnas existían desde v7 y no había forma de cargarlas, así que
        # quedaban en 0 y el prorrateo caía siempre en reparto lineal.
        'porcentaje_a': d.get('porcentaje_a') or 0,
        'porcentaje_b': d.get('porcentaje_b') or 0,
        'porcentaje_c': d.get('porcentaje_c') or 0,
        'porcentaje_e': d.get('porcentaje_e') or 0,
    }
    res = supabase.table('unidades_funcionales').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/consorcios/<cid>/unidades/<uid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_ufs_update(cid, uid):
    consorcio_propio(cid, 'id')
    d = request.json
    payload = {k: v for k, v in {
        'numero': d.get('numero'),
        'piso': d.get('piso'),
        'tipo': d.get('tipo'),
        'superficie_m2': d.get('superficie_m2'),
        'vecino_nombre': d.get('vecino_nombre'),
        'vecino_email': d.get('vecino_email'),
        'porcentaje_a': d.get('porcentaje_a'),
        'porcentaje_b': d.get('porcentaje_b'),
        'porcentaje_c': d.get('porcentaje_c'),
        'porcentaje_e': d.get('porcentaje_e'),
    }.items() if v is not None}
    res = supabase.table('unidades_funcionales').update(payload).eq('id', uid).eq('consorcio_id', cid).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/consorcios/<cid>/unidades/<uid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_ufs_delete(cid, uid):
    consorcio_propio(cid, 'id')
    supabase.table('unidades_funcionales').delete().eq('id', uid).eq('consorcio_id', cid).execute()
    return jsonify({'ok': True})


@app.route('/api/consorcios/plantilla')
@require_auth(allowed_roles=['admin'])
def descargar_plantilla_carga_masiva():
    admin_id = get_admin_id()
    existentes = supabase.table('consorcios').select('nombre,direccion').eq('admin_id', admin_id).order('nombre').execute().data or []
    wb = build_carga_masiva_template(existentes)
    return excel_response(wb, 'plantilla_carga_masiva.xlsx')


@app.route('/api/consorcios/carga-masiva', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_carga_masiva():
    admin_id = get_admin_id()
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No se envió archivo'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception:
        return jsonify({'error': 'No se pudo leer el archivo. Verificá que sea el .xlsx de la plantilla.'}), 400

    errores = []

    # ── Paso A: hoja "Consorcios" ────────────────────────────────────────────
    existentes_res = supabase.table('consorcios').select('id,nombre').eq('admin_id', admin_id).execute().data or []
    mapa_consorcios = {c['nombre'].strip().lower(): c['id'] for c in existentes_res}
    ids_originales = {c['id'] for c in existentes_res}
    ids_reutilizados = set()
    nuevos_consorcios = []

    if 'Consorcios' in wb.sheetnames:
        ws_c = wb['Consorcios']
        for i, row in enumerate(ws_c.iter_rows(min_row=2, values_only=True), 2):
            if not row or all(v in (None, '') for v in row):
                continue
            nombre = str(row[0]).strip() if row[0] else ''
            if es_fila_ejemplo(nombre):
                continue
            if not nombre:
                errores.append({'hoja': 'Consorcios', 'fila': i, 'mensaje': 'Falta el nombre del consorcio'})
                continue
            key = nombre.lower()
            if key in mapa_consorcios:
                ids_reutilizados.add(mapa_consorcios[key])
                continue
            nuevos_consorcios.append({
                'nombre': nombre,
                'direccion': row[1] or '',
                'cuit': row[2] or '',
                'pisos': row[3] or None,
                'unidades_totales': row[4] or None,
                'encargado_nombre': row[5] or '',
                'encargado_tel': row[6] or '',
                'admin_id': admin_id,
            })
            mapa_consorcios[key] = None  # placeholder hasta insertar, evita duplicar dentro del mismo archivo

    if nuevos_consorcios:
        creados = supabase.table('consorcios').insert(nuevos_consorcios).execute().data or []
        for c in creados:
            mapa_consorcios[c['nombre'].strip().lower()] = c['id']

    # ── Paso B: hoja "Unidades" ──────────────────────────────────────────────
    nuevas_ufs = []
    consorcio_ids_tocados = set()

    if 'Unidades' in wb.sheetnames:
        ws_u = wb['Unidades']
        filas_unidades = [(i, row) for i, row in enumerate(ws_u.iter_rows(min_row=2, values_only=True), 2)
                           if row and not all(v in (None, '') for v in row)]
        for i, row in filas_unidades:
            nombre_con = str(row[0]).strip() if row[0] else ''
            if es_fila_ejemplo(nombre_con):
                continue
            con_id = mapa_consorcios.get(nombre_con.lower())
            if not nombre_con or not con_id:
                errores.append({'hoja': 'Unidades', 'fila': i, 'mensaje': f'Consorcio no encontrado: "{nombre_con}"'})
                continue
            if con_id in ids_originales:
                ids_reutilizados.add(con_id)
            numero = str(row[1]).strip() if row[1] else ''
            if not numero:
                errores.append({'hoja': 'Unidades', 'fila': i, 'mensaje': 'Falta el número de unidad'})
                continue
            tipo = str(row[3]).strip().lower() if row[3] else 'departamento'
            if tipo not in TIPOS_UF_VALIDOS:
                tipo = 'departamento'
            consorcio_ids_tocados.add(con_id)
            nuevas_ufs.append({
                'consorcio_id': con_id,
                'numero': numero,
                'piso': str(row[2]) if row[2] not in (None, '') else '',
                'tipo': tipo,
                'superficie_m2': row[4] or None,
                'vecino_nombre': row[5] or '',
                'vecino_email': row[6] or '',
            })

    # Evitar duplicar UF ya existentes en el mismo consorcio
    numeros_existentes = set()
    if consorcio_ids_tocados:
        existentes_uf = supabase.table('unidades_funcionales').select('consorcio_id,numero') \
            .in_('consorcio_id', list(consorcio_ids_tocados)).execute().data or []
        numeros_existentes = {(u['consorcio_id'], u['numero'].strip().lower()) for u in existentes_uf}

    ufs_a_insertar = []
    unidades_omitidas = 0
    vistas_en_archivo = set()
    for uf in nuevas_ufs:
        key = (uf['consorcio_id'], uf['numero'].strip().lower())
        if key in numeros_existentes or key in vistas_en_archivo:
            unidades_omitidas += 1
            continue
        vistas_en_archivo.add(key)
        ufs_a_insertar.append(uf)

    if ufs_a_insertar:
        supabase.table('unidades_funcionales').insert(ufs_a_insertar).execute()

    return jsonify({
        'consorcios_creados': len(nuevos_consorcios),
        'consorcios_reutilizados': len(ids_reutilizados),
        'unidades_creadas': len(ufs_a_insertar),
        'unidades_omitidas': unidades_omitidas,
        'errores': errores,
    })


@app.route('/api/consorcios/<cid>/export/excel')
@require_auth(allowed_roles=['admin'])
def export_consorcios_excel(cid):
    con = consorcio_propio(cid)
    ufs = supabase.table('unidades_funcionales').select('*').eq('consorcio_id', cid).order('numero').execute().data
    headers = ['UF', 'Piso', 'Tipo', 'Superficie m²', 'Vecino', 'Email']
    rows = [[u['numero'], u.get('piso',''), u.get('tipo',''), u.get('superficie_m2',''),
             u.get('vecino_nombre',''), u.get('vecino_email','')] for u in ufs]
    wb = make_excel(headers, rows, 'Unidades')
    return excel_response(wb, f"consorcio_{con.get('nombre','')}.xlsx")


@app.route('/api/consorcios/<cid>/export/pdf')
@require_auth(allowed_roles=['admin'])
def export_consorcios_pdf(cid):
    con = consorcio_propio(cid)
    ufs = supabase.table('unidades_funcionales').select('*').eq('consorcio_id', cid).order('numero').execute().data
    headers = ['UF', 'Piso', 'Tipo', 'Sup. m²', 'Vecino', 'Email']
    rows = [[u['numero'], u.get('piso',''), u.get('tipo',''), str(u.get('superficie_m2','')),
             u.get('vecino_nombre',''), u.get('vecino_email','')] for u in ufs]
    buf = make_pdf(f"Consorcio: {con.get('nombre','')}", headers, rows)
    return pdf_response(buf, f"consorcio_{con.get('nombre','')}.pdf")


# ══════════════════════════════════════════════════════════════════════════════
# API — PROVEEDORES
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/proveedores', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_list():
    admin_id = get_admin_id()
    res = supabase.table('proveedores').select('*').eq('admin_id', admin_id).order('nombre').execute()
    return jsonify(res.data)


@app.route('/api/proveedores', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_create():
    admin_id = get_admin_id()
    d = request.json
    payload = {
        'nombre': d.get('nombre', '').strip(),
        'cuit': d.get('cuit', ''),
        'rubro': d.get('rubro', ''),
        'email': d.get('email', ''),
        'telefono': d.get('telefono', ''),
        'admin_id': admin_id,
    }
    res = supabase.table('proveedores').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/proveedores/<pid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_update(pid):
    admin_id = get_admin_id()
    d = request.json
    payload = {k: v for k, v in d.items() if k in ('nombre','cuit','rubro','email','telefono') and v is not None}
    res = supabase.table('proveedores').update(payload).eq('id', pid).eq('admin_id', admin_id).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/proveedores/<pid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_proveedores_delete(pid):
    admin_id = get_admin_id()
    supabase.table('proveedores').delete().eq('id', pid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/proveedores/<pid>/gastos')
@require_auth(allowed_roles=['admin'])
def api_proveedores_gastos(pid):
    # El filtro por `admin_id` es el que acota: sin él, el historial de un
    # proveedor compartido —el mismo ascensorista factura a varios edificios—
    # mezclaba los gastos de todos los administradores que lo tuvieran cargado.
    res = supabase.table('gastos').select('*, consorcios(nombre)') \
        .eq('proveedor_id', pid).eq('admin_id', get_admin_id()) \
        .order('fecha_gasto', desc=True).execute()
    return jsonify(res.data)


# ══════════════════════════════════════════════════════════════════════════════
# API — GASTOS
# ══════════════════════════════════════════════════════════════════════════════
def _falta_schema_v12(msg):
    """¿El error es "no existe la columna" de las que agrega v12?

    Mismo criterio que `_falta_schema_v9`: las cuatro columnas de gastos
    recurrentes llegan juntas, así que un fallo por cualquiera se responde
    igual — hay que correr el SQL.
    """
    return (('dia_carga' in msg or 'gasto_origen_id' in msg
             or 'periodo_generado' in msg or 'tarifa_confirmada' in msg)
            and ('does not exist' in msg or '42703' in msg
                 or 'PGRST204' in msg or 'schema cache' in msg))


ERROR_FALTA_V12 = ('Falta correr supabase_schema_v12.sql en Supabase: la base todavía '
                   'no tiene las columnas que generan los gastos recurrentes mes a mes.')


def _es_duplicado(msg):
    """¿El INSERT chocó contra el índice único de (origen, período)?

    Significa que otra request ya generó ese gasto entre nuestro chequeo y
    nuestra inserción. No es un error a reportar: el gasto existe, que es el
    resultado que buscábamos.
    """
    return '23505' in msg or 'duplicate key value' in msg


# Campos que el hijo hereda de su plantilla. `monto` entra acá a propósito:
# copiar la tarifa del período anterior es lo que hace que el gasto sirva de
# algo apenas se genera, y el cartel de confirmación es el que se encarga de
# que nadie la dé por buena sin mirarla.
_CAMPOS_HEREDADOS = ('consorcio_id', 'proveedor_id', 'unidad_id', 'descripcion',
                     'categoria', 'monto', 'metodo_pago', 'frecuencia', 'notas')


def generar_recurrentes(admin_id, hoy=None):
    """Crea los gastos que las plantillas recurrentes deberían haber generado.

    Es idempotente y perezosa: la dispara la request que entra a Gastos o la
    que abre el selector de liquidación, en vez de un cron. En una app de
    administración de consorcios el único escenario donde un cron ganaría es
    "nadie entró en tres meses y quiero los gastos igual", y si nadie entró
    tampoco hay liquidación que emitir. A cambio no hay endpoints públicos ni
    secretos de cron que cuidar.

    Devuelve cuántos gastos creó.
    """
    hoy = hoy or date.today()

    plantillas = supabase.table('gastos') \
        .select('id, ' + ', '.join(_CAMPOS_HEREDADOS) + ', fecha_gasto, dia_carga') \
        .eq('admin_id', admin_id).eq('recurrente', True).execute().data or []
    plantillas = [p for p in plantillas if p.get('dia_carga')]
    if not plantillas:
        return 0

    hijos = supabase.table('gastos').select('gasto_origen_id, periodo_generado') \
        .in_('gasto_origen_id', [p['id'] for p in plantillas]).execute().data or []
    ya_generados = {}
    for h in hijos:
        ya_generados.setdefault(h['gasto_origen_id'], set()).add(h['periodo_generado'])

    creados = 0
    for p in plantillas:
        pendientes = periodos_pendientes(
            fecha_inicio=date.fromisoformat(str(p['fecha_gasto'])),
            dia_carga=p.get('dia_carga'),
            frecuencia=p.get('frecuencia'),
            hoy=hoy,
            ya_generados=ya_generados.get(p['id'], set()),
        )
        for periodo, fecha in pendientes:
            hijo = {c: p.get(c) for c in _CAMPOS_HEREDADOS}
            hijo.update({
                'fecha_gasto': str(fecha),
                'admin_id': admin_id,
                'pagado': False,
                'fecha_pago': None,
                # El hijo NO es recurrente: si heredara el tilde generaría sus
                # propios hijos y el mes siguiente habría una progresión
                # geométrica de gastos duplicados.
                'recurrente': False,
                'gasto_origen_id': p['id'],
                'periodo_generado': periodo,
                'tarifa_confirmada': False,
            })
            try:
                supabase.table('gastos').insert(hijo).execute()
                creados += 1
            except Exception as e:
                if _es_duplicado(str(e)):
                    continue
                raise

    return creados


def _generar_recurrentes_silencioso(admin_id):
    """Genera recurrentes sin dejar que un fallo tumbe la pantalla.

    Poder ver y cargar gastos importa más que generar los del mes: si esto
    falla, la lista se muestra igual y el barrido se reintenta en la próxima
    entrada. La excepción se imprime para que quede en los logs de Vercel.
    """
    try:
        return generar_recurrentes(admin_id)
    except Exception as e:
        app.logger.warning('No se pudieron generar los gastos recurrentes: %s', e)
        return 0


@app.route('/api/gastos', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_gastos_list():
    admin_id = get_admin_id()
    _generar_recurrentes_silencioso(admin_id)
    q = supabase.table('gastos') \
        .select('*, consorcios(nombre), proveedores(nombre), unidades_funcionales(numero, piso)') \
        .eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    if request.args.get('desde'):
        q = q.gte('fecha_gasto', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('fecha_gasto', request.args['hasta'])
    res = q.order('fecha_gasto', desc=True).execute()
    return jsonify(res.data)


def _falta_schema_v9(msg):
    """¿El error de Supabase es "no existe la columna" de las que agrega v9?

    Las tres columnas nuevas (gastos.unidad_id, liquidacion_items.unidad_id y
    liquidacion_prorrateo.gastos_particulares) llegan todas juntas en esa
    migración, así que un fallo por cualquiera de ellas se responde igual: hay
    que correr el SQL. Sin este chequeo el admin ve un 500 genérico y no tiene
    forma de saber que le falta un paso de setup.
    """
    return (('unidad_id' in msg or 'gastos_particulares' in msg)
            and ('does not exist' in msg or '42703' in msg
                 or 'PGRST204' in msg or 'schema cache' in msg))


ERROR_FALTA_V9 = ('Falta correr supabase_schema_v9.sql en Supabase: la base todavía no '
                  'tiene las columnas que separan los gastos generales de los '
                  'específicos de una UF.')


def _dia_carga(d):
    """Día del mes elegido para un gasto recurrente, o None.

    Solo tiene sentido en gastos recurrentes: un gasto común con día de carga
    sería una plantilla que nunca genera nada, y quedaría como ruido esperando
    a que alguien tilde "recurrente" y arranque a generar meses sin querer.
    """
    if d.get('recurrente') not in (True, 'true', 'on', '1'):
        return None
    try:
        dia = int(d.get('dia_carga') or 0)
    except (TypeError, ValueError):
        return None
    return dia if 1 <= dia <= 31 else None


def _unidad_es_del_consorcio(unidad_id, consorcio_id):
    """La UF a la que se imputa un gasto tiene que ser de ese mismo consorcio.

    Si no, el gasto se prorratearía sobre un edificio y se cobraría en otro: la
    liquidación del consorcio dueño del gasto nunca lo repartiría y la unidad
    ajena aparecería con un cargo que nadie puede explicar.
    """
    if not unidad_id or not consorcio_id:
        return True
    res = supabase.table('unidades_funcionales').select('id') \
        .eq('id', unidad_id).eq('consorcio_id', consorcio_id).execute()
    return bool(res.data)


@app.route('/api/gastos', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_gastos_create():
    admin_id = get_admin_id()
    # Soporte multipart/form-data para archivos adjuntos
    d = request.form if request.content_type and 'multipart' in request.content_type else request.json or {}
    payload = {
        'consorcio_id': d.get('consorcio_id') or d.get('consorcio_id', ''),
        'proveedor_id': d.get('proveedor_id') or None,
        # NULL = gasto general del consorcio (el caso normal). Con unidad_id el
        # gasto no se prorratea: se le carga entero a esa UF.
        'unidad_id': d.get('unidad_id') or None,
        'descripcion': (d.get('descripcion') or '').strip(),
        'categoria': d.get('categoria', ''),
        'monto': float(d.get('monto', 0)),
        'fecha_gasto': d.get('fecha_gasto', str(date.today())),
        'fecha_vencimiento': d.get('fecha_vencimiento') or None,
        'pagado': d.get('pagado') in (True, 'true', 'on', '1'),
        'fecha_pago': d.get('fecha_pago') or None,
        'metodo_pago': d.get('metodo_pago', ''),
        'recurrente': d.get('recurrente') in (True, 'true', 'on', '1'),
        'frecuencia': d.get('frecuencia', ''),
        'dia_carga': _dia_carga(d),
        'notas': d.get('notas', ''),
        # Por qué coeficiente se reparte, y el comprobante que la Ley 941 pide
        # rendir junto al gasto. 'A' es como se repartía todo hasta ahora.
        'coeficiente': (d.get('coeficiente') or 'A').upper(),
        'comprobante_tipo': d.get('comprobante_tipo', ''),
        'comprobante_numero': d.get('comprobante_numero', ''),
        'admin_id': admin_id,
    }
    # El gasto queda con el `admin_id` de quien lo carga, así que un consorcio
    # ajeno no lo haría aparecer en la lista de su administrador — pero sí en la
    # pantalla de gastos de sus vecinos, que filtra por consorcio.
    #
    # El campo vacío se responde aparte: es el formulario mandado sin elegir
    # edificio, y un "No encontrado" ahí no le dice a nadie qué le falta.
    if not payload['consorcio_id']:
        return jsonify({'error': 'Elegí a qué consorcio corresponde el gasto'}), 400
    consorcio_propio(payload['consorcio_id'], 'id')

    if not _unidad_es_del_consorcio(payload['unidad_id'], payload['consorcio_id']):
        return jsonify({'error': 'La unidad funcional elegida no pertenece a ese consorcio'}), 400

    try:
        res = supabase.table('gastos').insert(payload).execute()
    except Exception as e:
        if _falta_schema_v9(str(e)):
            return jsonify({'error': f'{ERROR_FALTA_V9} [{e}]'}), 409
        if _falta_schema_v12(str(e)):
            return jsonify({'error': f'{ERROR_FALTA_V12} [{e}]'}), 409
        raise
    gasto = res.data[0] if res.data else {}

    # Guardar comprobante si se adjuntó
    archivo = request.files.get('comprobante')
    if archivo and archivo.filename and gasto.get('id'):
        try:
            file_bytes, mime, nombre = validar_adjunto(archivo)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        b64 = base64.b64encode(file_bytes).decode('utf-8')
        supabase.table('comprobantes_gastos').insert({
            'gasto_id': gasto['id'],
            'archivo_nombre': nombre,
            'archivo_base64': b64,
            'mime_type': mime,
        }).execute()
        supabase.table('gastos').update({'archivo_nombre': nombre}).eq('id', gasto['id']).execute()
        gasto['archivo_nombre'] = nombre

    return jsonify(gasto), 201


@app.route('/api/gastos/<gid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_gastos_update(gid):
    admin_id = get_admin_id()
    d = request.form if request.content_type and 'multipart' in request.content_type else request.json or {}
    allowed = ('consorcio_id','proveedor_id','unidad_id','descripcion','categoria','monto','fecha_gasto',
                'fecha_vencimiento','pagado','fecha_pago','metodo_pago','recurrente','frecuencia',
                'dia_carga','notas','coeficiente','comprobante_tipo','comprobante_numero')
    payload = {}
    for k in allowed:
        if k in d:
            v = d[k]
            if k == 'monto':
                v = float(v)
            elif k in ('pagado', 'recurrente'):
                v = v in (True, 'true', 'on', '1')
            elif k == 'dia_carga':
                v = int(v) if str(v).strip().isdigit() and 1 <= int(v) <= 31 else None
            elif k in ('proveedor_id', 'unidad_id', 'fecha_vencimiento', 'fecha_pago'):
                v = v or None
            payload[k] = v

    # Destildar "recurrente" tiene que apagar la generación, no dejar la
    # plantilla con día cargado esperando a que alguien la vuelva a tildar.
    if payload.get('recurrente') is False:
        payload['dia_carga'] = None

    # Editar el monto a mano es una confirmación de tarifa: el administrador
    # acaba de escribir ese número mirando la factura. Dejarlo igual haría que
    # el cartel le vuelva a preguntar por algo que ya resolvió.
    if 'monto' in payload:
        payload['tarifa_confirmada'] = True

    # Mudar el gasto a otro consorcio es válido, pero solo entre los propios: sin
    # el chequeo, un PUT con un `consorcio_id` ajeno metía el gasto en un edificio
    # de otro administrador.
    if payload.get('consorcio_id'):
        consorcio_propio(payload['consorcio_id'], 'id')

    # El consorcio puede no venir en un PUT parcial; se lee el guardado para poder
    # validar igual que en el alta.
    if payload.get('unidad_id'):
        cid = payload.get('consorcio_id')
        if not cid:
            actual = supabase.table('gastos').select('consorcio_id') \
                .eq('id', gid).eq('admin_id', admin_id).execute().data
            cid = actual[0]['consorcio_id'] if actual else None
        if not _unidad_es_del_consorcio(payload['unidad_id'], cid):
            return jsonify({'error': 'La unidad funcional elegida no pertenece a ese consorcio'}), 400

    try:
        res = supabase.table('gastos').update(payload).eq('id', gid).eq('admin_id', admin_id).execute()
    except Exception as e:
        if _falta_schema_v9(str(e)):
            return jsonify({'error': f'{ERROR_FALTA_V9} [{e}]'}), 409
        if _falta_schema_v12(str(e)):
            return jsonify({'error': f'{ERROR_FALTA_V12} [{e}]'}), 409
        raise
    gasto = res.data[0] if res.data else {}

    # Guardar/reemplazar comprobante si se adjuntó
    archivo = request.files.get('comprobante')
    if archivo and archivo.filename:
        try:
            file_bytes, mime, nombre = validar_adjunto(archivo)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        b64 = base64.b64encode(file_bytes).decode('utf-8')
        # Eliminar comprobante anterior si existe
        supabase.table('comprobantes_gastos').delete().eq('gasto_id', gid).execute()
        supabase.table('comprobantes_gastos').insert({
            'gasto_id': gid,
            'archivo_nombre': nombre,
            'archivo_base64': b64,
            'mime_type': mime,
        }).execute()
        supabase.table('gastos').update({'archivo_nombre': nombre}).eq('id', gid).execute()
        gasto['archivo_nombre'] = nombre

    return jsonify(gasto)


@app.route('/api/gastos/<gid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_gastos_delete(gid):
    admin_id = get_admin_id()
    supabase.table('gastos').delete().eq('id', gid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/gastos/tarifas-pendientes')
@require_auth(allowed_roles=['admin'])
def api_tarifas_pendientes():
    """Gastos generados automáticamente cuya tarifa nadie confirmó todavía.

    Alimenta el cartel que salta al entrar a Gastos. Van agrupados por
    consorcio porque un administrador con seis edificios los revisa por
    edificio, no en una lista plana de treinta filas.
    """
    admin_id = get_admin_id()
    try:
        gastos = supabase.table('gastos') \
            .select('id, descripcion, categoria, monto, fecha_gasto, periodo_generado, '
                    'consorcio_id, consorcios(nombre), proveedores(nombre)') \
            .eq('admin_id', admin_id).eq('tarifa_confirmada', False) \
            .order('fecha_gasto').execute().data or []
    except Exception as e:
        if _falta_schema_v12(str(e)):
            return jsonify({'error': f'{ERROR_FALTA_V12} [{e}]'}), 409
        raise

    por_consorcio = {}
    for gasto in gastos:
        cid = gasto['consorcio_id']
        if cid not in por_consorcio:
            por_consorcio[cid] = {
                'consorcio_id': cid,
                'consorcio': (gasto.get('consorcios') or {}).get('nombre', 'Sin consorcio'),
                'gastos': [],
            }
        por_consorcio[cid]['gastos'].append(gasto)

    return jsonify({'total': len(gastos), 'consorcios': list(por_consorcio.values())})


@app.route('/api/gastos/confirmar-tarifas', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_confirmar_tarifas():
    """Marca tarifas como confirmadas, hayan cambiado de monto o no.

    Que "sigue igual" también confirme es deliberado. Si solo apagara el cartel
    un cambio de monto, un gasto con tarifa estable — un seguro, una cuota fija
    — pediría confirmación todos los meses para siempre, y a la tercera vez el
    administrador aprende a cerrar el cartel sin leerlo. Ahí la alerta deja de
    servir justo para el mes en que la tarifa sí cambió.
    """
    admin_id = get_admin_id()
    items = (request.json or {}).get('gastos') or []
    if not items:
        return jsonify({'error': 'No se recibió ningún gasto para confirmar'}), 400

    actualizados = 0
    for item in items:
        gid = item.get('id')
        if not gid:
            continue
        payload = {'tarifa_confirmada': True}
        if item.get('monto') not in (None, ''):
            try:
                payload['monto'] = float(item['monto'])
            except (TypeError, ValueError):
                return jsonify({'error': f'Monto inválido para el gasto {gid}'}), 400
        try:
            res = supabase.table('gastos').update(payload) \
                .eq('id', gid).eq('admin_id', admin_id).execute()
        except Exception as e:
            if _falta_schema_v12(str(e)):
                return jsonify({'error': f'{ERROR_FALTA_V12} [{e}]'}), 409
            raise
        actualizados += len(res.data or [])

    return jsonify({'ok': True, 'actualizados': actualizados})


@app.route('/api/gastos/<gid>/comprobante')
@require_auth()
def api_gasto_comprobante(gid):
    """Servir el comprobante adjunto de un gasto (PDF o imagen).

    La factura lleva el CUIT del proveedor, el importe y a veces el CBU donde se
    pagó. La ruta la usan las dos pantallas —la del administrador y la del
    vecino— así que el permiso se resuelve por el consorcio del gasto y no por
    el rol: cada uno ve los comprobantes de su edificio.
    """
    gasto = supabase.table('gastos').select('consorcio_id').eq('id', gid).execute().data
    if not gasto:
        _no_es_tuyo()
    consorcio_visible(gasto[0].get('consorcio_id'))

    res = supabase.table('comprobantes_gastos').select('*').eq('gasto_id', gid).single().execute()
    if not res.data:
        return jsonify({'error': 'No hay comprobante adjunto para este gasto'}), 404
    comp = res.data
    return enviar_adjunto(comp['archivo_base64'],
                          comp.get('archivo_nombre', 'comprobante.pdf'),
                          comp.get('mime_type'))


@app.route('/api/gastos/extract', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_gastos_extract():
    """Extrae datos de un comprobante usando Groq Vision (Qwen), gratis sin tarjeta."""
    from groq import Groq

    api_key = os.environ.get('GROQ_API_KEY', '')
    if not api_key:
        return jsonify({'error': 'GROQ_API_KEY no configurada'}), 500

    client = Groq(api_key=api_key)

    file = request.files.get('file')
    if not file or not file.filename:
        return jsonify({'error': 'Se requiere un archivo'}), 400

    file_bytes = file.read()
    mime = file.content_type or 'image/jpeg'

    if file.filename.lower().endswith('.pdf'):
        mime = 'image/png'
        try:
            import fitz  # PyMuPDF
            pdf = fitz.open(stream=file_bytes, filetype='pdf')
            pix = pdf[0].get_pixmap(dpi=200)
            file_bytes = pix.tobytes('png')
        except Exception as e:
            return jsonify({'error': f'No se pudo convertir el PDF: {str(e)}'}), 500
    elif file.filename.lower().endswith('.png'):
        mime = 'image/png'
    elif file.filename.lower().endswith(('.jpg', '.jpeg')):
        mime = 'image/jpeg'

    prompt = """Analizá esta factura/comprobante de un gasto de un consorcio de propiedad horizontal en Argentina.
Extraé los siguientes campos y devolvé SOLO un JSON válido (sin markdown, sin texto adicional):

{
  "descripcion": "descripción breve del gasto (ej: 'Factura Edesur junio 2026')",
  "monto": número decimal sin símbolo de moneda (ej: 15430.50),
  "categoria": una de estas opciones exactas: "electricidad", "gas", "agua", "limpieza", "ascensor", "seguro", "honorarios", "impuesto", "mantenimiento", "sueldos", "otro",
  "fecha_gasto": "YYYY-MM-DD" (fecha de emisión de la factura),
  "fecha_vencimiento": "YYYY-MM-DD" o null (fecha de vencimiento de pago),
  "notas": "datos adicionales relevantes (número de factura, cliente, medidor, etc.)"
}

IMPORTANTE:
- El monto debe ser un número, NO un string. Sin puntos de miles, con punto decimal.
- Las fechas en formato YYYY-MM-DD.
- Si no podés determinar un campo, poné null.
- Respondé SOLO el JSON, sin markdown ni explicaciones."""

    try:
        image_b64 = base64.b64encode(file_bytes).decode('utf-8')
        completion = client.chat.completions.create(
            model='qwen/qwen3.6-27b',
            messages=[{
                'role': 'user',
                'content': [
                    {'type': 'text', 'text': prompt},
                    {'type': 'image_url', 'image_url': {'url': f'data:{mime};base64,{image_b64}'}}
                ]
            }],
            response_format={'type': 'json_object'},
            max_completion_tokens=1024
        )

        # Parse the response - clean markdown if present
        text = completion.choices[0].message.content.strip()
        if text.startswith('```'):
            text = text.split('\n', 1)[1] if '\n' in text else text[3:]
            if text.endswith('```'):
                text = text[:-3].strip()
            elif '```' in text:
                text = text[:text.rfind('```')].strip()

        data = json.loads(text)

        # Sanitize and validate
        result = {
            'descripcion': str(data.get('descripcion', '') or '').strip(),
            'monto': None,
            'categoria': '',
            'fecha_gasto': '',
            'fecha_vencimiento': '',
            'notas': str(data.get('notas', '') or '').strip(),
        }

        # Monto
        try:
            result['monto'] = round(float(data.get('monto', 0) or 0), 2)
        except (ValueError, TypeError):
            result['monto'] = None

        # Categoria
        cat = str(data.get('categoria', '') or '').lower().strip()
        result['categoria'] = cat if cat in CATEGORIAS_GASTO else 'otro'

        # Fechas
        for fk in ('fecha_gasto', 'fecha_vencimiento'):
            val = data.get(fk)
            if val and val != 'null':
                try:
                    datetime.strptime(str(val), '%Y-%m-%d')
                    result[fk] = str(val)
                except ValueError:
                    result[fk] = ''

        return jsonify(result)

    except json.JSONDecodeError:
        return jsonify({'error': 'La IA no devolvió datos válidos. Intentá con otra imagen.'}), 422
    except Exception as e:
        return jsonify({'error': f'Error al procesar: {str(e)}'}), 500


@app.route('/api/gastos/export')
@require_auth(allowed_roles=['admin'])
def api_gastos_export():
    admin_id = get_admin_id()
    q = supabase.table('gastos').select('*, consorcios(nombre), proveedores(nombre)').eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    if request.args.get('desde'):
        q = q.gte('fecha_gasto', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('fecha_gasto', request.args['hasta'])
    data = q.order('fecha_gasto', desc=True).execute().data

    headers = ['Fecha', 'Consorcio', 'Descripción', 'Categoría', 'Proveedor', 'Monto', 'Pagado', 'Método Pago', 'Recurrente']
    rows = [[
        g.get('fecha_gasto',''), (g.get('consorcios') or {}).get('nombre',''),
        g.get('descripcion',''), g.get('categoria',''),
        (g.get('proveedores') or {}).get('nombre',''),
        g.get('monto',0), 'Sí' if g.get('pagado') else 'No',
        g.get('metodo_pago',''), 'Sí' if g.get('recurrente') else 'No'
    ] for g in data]

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'pdf':
        buf = make_pdf('Historial de Gastos', headers, [list(map(str, r)) for r in rows])
        return pdf_response(buf, 'gastos.pdf')
    wb = make_excel(headers, rows, 'Gastos')
    return excel_response(wb, 'gastos.xlsx')


@app.route('/api/gastos/plantilla')
@require_auth(allowed_roles=['admin'])
def descargar_plantilla_gastos():
    admin_id = get_admin_id()
    consorcios = supabase.table('consorcios').select('id,nombre,direccion') \
        .eq('admin_id', admin_id).order('nombre').execute().data or []
    proveedores = supabase.table('proveedores').select('nombre,rubro') \
        .eq('admin_id', admin_id).order('nombre').execute().data or []

    # Las UF van en la plantilla para que el administrador sepa qué escribir en
    # la columna "unidad": sin la lista, el gasto particular se carga a ciegas.
    unidades = []
    if consorcios:
        nombres = {c['id']: c['nombre'] for c in consorcios}
        filas = supabase.table('unidades_funcionales').select('consorcio_id,numero,piso') \
            .in_('consorcio_id', list(nombres)).order('numero').execute().data or []
        unidades = [{'consorcio': nombres.get(u['consorcio_id'], ''),
                     'numero': u.get('numero', ''), 'piso': u.get('piso', '')} for u in filas]
        unidades.sort(key=lambda u: (u['consorcio'], u['numero']))

    wb = build_gastos_template(consorcios, proveedores, unidades)
    return excel_response(wb, 'plantilla_gastos.xlsx')


@app.route('/api/gastos/carga-masiva', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_gastos_carga_masiva():
    """Importa una planilla de gastos, fila por fila y sin crear referencias.

    Una fila con un error no aborta la carga: se reporta y las demás entran. Es
    la diferencia entre que el administrador corrija tres renglones y que
    vuelva a empezar de cero por un typo en la fila 148.
    """
    admin_id = get_admin_id()
    file = request.files.get('file')
    if not file:
        return jsonify({'error': 'No se envió archivo'}), 400
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(file.read()), data_only=True)
    except Exception:
        return jsonify({'error': 'No se pudo leer el archivo. Verificá que sea el .xlsx de la plantilla.'}), 400

    # "Gastos" es la hoja de la plantilla; si no está se usa la primera, que es
    # lo que llega cuando el archivo salió de exportar y tiene una sola hoja.
    ws = wb['Gastos'] if 'Gastos' in wb.sheetnames else wb[wb.sheetnames[0]]
    columnas, filas = leer_filas_por_encabezado(ws, ALIAS_COLUMNAS_GASTO)

    faltantes = [c for c in COLUMNAS_GASTO_OBLIGATORIAS if c not in columnas]
    if faltantes:
        nombres = {'consorcio': 'consorcio', 'fecha_gasto': 'fecha',
                   'descripcion': 'descripcion', 'monto': 'monto'}
        return jsonify({'error': 'Al archivo le faltan columnas obligatorias: '
                        + ', '.join(nombres[c] for c in faltantes)
                        + '. Descargá la plantilla y usá esos encabezados.'}), 400

    # ── Referencias: nada de esto se crea, sólo se busca ──────────────────────
    consorcios = supabase.table('consorcios').select('id,nombre').eq('admin_id', admin_id).execute().data or []
    mapa_consorcios = {c['nombre'].strip().lower(): c['id'] for c in consorcios}
    proveedores = supabase.table('proveedores').select('id,nombre').eq('admin_id', admin_id).execute().data or []
    mapa_proveedores = {p['nombre'].strip().lower(): p['id'] for p in proveedores}

    mapa_unidades = {}
    if consorcios:
        ufs = supabase.table('unidades_funcionales').select('id,consorcio_id,numero') \
            .in_('consorcio_id', [c['id'] for c in consorcios]).execute().data or []
        mapa_unidades = {(u['consorcio_id'], (u.get('numero') or '').strip().lower()): u['id'] for u in ufs}

    errores = []
    pendientes = []   # (clave_dedup, payload)
    consorcios_tocados = set()

    for i, fila in filas:
        nombre_con = _texto_celda(fila.get('consorcio'))
        descripcion = _texto_celda(fila.get('descripcion'))
        if es_fila_ejemplo(nombre_con) or es_fila_ejemplo(descripcion):
            continue

        if not nombre_con:
            errores.append({'fila': i, 'mensaje': 'Falta el consorcio'})
            continue
        con_id = mapa_consorcios.get(nombre_con.lower())
        if not con_id:
            errores.append({'fila': i, 'mensaje': f'Consorcio no encontrado: "{nombre_con}"'})
            continue
        if not descripcion:
            errores.append({'fila': i, 'mensaje': 'Falta la descripción del gasto'})
            continue

        try:
            monto = _monto_excel(fila.get('monto'))
            fecha_gasto = _fecha_excel(fila.get('fecha_gasto'))
            if not fecha_gasto:
                raise ValueError('Falta la fecha del gasto')
            fecha_vto = _fecha_excel(fila.get('fecha_vencimiento'))
            fecha_pago = _fecha_excel(fila.get('fecha_pago'))
            pagado = _bool_excel(fila.get('pagado'), 'pagado')
            recurrente = _bool_excel(fila.get('recurrente'), 'recurrente')
        except ValueError as e:
            errores.append({'fila': i, 'mensaje': str(e)})
            continue

        prov_id = None
        nombre_prov = _texto_celda(fila.get('proveedor'))
        if nombre_prov:
            prov_id = mapa_proveedores.get(nombre_prov.lower())
            if not prov_id:
                errores.append({'fila': i, 'mensaje': f'Proveedor no encontrado: "{nombre_prov}". '
                                                      'Cargalo en Proveedores antes de importar.'})
                continue

        unidad_id = None
        nombre_uf = _texto_celda(fila.get('unidad'))
        if nombre_uf:
            unidad_id = mapa_unidades.get((con_id, nombre_uf.lower()))
            if not unidad_id:
                errores.append({'fila': i, 'mensaje': f'La unidad "{nombre_uf}" no existe en '
                                                      f'"{nombre_con}"'})
                continue

        categoria = _texto_celda(fila.get('categoria')).lower()
        if categoria not in CATEGORIAS_GASTO:
            categoria = 'otro'

        # Una fecha de pago cargada es la respuesta a "¿está pagado?": pedirle
        # además que tilde la columna sería preguntar dos veces lo mismo.
        if fecha_pago:
            pagado = True

        frecuencia, dia_carga = '', None
        if recurrente:
            frecuencia = _texto_celda(fila.get('frecuencia')).lower()
            if frecuencia not in FRECUENCIAS_GASTO:
                frecuencia = 'mensual'
            dia = _texto_celda(fila.get('dia_carga'))
            if dia.isdigit() and 1 <= int(dia) <= 31:
                dia_carga = int(dia)

        # Un coeficiente que no existe se corrige a A en vez de rechazar la fila:
        # es la diferencia entre importar 299 gastos y no importar ninguno por
        # una letra mal tipeada. A es además como se repartía todo antes de v18.
        coeficiente = _texto_celda(fila.get('coeficiente')).upper()
        if coeficiente not in COEFICIENTES:
            coeficiente = 'A'

        consorcios_tocados.add(con_id)
        pendientes.append(((con_id, fecha_gasto, descripcion.lower(), monto), {
            'consorcio_id': con_id,
            'proveedor_id': prov_id,
            'unidad_id': unidad_id,
            'descripcion': descripcion,
            'categoria': categoria,
            'monto': monto,
            'fecha_gasto': fecha_gasto,
            'fecha_vencimiento': fecha_vto,
            'pagado': pagado,
            'fecha_pago': fecha_pago,
            'metodo_pago': _texto_celda(fila.get('metodo_pago')).lower(),
            'recurrente': recurrente,
            'frecuencia': frecuencia,
            'dia_carga': dia_carga,
            'coeficiente': coeficiente,
            'comprobante_tipo': _texto_celda(fila.get('comprobante_tipo')),
            'comprobante_numero': _texto_celda(fila.get('comprobante_numero')),
            'notas': _texto_celda(fila.get('notas')),
            'admin_id': admin_id,
        }))

    # ── Deduplicación ────────────────────────────────────────────────────────
    # Mismo consorcio, misma fecha, misma descripción y mismo monto es el mismo
    # gasto. Sin esto, el archivo subido dos veces —el caso normal cuando la
    # primera vez trajo errores y el administrador corrige y reintenta— deja el
    # edificio con todos los gastos duplicados y la liquidación al doble.
    ya_existen = set()
    if consorcios_tocados:
        # Acotado al rango de fechas del archivo y no al historial entero: una
        # importación es casi siempre un período, y PostgREST corta los
        # resultados en 1000 filas. Un administrador con años de gastos
        # cargados recibiría una lista truncada y la deduplicación dejaría
        # pasar justo los repetidos más viejos.
        fechas = [clave[1] for clave, _ in pendientes]
        previos = supabase.table('gastos').select('consorcio_id,fecha_gasto,descripcion,monto') \
            .eq('admin_id', admin_id).in_('consorcio_id', list(consorcios_tocados)) \
            .gte('fecha_gasto', min(fechas)).lte('fecha_gasto', max(fechas)).execute().data or []
        for g in previos:
            try:
                monto_previo = round(float(g.get('monto') or 0), 2)
            except (TypeError, ValueError):
                continue
            ya_existen.add((g.get('consorcio_id'), g.get('fecha_gasto'),
                            (g.get('descripcion') or '').strip().lower(), monto_previo))

    a_insertar, omitidos, vistos = [], 0, set()
    for clave, payload in pendientes:
        if clave in ya_existen or clave in vistos:
            omitidos += 1
            continue
        vistos.add(clave)
        a_insertar.append(payload)

    if a_insertar:
        try:
            supabase.table('gastos').insert(a_insertar).execute()
        except Exception as e:
            if _falta_schema_v9(str(e)):
                return jsonify({'error': f'{ERROR_FALTA_V9} [{e}]'}), 409
            if _falta_schema_v12(str(e)):
                return jsonify({'error': f'{ERROR_FALTA_V12} [{e}]'}), 409
            raise

    return jsonify({
        'gastos_creados': len(a_insertar),
        'gastos_omitidos': omitidos,
        'consorcios_afectados': len({g['consorcio_id'] for g in a_insertar}),
        'monto_total': round(sum(g['monto'] for g in a_insertar), 2),
        'errores': errores,
    })


# ══════════════════════════════════════════════════════════════════════════════
# API — COBROS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/cobros', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_cobros_list():
    # `cobros` no tiene `admin_id`: el dueño se deduce del consorcio. Sin filtro
    # explícito este listado devolvía las expensas de toda la plataforma.
    q = supabase.table('cobros').select('*, unidades_funcionales(numero, vecino_nombre, vecino_email), consorcios(nombre)')
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', consorcio_propio(request.args['consorcio_id'], 'id')['id'])
    else:
        cids = consorcios_propios_ids()
        if not cids:
            return jsonify([])
        q = q.in_('consorcio_id', cids)
    if request.args.get('periodo'):
        q = q.eq('periodo', request.args['periodo'])
    if request.args.get('estado'):
        q = q.eq('estado', request.args['estado'])
    res = q.order('created_at', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/cobros/generar', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_cobros_generar():
    """Genera un cobro para cada UF del consorcio en el período dado."""
    d = request.json
    consorcio_id = consorcio_propio(d.get('consorcio_id'), 'id')['id']
    periodo = d['periodo']
    monto_base = float(d['monto_base'])
    fecha_vencimiento = d.get('fecha_vencimiento')

    # Traer todas las UFs del consorcio
    ufs = supabase.table('unidades_funcionales').select('id').eq('consorcio_id', consorcio_id).execute().data

    rows = []
    for uf in ufs:
        rows.append({
            'unidad_id': uf['id'],
            'consorcio_id': consorcio_id,
            'periodo': periodo,
            'monto_base': monto_base,
            'interes_mora': 0,
            'total': monto_base,
            'estado': 'pendiente',
            'fecha_vencimiento': fecha_vencimiento,
        })
    if rows:
        supabase.table('cobros').insert(rows).execute()
    return jsonify({'generados': len(rows)})


@app.route('/api/cobros/<rid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_cobros_update(rid):
    fila_de_consorcio_propio('cobros', rid)
    d = request.json
    allowed = ('estado','fecha_pago','interes_mora','total','notas','comprobante_nombre')
    payload = {k: v for k, v in d.items() if k in allowed}
    res = supabase.table('cobros').update(payload).eq('id', rid).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/cobros/mora')
@require_auth(allowed_roles=['admin'])
def api_cobros_mora():
    """Cobros vencidos o en mora para el panel de morosidad."""
    q = supabase.table('cobros').select('*, unidades_funcionales(numero, vecino_nombre, vecino_email), consorcios(nombre)')
    q = q.in_('estado', ['vencido', 'en_mora'])
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', consorcio_propio(request.args['consorcio_id'], 'id')['id'])
    else:
        cids = consorcios_propios_ids()
        if not cids:
            return jsonify([])
        q = q.in_('consorcio_id', cids)
    res = q.order('fecha_vencimiento').execute()
    return jsonify(res.data)


@app.route('/api/cobros/export')
@require_auth(allowed_roles=['admin'])
def api_cobros_export():
    q = supabase.table('cobros').select('*, unidades_funcionales(numero, vecino_nombre), consorcios(nombre)')
    propios = True
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', consorcio_propio(request.args['consorcio_id'], 'id')['id'])
    else:
        cids = consorcios_propios_ids()
        propios = bool(cids)
        if propios:
            q = q.in_('consorcio_id', cids)
    if request.args.get('periodo'):
        q = q.eq('periodo', request.args['periodo'])
    # Sin consorcios propios el archivo sale vacío, pero sale: un administrador
    # recién aprobado que exporta espera una planilla sin filas, no un error.
    data = q.order('created_at', desc=True).execute().data if propios else []

    headers = ['Consorcio', 'UF', 'Vecino', 'Período', 'Monto Base', 'Interés', 'Total', 'Estado', 'Vencimiento', 'Fecha Pago']
    rows = [[
        (c.get('consorcios') or {}).get('nombre',''),
        (c.get('unidades_funcionales') or {}).get('numero',''),
        (c.get('unidades_funcionales') or {}).get('vecino_nombre',''),
        c.get('periodo',''), c.get('monto_base',0), c.get('interes_mora',0),
        c.get('total',0), c.get('estado',''),
        c.get('fecha_vencimiento',''), c.get('fecha_pago',''),
    ] for c in data]

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'pdf':
        buf = make_pdf('Cobros / Expensas', headers, [list(map(str, r)) for r in rows])
        return pdf_response(buf, 'cobros.pdf')
    wb = make_excel(headers, rows, 'Cobros')
    return excel_response(wb, 'cobros.xlsx')


# ══════════════════════════════════════════════════════════════════════════════
# API — BALANCE
# ══════════════════════════════════════════════════════════════════════════════
def _cids_del_balance(consorcio_id):
    """Consorcios que entran en el balance: el pedido, o todos los propios.

    Devuelve siempre una lista. Vacía significa "ningún consorcio que mirar" y
    quien la reciba tiene que saltear la consulta, no correrla sin filtro.
    """
    if consorcio_id:
        return [consorcio_propio(consorcio_id, 'id')['id']]
    return consorcios_propios_ids()



@app.route('/api/balance')
@require_auth(allowed_roles=['admin'])
def api_balance():
    admin_id = get_admin_id()
    consorcio_id = request.args.get('consorcio_id')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    # Ingresos (cobros pagados). Los gastos ya venían filtrados por `admin_id`,
    # pero los cobros no tienen esa columna y quedaban sin acotar: el balance
    # sumaba como ingreso propio cada expensa pagada de la plataforma.
    cids = _cids_del_balance(consorcio_id)
    q_cobros = supabase.table('cobros').select('total, periodo, consorcios(nombre)')
    q_cobros = q_cobros.eq('estado', 'pagado')
    if desde:
        q_cobros = q_cobros.gte('fecha_pago', desde)
    if hasta:
        q_cobros = q_cobros.lte('fecha_pago', hasta)
    cobros = q_cobros.in_('consorcio_id', cids).execute().data if cids else []

    # Egresos (gastos)
    q_gastos = supabase.table('gastos').select('monto, fecha_gasto, descripcion, categoria, consorcios(nombre)').eq('admin_id', admin_id)
    if consorcio_id:
        q_gastos = q_gastos.eq('consorcio_id', consorcio_id)
    if desde:
        q_gastos = q_gastos.gte('fecha_gasto', desde)
    if hasta:
        q_gastos = q_gastos.lte('fecha_gasto', hasta)
    gastos = q_gastos.execute().data

    total_ingresos = sum(c.get('total', 0) or 0 for c in cobros)
    total_egresos  = sum(g.get('monto', 0) or 0 for g in gastos)

    return jsonify({
        'ingresos': total_ingresos,
        'egresos': total_egresos,
        'resultado': total_ingresos - total_egresos,
        'cobros': cobros,
        'gastos': gastos,
    })


@app.route('/api/balance/export')
@require_auth(allowed_roles=['admin'])
def api_balance_export():
    admin_id = get_admin_id()
    consorcio_id = request.args.get('consorcio_id')
    desde = request.args.get('desde')
    hasta = request.args.get('hasta')

    cids = _cids_del_balance(consorcio_id)
    q_cobros = supabase.table('cobros').select('total, periodo, consorcios(nombre)').eq('estado', 'pagado')
    if desde: q_cobros = q_cobros.gte('fecha_pago', desde)
    if hasta: q_cobros = q_cobros.lte('fecha_pago', hasta)
    cobros = q_cobros.in_('consorcio_id', cids).execute().data if cids else []

    q_gastos = supabase.table('gastos').select('monto, fecha_gasto, descripcion, categoria, consorcios(nombre)').eq('admin_id', admin_id)
    if consorcio_id: q_gastos = q_gastos.eq('consorcio_id', consorcio_id)
    if desde: q_gastos = q_gastos.gte('fecha_gasto', desde)
    if hasta: q_gastos = q_gastos.lte('fecha_gasto', hasta)
    gastos = q_gastos.execute().data

    headers = ['Tipo', 'Consorcio', 'Descripción/Período', 'Categoría', 'Monto']
    rows = []
    for c in cobros:
        rows.append(['INGRESO', (c.get('consorcios') or {}).get('nombre',''), c.get('periodo',''), 'Expensas', str(c.get('total',0))])
    for g in gastos:
        rows.append(['EGRESO', (g.get('consorcios') or {}).get('nombre',''), g.get('descripcion',''), g.get('categoria',''), str(g.get('monto',0))])

    fmt = request.args.get('fmt', 'excel')
    if fmt == 'pdf':
        buf = make_pdf('Balance Financiero', headers, rows)
        return pdf_response(buf, 'balance.pdf')
    wb = make_excel(headers, rows, 'Balance')
    return excel_response(wb, 'balance.xlsx')


# ══════════════════════════════════════════════════════════════════════════════
# API — AMENITIES & RESERVAS DE AMENITIES
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/consorcios/<cid>/amenities', methods=['GET'])
@require_auth()
def api_amenities_list(cid):
    consorcio_visible(cid)
    res = supabase.table('amenities').select('*').eq('consorcio_id', cid).order('nombre').execute()
    return jsonify(res.data)


@app.route('/api/consorcios/<cid>/amenities', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_amenities_create(cid):
    consorcio_propio(cid, 'id')
    d = request.json
    payload = {
        'consorcio_id': cid,
        'nombre': d.get('nombre', '').strip(),
        'descripcion': d.get('descripcion', ''),
        'condiciones_uso': d.get('condiciones_uso', ''),
        'capacidad_maxima': d.get('capacidad_maxima') or None,
        'max_reservas_mes': d.get('max_reservas_mes') or None,
    }
    res = supabase.table('amenities').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/consorcios/<cid>/amenities/<aid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_amenities_update(cid, aid):
    consorcio_propio(cid, 'id')
    d = request.json
    payload = {k: v for k, v in {
        'nombre': d.get('nombre'),
        'descripcion': d.get('descripcion'),
        'condiciones_uso': d.get('condiciones_uso'),
        'capacidad_maxima': d.get('capacidad_maxima'),
        'max_reservas_mes': d.get('max_reservas_mes'),
    }.items() if v is not None}
    res = supabase.table('amenities').update(payload).eq('id', aid).eq('consorcio_id', cid).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/consorcios/<cid>/amenities/<aid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_amenities_delete(cid, aid):
    consorcio_propio(cid, 'id')
    supabase.table('amenities').delete().eq('id', aid).eq('consorcio_id', cid).execute()
    return jsonify({'ok': True})


@app.route('/api/reservas_amenities', methods=['GET'])
@require_auth()
def api_reservas_list():
    user = session['user']
    consorcio_id = request.args.get('consorcio_id')
    amenity_id = request.args.get('amenity_id')
    fecha = request.args.get('fecha')

    q = supabase.table('reservas_amenities').select('*, amenities(nombre), vecinos(nombre, unidad)')

    if user['role'] == 'admin':
        if amenity_id:
            amenity_visible(amenity_id)
            q = q.eq('amenity_id', amenity_id)
        else:
            # Sin `amenity_id` la consulta salía sin filtrar: devolvía las
            # reservas de toda la plataforma, con nombre y unidad de cada vecino.
            cids = [consorcio_propio(consorcio_id, 'id')['id']] if consorcio_id \
                else consorcios_propios_ids()
            if not cids:
                return jsonify([])
            res_amenities = supabase.table('amenities').select('id').in_('consorcio_id', cids).execute()
            ids = [a['id'] for a in (res_amenities.data or [])]
            if not ids:
                return jsonify([])
            q = q.in_('amenity_id', ids)
    else:
        vecino_id = get_vecino_id()
        if request.args.get('only_mine') == 'true':
            q = q.eq('vecino_id', vecino_id)
        elif amenity_id:
            amenity_visible(amenity_id)
            q = q.eq('amenity_id', amenity_id)
        else:
            q = q.eq('vecino_id', vecino_id)

    if fecha:
        q = q.eq('fecha', fecha)

    res = q.order('fecha').order('hora_inicio').execute()
    return jsonify(res.data)


@app.route('/api/reservas_amenities', methods=['POST'])
@require_auth()
def api_reservas_create():
    user = session['user']
    d = request.json
    amenity_id = d['amenity_id']
    fecha = d['fecha']
    hora_inicio = d['hora_inicio']
    hora_fin = d['hora_fin']

    if user['role'] == 'admin':
        vecino_id = d.get('vecino_id') or None
    else:
        vecino_id = get_vecino_id()

    # Validar formato
    if not amenity_id or not fecha or not hora_inicio or not hora_fin:
        return jsonify({'error': 'Faltan datos obligatorios'}), 400

    # Reservar el quincho de otro edificio es ocuparlo: la validación de
    # conflictos de horario lo daría por bueno igual, porque solo mira el amenity.
    amenity_visible(amenity_id)

    # Comprobar conflictos de horario
    conflicts_res = supabase.table('reservas_amenities').select('*')\
        .eq('amenity_id', amenity_id)\
        .eq('fecha', fecha)\
        .eq('estado', 'confirmada')\
        .execute()

    def to_minutes(t_str):
        parts = list(map(int, t_str.split(':')[:2]))
        return parts[0] * 60 + parts[1]

    try:
        new_start = to_minutes(hora_inicio)
        new_end = to_minutes(hora_fin)
    except Exception:
        return jsonify({'error': 'Formato de hora inválido'}), 400

    if new_start >= new_end:
        return jsonify({'error': 'La hora de inicio debe ser anterior a la hora de fin'}), 400

    # Reservar para ayer no es una reserva. Se comparaba solamente contra las
    # otras reservas, así que una fecha pasada entraba sin chistar y quedaba
    # ocupando el calendario hacia atrás.
    try:
        dia = date.fromisoformat(str(fecha)[:10])
    except ValueError:
        return jsonify({'error': 'Fecha inválida'}), 400
    if dia < date.today():
        return jsonify({'error': 'No se puede reservar una fecha que ya pasó'}), 400

    for r in conflicts_res.data:
        est_start = to_minutes(r['hora_inicio'])
        est_end = to_minutes(r['hora_fin'])
        if new_start < est_end and new_end > est_start:
            return jsonify({'error': 'El horario seleccionado entra en conflicto con otra reserva'}), 400

    # Las dos reglas del amenity. Las dos son opcionales: si el admin no cargó
    # el número, el espacio no tiene tope y se comporta como hasta ahora.
    amenity = supabase.table('amenities') \
        .select('nombre, capacidad_maxima, max_reservas_mes') \
        .eq('id', amenity_id).single().execute().data or {}

    personas = d.get('cantidad_personas')
    if personas is not None:
        try:
            personas = int(personas)
        except (TypeError, ValueError):
            return jsonify({'error': 'La cantidad de personas tiene que ser un número'}), 400
        if personas < 1:
            return jsonify({'error': 'La cantidad de personas tiene que ser al menos 1'}), 400

    cupo = amenity.get('capacidad_maxima')
    if cupo and personas and personas > int(cupo):
        return jsonify({'error': f'{amenity.get("nombre") or "El espacio"} admite hasta '
                                 f'{cupo} personas y estás anotando {personas}'}), 400

    # El tope mensual se cuenta por vecino y por espacio, sobre el mes de la
    # fecha que se está reservando y no sobre el mes en curso: reservar en
    # octubre no debería gastar el cupo de septiembre.
    tope = amenity.get('max_reservas_mes')
    if tope and vecino_id:
        desde = dia.replace(day=1).isoformat()
        hasta = (dia.replace(day=28) + timedelta(days=4)).replace(day=1).isoformat()
        delmes = supabase.table('reservas_amenities').select('id') \
            .eq('amenity_id', amenity_id).eq('vecino_id', vecino_id) \
            .eq('estado', 'confirmada') \
            .gte('fecha', desde).lt('fecha', hasta).execute().data or []
        if len(delmes) >= int(tope):
            return jsonify({'error': f'Ya tenés {len(delmes)} reserva(s) de '
                                     f'{amenity.get("nombre") or "este espacio"} este mes '
                                     f'y el máximo es {tope}'}), 400

    payload = {
        'amenity_id': amenity_id,
        'vecino_id': vecino_id,
        'fecha': fecha,
        'hora_inicio': hora_inicio,
        'hora_fin': hora_fin,
        'cantidad_personas': personas,
        'estado': 'confirmada'
    }
    res = supabase.table('reservas_amenities').insert(payload).execute()
    reserva = res.data[0] if res.data else {}

    # La reserva ya está guardada. Avisar es lo último y lo menos importante:
    # si armar el mail falla —un amenity borrado entre medio, Resend caído—, el
    # vecino tiene que ver su reserva hecha igual, no un 500 sobre algo que sí
    # pasó. `_enviar_mail` ya se traga lo suyo; esto cubre las consultas de
    # alrededor, que también pueden romper.
    try:
        _mail_reserva_confirmada(reserva, amenity_id)
    except Exception:
        app.logger.exception('Falló el aviso de la reserva %s', reserva.get('id'))

    return jsonify(reserva), 201


@app.route('/api/reservas_amenities/<rid>', methods=['DELETE'])
@require_auth()
def api_reservas_delete(rid):
    user = session['user']
    if user['role'] == 'admin':
        # Sin el salto por el amenity, cualquier administrador cancelaba la
        # reserva de cualquier edificio de la plataforma.
        reserva = supabase.table('reservas_amenities').select('amenity_id').eq('id', rid).execute().data
        if not reserva:
            _no_es_tuyo()
        amenity_visible(reserva[0].get('amenity_id'))
        supabase.table('reservas_amenities').delete().eq('id', rid).execute()
    else:
        vecino_id = get_vecino_id()
        booking = supabase.table('reservas_amenities').select('vecino_id').eq('id', rid).single().execute()
        if booking.data and booking.data['vecino_id'] == vecino_id:
            supabase.table('reservas_amenities').delete().eq('id', rid).execute()
        else:
            return jsonify({'error': 'Sin permiso para cancelar esta reserva'}), 403

    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# API — ASOCIACIÓN DE VECINOS
# ══════════════════════════════════════════════════════════════════════════════
@app.route('/api/public/consorcios', methods=['GET'])
@require_auth()
def api_public_consorcios():
    """La lista para el desplegable del alta de vecino.

    Devolvía todos los consorcios de la plataforma a cualquiera con sesión. Un
    vecino ya asociado no tiene nada que elegir acá, y un administrador sólo
    puede querer los suyos.
    """
    if (session.get('user') or {}).get('role') == 'admin':
        res = supabase.table('consorcios').select('id, nombre') \
            .eq('admin_id', get_admin_id()).order('nombre').execute()
        return jsonify(res.data)
    if not _vecino_sin_consorcio():
        _no_es_tuyo()
    res = supabase.table('consorcios').select('id, nombre').order('nombre').execute()
    return jsonify(res.data)


@app.route('/api/public/consorcios/<cid>/unidades-libres', methods=['GET'])
@require_auth()
def api_public_unidades_libres(cid):
    """Las unidades de un consorcio, para elegir la propia al darse de alta.

    La usan dos pantallas: el alta del vecino, que necesita ver un edificio que
    todavía no es suyo, y el panel del administrador sobre uno de los suyos. El
    permiso se resuelve distinto en cada caso; lo que no puede seguir pasando
    es que un vecino ya asociado se baje el padrón de cualquier edificio.
    """
    if (session.get('user') or {}).get('role') == 'admin':
        consorcio_propio(cid, 'id')
    elif not _vecino_sin_consorcio():
        _no_es_tuyo()
    res = supabase.table('unidades_funcionales')\
        .select('id, numero, piso, tipo')\
        .eq('consorcio_id', cid)\
        .order('numero')\
        .execute()
    return jsonify(res.data)


@app.route('/api/vecinos/asociar', methods=['POST'])
@require_auth(allowed_roles=['vecino'])
def api_vecinos_asociar():
    """El vecino pide entrar a una unidad. No entra: queda pendiente.

    Antes esta misma request lo metía adentro. Con un login alcanzaba para
    quedarse con cualquier unidad libre de cualquier edificio de la lista y
    ver sus expensas al instante.

    Lo importante de acá abajo es lo que NO se escribe: `consorcio_id` y
    `unidad_id` siguen en NULL hasta que el administrador apruebe. Son las dos
    columnas de las que cuelga todo el acceso del vecino, así que dejarlas
    vacías es el gate — no hace falta que cada endpoint se acuerde de filtrar
    por estado.
    """
    d = request.json or {}
    consorcio_id = d.get('consorcio_id')
    unidad_id = d.get('unidad_id')
    rol = d.get('rol', 'propietario')

    if not consorcio_id:
        return jsonify({'error': 'El Consorcio es un campo requerido'}), 400

    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({'error': 'No se pudo identificar tu perfil de vecino'}), 404

    actual = supabase.table('vecinos').select('estado_asociacion, consorcio_id') \
        .eq('id', vecino_id).single().execute().data or {}
    if actual.get('consorcio_id'):
        return jsonify({'error': 'Ya estás asociado a una unidad'}), 409
    if actual.get('estado_asociacion') == 'pendiente':
        return jsonify({'error': 'Ya tenés una solicitud esperando aprobación'}), 409

    if unidad_id:
        uf = supabase.table('unidades_funcionales').select('id, numero, vecino_id') \
            .eq('id', unidad_id).eq('consorcio_id', consorcio_id).execute().data
        if not uf:
            return jsonify({'error': 'La unidad seleccionada no existe'}), 400

    supabase.table('vecinos').update({
        'estado_asociacion': 'pendiente',
        'consorcio_solicitado_id': consorcio_id,
        'unidad_solicitada_id': unidad_id or None,
        'solicitud_at': now_iso(),
        'motivo_rechazo': None,
        'rol': rol,
        # `unidad` es el número de la unidad y se escribe recién al aprobar.
        # Antes acá iba el string 'Pendiente', que era el modo viejo de marcar
        # la espera; ahora eso lo dice `estado_asociacion` y nadie lo lee más.
    }).eq('id', vecino_id).execute()

    _mail_solicitud_vecino(vecino_id, consorcio_id, unidad_id)
    return jsonify({'ok': True, 'estado': 'pendiente'})


@app.route('/api/vecinos/mi-solicitud')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_mi_solicitud():
    """En qué quedó la solicitud, para que la espera no sea una pared muda."""
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({})
    v = supabase.table('vecinos') \
        .select('estado_asociacion, consorcio_solicitado_id, unidad_solicitada_id, '
                'solicitud_at, motivo_rechazo, consorcio_id') \
        .eq('id', vecino_id).single().execute().data or {}

    consorcio = unidad = None
    if v.get('consorcio_solicitado_id'):
        c = supabase.table('consorcios').select('nombre') \
            .eq('id', v['consorcio_solicitado_id']).execute().data
        consorcio = c[0]['nombre'] if c else None
    if v.get('unidad_solicitada_id'):
        u = supabase.table('unidades_funcionales').select('numero') \
            .eq('id', v['unidad_solicitada_id']).execute().data
        unidad = u[0]['numero'] if u else None

    return jsonify({
        'estado': v.get('estado_asociacion'),
        'consorcio': consorcio,
        'unidad': unidad,
        'solicitud_at': v.get('solicitud_at'),
        'motivo_rechazo': v.get('motivo_rechazo'),
    })


@app.route('/api/consorcios/<cid>/vecinos/pendientes', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_consorcio_vecinos_pendientes(cid):
    """Los que están esperando que este administrador los deje entrar."""
    consorcio_propio(cid, 'id')
    pendientes = supabase.table('vecinos') \
        .select('id, nombre, email, telefono, rol, unidad_solicitada_id, solicitud_at') \
        .eq('consorcio_solicitado_id', cid) \
        .eq('estado_asociacion', 'pendiente') \
        .execute().data or []

    # El número de la unidad que pidieron, en una query y no una por vecino.
    ids = [v['unidad_solicitada_id'] for v in pendientes if v.get('unidad_solicitada_id')]
    numeros = {}
    if ids:
        ufs = supabase.table('unidades_funcionales').select('id, numero') \
            .in_('id', ids).execute().data or []
        numeros = {u['id']: u['numero'] for u in ufs}
    for v in pendientes:
        v['unidad_solicitada'] = numeros.get(v.get('unidad_solicitada_id'))
    return jsonify(pendientes)


def _aprobar_vecino(cid, vid, unidad_id=None):
    """Mete al vecino adentro: es el único lugar que escribe consorcio_id.

    `unidad_id` lo manda el administrador cuando corrige la unidad pedida, o
    cuando el vecino dijo "no encuentro la mía" y no pidió ninguna.
    """
    vecino = supabase.table('vecinos') \
        .select('id, nombre, email, estado_asociacion, consorcio_solicitado_id, unidad_solicitada_id') \
        .eq('id', vid).execute().data
    vecino = vecino[0] if vecino else None
    if not vecino or vecino.get('consorcio_solicitado_id') != cid:
        _no_es_tuyo()
    if vecino.get('estado_asociacion') != 'pendiente':
        return None, (jsonify({'error': 'Esta solicitud ya fue resuelta'}), 409)

    unidad_id = unidad_id or vecino.get('unidad_solicitada_id')
    if not unidad_id:
        return None, (jsonify({'error': 'Elegí qué unidad asignarle: este vecino no supo cuál era la suya'}), 400)

    uf = supabase.table('unidades_funcionales').select('id, numero, vecino_id') \
        .eq('id', unidad_id).eq('consorcio_id', cid).execute().data
    if not uf:
        return None, (jsonify({'error': 'La unidad seleccionada no existe en este consorcio'}), 400)
    uf = uf[0]

    supabase.table('vecinos').update({
        'estado_asociacion': 'aprobado',
        'consorcio_id': cid,
        'unidad_id': unidad_id,
        'unidad': uf['numero'],
        'motivo_rechazo': None,
    }).eq('id', vid).execute()

    # Compatibilidad con el vecino "principal" de la unidad, que es anterior a
    # vecinos_unidades y todavía lo leen algunas pantallas.
    if not uf.get('vecino_id'):
        supabase.table('unidades_funcionales').update({'vecino_id': vid}) \
            .eq('id', unidad_id).execute()

    return {'vecino': vecino, 'uf': uf}, None


@app.route('/api/consorcios/<cid>/vecinos/<vid>/aprobar', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_consorcio_vecino_aprobar(cid, vid):
    consorcio_propio(cid, 'id')
    datos, error = _aprobar_vecino(cid, vid, (request.json or {}).get('unidad_id'))
    if error:
        return error
    try:
        _mail_alta_vecino_resuelta(datos['vecino'], True, cid, datos['uf']['numero'])
    except Exception:
        app.logger.exception('Falló el aviso de aprobación del vecino %s', vid)
    return jsonify({'ok': True, 'unidad': datos['uf']['numero']})


@app.route('/api/consorcios/<cid>/vecinos/<vid>/rechazar', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_consorcio_vecino_rechazar(cid, vid):
    """Rechaza la solicitud y deja al vecino en condiciones de volver a pedir.

    No se le borra la cuenta ni se lo bloquea: puede haberse equivocado de
    unidad. Se limpian las columnas de solicitud para que el alta le vuelva a
    aparecer, con el motivo a la vista.
    """
    consorcio_propio(cid, 'id')
    motivo = ((request.json or {}).get('motivo') or '').strip()

    vecino = supabase.table('vecinos') \
        .select('id, nombre, email, estado_asociacion, consorcio_solicitado_id') \
        .eq('id', vid).execute().data
    vecino = vecino[0] if vecino else None
    if not vecino or vecino.get('consorcio_solicitado_id') != cid:
        _no_es_tuyo()
    if vecino.get('estado_asociacion') != 'pendiente':
        return jsonify({'error': 'Esta solicitud ya fue resuelta'}), 409

    supabase.table('vecinos').update({
        'estado_asociacion': 'rechazado',
        'consorcio_solicitado_id': None,
        'unidad_solicitada_id': None,
        'motivo_rechazo': motivo or None,
        'unidad': None,
    }).eq('id', vid).execute()

    try:
        _mail_alta_vecino_resuelta(vecino, False, cid, None, motivo)
    except Exception:
        app.logger.exception('Falló el aviso de rechazo del vecino %s', vid)
    return jsonify({'ok': True})


@app.route('/api/consorcios/<cid>/vecinos/<vid>/asignar-unidad', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_consorcio_vecino_asignar(cid, vid):
    """Alias de /aprobar con unidad explícita.

    Se mantiene porque es lo que llama el panel desde antes de que existiera la
    aprobación; ahora aprueba, que es lo que el administrador creía estar
    haciendo cuando asignaba la unidad.
    """
    consorcio_propio(cid, 'id')
    unidad_id = (request.json or {}).get('unidad_id')
    if not unidad_id:
        return jsonify({'error': 'La unidad es requerida'}), 400
    datos, error = _aprobar_vecino(cid, vid, unidad_id)
    if error:
        return error
    try:
        _mail_alta_vecino_resuelta(datos['vecino'], True, cid, datos['uf']['numero'])
    except Exception:
        app.logger.exception('Falló el aviso de aprobación del vecino %s', vid)
    return jsonify({'ok': True})


# ── API: gastos para vecinos ───────────────────────────────────────────────────
@app.route('/api/vecinos/gastos')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_gastos():
    """Lista gastos del consorcio del vecino logueado (datos seguros, sin datos de admin)."""
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    vecino = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).single().execute()
    if not vecino.data or not vecino.data.get('consorcio_id'):
        return jsonify([])
    cid = vecino.data['consorcio_id']
    res = supabase.table('gastos')\
        .select('id, descripcion, categoria, monto, fecha_gasto, pagado, archivo_nombre')\
        .eq('consorcio_id', cid)\
        .order('fecha_gasto', desc=True)\
        .execute()
    return jsonify(res.data)


# ── API: datos del dashboard ───────────────────────────────────────────────────
@app.route('/api/dashboard/kpis')
@require_auth(allowed_roles=['admin'])
def api_dashboard_kpis():
    admin_id = get_admin_id()
    consorcios_count = len(supabase.table('consorcios').select('id').eq('admin_id', admin_id).execute().data)
    gastos_pendientes = len(supabase.table('gastos').select('id').eq('admin_id', admin_id).eq('pagado', False).execute().data)
    mora_count = len(supabase.table('cobros').select('id').in_('estado', ['vencido','en_mora']).execute().data)
    return jsonify({'consorcios': consorcios_count, 'gastos_pendientes': gastos_pendientes, 'en_mora': mora_count})


# ── API: perfil ────────────────────────────────────────────────────────────────
@app.route('/api/me')
@require_auth()
def api_me():
    user = session['user']
    table = 'administradores' if user['role'] == 'admin' else 'vecinos'
    result = supabase.table(table).select('*').eq('auth0_id', user['sub']).execute()
    if not result.data:
        # La sesión es válida pero no hay fila de perfil (usuario borrado, DB
        # reseteada, etc.). Cerramos la sesión en vez de devolver un 500: si
        # no, el frontend reintenta, /login ve la sesión viva y redirige de
        # vuelta al dashboard -> loop infinito de refrescos.
        session.clear()
        return jsonify({'error': 'Perfil no encontrado'}), 401
    return jsonify(result.data[0])


# Lo que cada rol puede cambiar de su propio perfil. Es una lista blanca y no
# una negra a propósito: con una negra, cada columna nueva del schema nace
# editable y hay que acordarse de prohibirla. Acá nace prohibida.
CAMPOS_PERFIL_EDITABLES = {
    'vecino': ('nombre', 'telefono'),
    'admin': ('nombre', 'nombre_contacto', 'telefono'),
}


@app.route('/api/me', methods=['PUT'])
@require_auth()
def api_me_update():
    """El usuario corrige su propio perfil.

    Nada de email, auth0_id, estado, es_superadmin, consorcio_id, unidad_id ni
    rol: son la identidad y los permisos, no el perfil. Cambiar el email acá,
    además, desincronizaría la fila de la cuenta de Auth0 que la ampara.
    """
    user = session['user']
    table = 'administradores' if user['role'] == 'admin' else 'vecinos'
    permitidos = CAMPOS_PERFIL_EDITABLES.get(user['role'], ())

    d = request.json or {}
    payload = {k: (v or '').strip() if isinstance(v, str) else v
               for k, v in d.items() if k in permitidos}
    if not payload:
        return jsonify({'error': f'No hay nada que actualizar. Campos editables: '
                                 f'{", ".join(permitidos)}'}), 400
    if 'nombre' in payload and not payload['nombre']:
        return jsonify({'error': 'El nombre no puede quedar vacío'}), 400

    res = supabase.table(table).update(payload).eq('auth0_id', user['sub']).execute()
    if not res.data:
        return jsonify({'error': 'Perfil no encontrado'}), 404
    return jsonify(res.data[0])




# ══════════════════════════════════════════════════════════════════════════════
# API — VECINOS DASHBOARD COMPLETO
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/vecinos/mis-unidades')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_mis_unidades():
    """Devuelve todas las unidades del vecino (multi-unidad vía vecinos_unidades o fallback vecinos)."""
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    try:
        res = supabase.table('vecinos_unidades')            .select('*, unidades_funcionales(id, numero, piso, tipo, superficie_m2), consorcios(id, nombre, direccion, encargado_nombre, encargado_tel)')            .eq('vecino_id', vecino_id).eq('activo', True).execute()
        if res.data:
            return jsonify(res.data)
    except Exception:
        pass
    vecino = supabase.table('vecinos').select('consorcio_id, unidad, unidad_id, rol').eq('id', vecino_id).single().execute()
    if not vecino.data or not vecino.data.get('consorcio_id'):
        return jsonify([])
    v = vecino.data
    uf_data = {}
    if v.get('unidad_id'):
        uf_res = supabase.table('unidades_funcionales').select('*').eq('id', v['unidad_id']).single().execute()
        uf_data = uf_res.data or {}
    con_res = supabase.table('consorcios').select('id, nombre, direccion, encargado_nombre, encargado_tel').eq('id', v['consorcio_id']).single().execute()
    return jsonify([{'vecino_id': vecino_id, 'unidad_id': v.get('unidad_id'), 'consorcio_id': v.get('consorcio_id'), 'rol': v.get('rol', 'propietario'), 'activo': True, 'unidades_funcionales': uf_data, 'consorcios': con_res.data or {}}])


@app.route('/api/vecinos/cobros')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_cobros():
    permitidas = unidades_del_vecino()
    unidad_id = request.args.get('unidad_id')
    if unidad_id:
        unidad_propia(unidad_id)
    else:
        unidad_id = permitidas[0] if permitidas else None
    if not unidad_id:
        return jsonify([])
    q = supabase.table('cobros').select('*').eq('unidad_id', unidad_id)
    if request.args.get('desde'):
        q = q.gte('created_at', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('created_at', request.args['hasta'])
    res = q.order('periodo', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/vecinos/cobro-actual')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_cobro_actual():
    permitidas = unidades_del_vecino()
    unidad_id = request.args.get('unidad_id')
    if unidad_id:
        unidad_propia(unidad_id)
    else:
        unidad_id = permitidas[0] if permitidas else None
    if not unidad_id:
        return jsonify(None)
    res = supabase.table('cobros').select('*').eq('unidad_id', unidad_id).in_('estado', ['pendiente', 'vencido', 'en_mora']).order('periodo', desc=True).limit(1).execute()
    return jsonify(res.data[0] if res.data else None)


@app.route('/api/vecinos/cobros/export')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_cobros_export():
    """Baja la cuenta corriente del vecino en Excel o PDF.

    Misma resolución de unidad que /api/vecinos/cobros: `unidad_id` se acepta
    pero pasa por `unidad_propia`, así que un UUID ajeno corta con 404 en vez de
    exportar la cuenta corriente del departamento de al lado. Sin el parámetro
    abre la de siempre, que es la primera de `unidades_del_vecino()`.
    """
    permitidas = unidades_del_vecino()
    unidad_id = request.args.get('unidad_id')
    if unidad_id:
        unidad_propia(unidad_id)
    else:
        unidad_id = permitidas[0] if permitidas else None
    if not unidad_id:
        return jsonify({'error': 'Todavía no tenés una unidad asignada'}), 400

    q = supabase.table('cobros').select('*').eq('unidad_id', unidad_id)
    if request.args.get('desde'):
        q = q.gte('created_at', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('created_at', request.args['hasta'])
    data = q.order('periodo', desc=True).execute().data or []

    headers = ['Período', 'Monto Base', 'Interés', 'Total', 'Estado', 'Vencimiento', 'Fecha Pago']
    rows = [[
        c.get('periodo', ''), c.get('monto_base', 0), c.get('interes_mora', 0),
        c.get('total', 0), c.get('estado', ''),
        c.get('fecha_vencimiento', ''), c.get('fecha_pago', ''),
    ] for c in data]

    if request.args.get('fmt') == 'pdf':
        buf = make_pdf('Mi cuenta corriente', headers, [[str(x) for x in r] for r in rows])
        return pdf_response(buf, 'mi_cuenta_corriente.pdf')
    wb = make_excel(headers, rows, 'Cuenta corriente')
    return excel_response(wb, 'mi_cuenta_corriente.xlsx')


@app.route('/api/vecinos/cobros/<rid>/cupon')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_cupon_pago(rid):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    vecino_id = get_vecino_id()
    cobro_res = supabase.table('cobros').select('*').eq('id', rid).execute()
    if not cobro_res.data:
        _no_es_tuyo()
    cobro = cobro_res.data[0]
    # El cupón trae el consorcio, el CUIT, la unidad y el importe. Sin este
    # chequeo alcanzaba con probar UUIDs para bajarse el de cualquier vecino.
    unidad_propia(cobro.get('unidad_id'))

    # El cupón lleva el importe, el estado de deuda y los datos del consorcio.
    # Sin este chequeo, un vecino se bajaba el de cualquier unidad cambiando el
    # id en la URL.
    vecino = supabase.table('vecinos').select('unidad_id').eq('id', vecino_id).execute().data
    if not vecino or not vecino[0].get('unidad_id') \
            or vecino[0]['unidad_id'] != cobro.get('unidad_id'):
        _no_es_tuyo()
    uf_data = {}
    if cobro.get('unidad_id'):
        uf_res = supabase.table('unidades_funcionales').select('*, consorcios(nombre, direccion, cuit)').eq('id', cobro['unidad_id']).single().execute()
        uf_data = uf_res.data or {}
    con = (uf_data.get('consorcios') or {})
    vecino = supabase.table('vecinos').select('nombre').eq('id', vecino_id).single().execute()
    v_data = vecino.data or {}
    buf = io.BytesIO()
    doc = SimpleDocTemplate(buf, pagesize=A4, leftMargin=2*cm, rightMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    styles = getSampleStyleSheet()
    elements = []
    elements.append(Paragraph('<b>CUPÓN DE PAGO DE EXPENSAS</b>', styles['Title']))
    elements.append(Spacer(1, 0.5*cm))
    elements.append(Paragraph(f"<b>Consorcio:</b> {con.get('nombre', '')} — {con.get('direccion', '')}", styles['Normal']))
    elements.append(Paragraph(f"<b>CUIT:</b> {con.get('cuit', 'N/A')}", styles['Normal']))
    elements.append(Spacer(1, 0.4*cm))
    elements.append(Paragraph(f"<b>Unidad:</b> {uf_data.get('numero', '')} (Piso {uf_data.get('piso', '')})", styles['Normal']))
    elements.append(Paragraph(f"<b>Vecino:</b> {v_data.get('nombre', '')}", styles['Normal']))
    elements.append(Spacer(1, 0.6*cm))
    data = [['Campo', 'Detalle'], ['Período', cobro.get('periodo', '')], ['Monto Base', f"$ {cobro.get('monto_base', 0):,.2f}"], ['Interés/Mora', f"$ {cobro.get('interes_mora', 0):,.2f}"], ['TOTAL A PAGAR', f"$ {cobro.get('total', 0):,.2f}"], ['Estado', str(cobro.get('estado', '')).upper()], ['Vencimiento', cobro.get('fecha_vencimiento', 'N/A')]]
    t = Table(data, colWidths=[8*cm, 9*cm])
    t.setStyle(TableStyle([('BACKGROUND', (0,0), (-1,0), colors.HexColor('#7C3AED')), ('TEXTCOLOR', (0,0), (-1,0), colors.white), ('FONTNAME', (0,0), (-1,0), 'Helvetica-Bold'), ('FONTNAME', (0,4), (-1,4), 'Helvetica-Bold'), ('FONTSIZE', (0,0), (-1,-1), 10), ('GRID', (0,0), (-1,-1), 0.5, colors.HexColor('#cccccc')), ('ROWBACKGROUNDS', (0,1), (-1,-1), [colors.white, colors.HexColor('#f5f0ff')]), ('ALIGN', (1,0), (1,-1), 'RIGHT')]))
    elements.append(t)
    elements.append(Spacer(1, 0.8*cm))
    elements.append(Paragraph('<i>Para informar su pago, ingrese al panel y use "Informar Pago".</i>', styles['Normal']))
    doc.build(elements)
    return pdf_response(buf, f"cupon_{cobro.get('periodo', '')}_UF{uf_data.get('numero', '')}.pdf")


@app.route('/api/vecinos/medios-pago')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_medios_pago():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    v = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).single().execute()
    cid = (v.data or {}).get('consorcio_id')
    if not cid:
        return jsonify([])
    res = supabase.table('medios_pago').select('*').eq('consorcio_id', cid).eq('activo', True).execute()
    return jsonify(res.data)


@app.route('/api/vecinos/gastos-reporte')
@require_auth(allowed_roles=['vecino'])
def api_vecinos_gastos_reporte():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    v = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).single().execute()
    cid = (v.data or {}).get('consorcio_id')
    if not cid:
        return jsonify([])
    q = supabase.table('gastos').select('id, descripcion, categoria, monto, fecha_gasto, fecha_vencimiento, pagado, metodo_pago, recurrente, frecuencia, notas, archivo_nombre, proveedores(nombre, rubro)').eq('consorcio_id', cid)
    if request.args.get('desde'):
        q = q.gte('fecha_gasto', request.args['desde'])
    if request.args.get('hasta'):
        q = q.lte('fecha_gasto', request.args['hasta'])
    if request.args.get('categoria'):
        q = q.eq('categoria', request.args['categoria'])
    res = q.order('fecha_gasto', desc=True).execute()
    return jsonify(res.data)


# ══════════════════════════════════════════════════════════════════════════════
# API — COMUNICADOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/comunicados')
@require_auth(allowed_roles=['vecino'])
def api_comunicados_list():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    v = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).single().execute()
    cid = (v.data or {}).get('consorcio_id')
    if not cid:
        return jsonify([])
    q = supabase.table('comunicados').select('*').eq('consorcio_id', cid)
    if request.args.get('importante') == 'true':
        q = q.eq('importante', True)
    comunicados = q.order('created_at', desc=True).execute().data or []
    leidos_res = supabase.table('comunicados_leidos').select('comunicado_id').eq('vecino_id', vecino_id).execute()
    leidos_set = {r['comunicado_id'] for r in (leidos_res.data or [])}
    for c in comunicados:
        c['leido'] = c['id'] in leidos_set
    if request.args.get('no_leidos') == 'true':
        comunicados = [c for c in comunicados if not c['leido']]
    return jsonify(comunicados)


@app.route('/api/comunicados/<cid_com>/leer', methods=['POST'])
@require_auth(allowed_roles=['vecino'])
def api_comunicados_leer(cid_com):
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({'error': 'No autenticado'}), 401
    try:
        supabase.table('comunicados_leidos').upsert({'comunicado_id': cid_com, 'vecino_id': vecino_id}, on_conflict='comunicado_id,vecino_id').execute()
    except Exception:
        pass
    return jsonify({'ok': True})


# ══════════════════════════════════════════════════════════════════════════════
# API — AVISOS DE PAGO
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/avisos_pago', methods=['GET'])
@require_auth(allowed_roles=['vecino'])
def api_avisos_pago_list():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    res = supabase.table('avisos_pago').select('id, monto, fecha_pago, medio_pago, estado, created_at, cobro_id').eq('vecino_id', vecino_id).order('created_at', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/avisos_pago', methods=['POST'])
@require_auth(allowed_roles=['vecino'])
def api_avisos_pago_create():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({'error': 'No autenticado'}), 401
    d = request.form if request.content_type and 'multipart' in request.content_type else request.json or {}
    v = supabase.table('vecinos').select('consorcio_id, unidad_id').eq('id', vecino_id).single().execute()
    v_data = v.data or {}
    payload = {'vecino_id': vecino_id, 'consorcio_id': v_data.get('consorcio_id'), 'unidad_id': v_data.get('unidad_id'), 'cobro_id': d.get('cobro_id') or None, 'monto': float(d.get('monto', 0)) if d.get('monto') else None, 'fecha_pago': d.get('fecha_pago') or None, 'medio_pago': d.get('medio_pago', ''), 'observaciones': d.get('observaciones', ''), 'estado': 'pendiente'}
    archivo = request.files.get('comprobante') if hasattr(request, 'files') and request.files else None
    if archivo and archivo.filename:
        try:
            file_bytes, mime, nombre = validar_adjunto(archivo)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        payload['adjunto_base64'] = base64.b64encode(file_bytes).decode('utf-8')
        payload['adjunto_nombre'] = nombre
        payload['adjunto_mime'] = mime
    res = supabase.table('avisos_pago').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


# ══════════════════════════════════════════════════════════════════════════════
# API — RECLAMOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/reclamos', methods=['GET'])
@require_auth(allowed_roles=['vecino'])
def api_reclamos_list():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    q = supabase.table('reclamos').select('id, titulo, descripcion, categoria, estado, respuesta_admin, adjunto_nombre, created_at, updated_at').eq('vecino_id', vecino_id)
    if request.args.get('estado'):
        q = q.eq('estado', request.args['estado'])
    res = q.order('created_at', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/reclamos', methods=['POST'])
@require_auth(allowed_roles=['vecino'])
def api_reclamos_create():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({'error': 'No autenticado'}), 401
    d = request.form if request.content_type and 'multipart' in request.content_type else request.json or {}
    titulo = (d.get('titulo') or '').strip()
    descripcion = (d.get('descripcion') or '').strip()
    if not titulo or not descripcion:
        return jsonify({'error': 'Título y descripción son obligatorios'}), 400
    v = supabase.table('vecinos').select('consorcio_id, unidad_id').eq('id', vecino_id).single().execute()
    v_data = v.data or {}
    payload = {'vecino_id': vecino_id, 'consorcio_id': v_data.get('consorcio_id'), 'unidad_id': v_data.get('unidad_id'), 'titulo': titulo, 'descripcion': descripcion, 'categoria': d.get('categoria', 'otro'), 'estado': 'activo'}
    archivo = request.files.get('adjunto') if hasattr(request, 'files') and request.files else None
    if archivo and archivo.filename:
        try:
            file_bytes, mime, nombre = validar_adjunto(archivo)
        except ValueError as e:
            return jsonify({'error': str(e)}), 400
        payload['adjunto_base64'] = base64.b64encode(file_bytes).decode('utf-8')
        payload['adjunto_nombre'] = nombre
        payload['adjunto_mime'] = mime
    res = supabase.table('reclamos').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/reclamos/<rid>', methods=['DELETE'])
@require_auth(allowed_roles=['vecino'])
def api_reclamos_delete(rid):
    vecino_id = get_vecino_id()
    reclamo = supabase.table('reclamos').select('vecino_id, estado').eq('id', rid).single().execute()
    if not reclamo.data or reclamo.data['vecino_id'] != vecino_id:
        return jsonify({'error': 'Sin permiso'}), 403
    if reclamo.data['estado'] not in ('activo',):
        return jsonify({'error': 'Solo se pueden cancelar reclamos activos'}), 400
    supabase.table('reclamos').update({'estado': 'cerrado', 'updated_at': now_iso()}).eq('id', rid).execute()
    return jsonify({'ok': True})


@app.route('/api/reclamos/<rid>/adjunto')
@require_auth(allowed_roles=['vecino'])
def api_reclamos_adjunto(rid):
    vecino_id = get_vecino_id()
    reclamo = supabase.table('reclamos').select('vecino_id, adjunto_base64, adjunto_nombre, adjunto_mime').eq('id', rid).single().execute()
    if not reclamo.data or reclamo.data['vecino_id'] != vecino_id:
        return jsonify({'error': 'No encontrado'}), 404
    if not reclamo.data.get('adjunto_base64'):
        return jsonify({'error': 'Sin adjunto'}), 404
    return enviar_adjunto(reclamo.data['adjunto_base64'],
                          reclamo.data.get('adjunto_nombre', 'adjunto'),
                          reclamo.data.get('adjunto_mime'))


# ══════════════════════════════════════════════════════════════════════════════
# API — VOTACIONES & VOTOS
# ══════════════════════════════════════════════════════════════════════════════

def _votacion_vigente(votacion) -> bool:
    """Si todavía se puede votar.

    `estado` no alcanza: nada lo mueve solo, así que una votación con la fecha
    límite vencida seguía figurando 'activa' y aceptando votos para siempre.
    `fecha_limite` es un DATE y se cuenta inclusive: el último día se vota.
    """
    if votacion.get('estado') != 'activa':
        return False
    limite = votacion.get('fecha_limite')
    if not limite:
        return True
    return str(limite)[:10] >= date.today().isoformat()


@app.route('/api/votaciones')
@require_auth(allowed_roles=['vecino'])
def api_votaciones_list():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    v = supabase.table('vecinos').select('consorcio_id, unidad_id').eq('id', vecino_id).single().execute()
    v_data = v.data or {}
    cid = v_data.get('consorcio_id')
    if not cid:
        return jsonify([])
    q = supabase.table('votaciones').select('*').eq('consorcio_id', cid)
    if request.args.get('estado'):
        q = q.eq('estado', request.args['estado'])
    votaciones = q.order('created_at', desc=True).execute().data or []
    unidad_id = v_data.get('unidad_id')
    vot_ids = [vot['id'] for vot in votaciones]
    votos_por_votacion = {}
    if vot_ids:
        todos_votos = supabase.table('votos').select('votacion_id, opcion, unidad_id').in_('votacion_id', vot_ids).execute().data or []
        for voto in todos_votos:
            votos_por_votacion.setdefault(voto['votacion_id'], []).append(voto)
    for vot in votaciones:
        votos = votos_por_votacion.get(vot['id'], [])
        conteo = {}
        for voto in votos:
            op = voto['opcion']
            conteo[op] = conteo.get(op, 0) + 1
        vot['conteo_votos'] = conteo
        vot['total_votos'] = len(votos)
        # Se calcula acá y no en el navegador para que la pantalla y el endpoint
        # que recibe el voto no puedan discrepar sobre si la votación sigue viva.
        vot['vigente'] = _votacion_vigente(vot)
        vot['vencida'] = (vot.get('estado') == 'activa' and not vot['vigente'])
        necesarios = vot.get('votos_necesarios')
        vot['falta_quorum'] = max(0, int(necesarios) - len(votos)) if necesarios else 0
        vot['ya_vote'] = False
        if unidad_id:
            mi_voto = next((voto for voto in votos if voto.get('unidad_id') == unidad_id), None)
            if mi_voto:
                vot['ya_vote'] = True
                vot['mi_opcion'] = mi_voto['opcion']
    return jsonify(votaciones)


@app.route('/api/votaciones/<vid>/votar', methods=['POST'])
@require_auth(allowed_roles=['vecino'])
def api_votaciones_votar(vid):
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({'error': 'No autenticado'}), 401
    d = request.json or {}
    opcion = (d.get('opcion') or '').strip()
    if not opcion:
        return jsonify({'error': 'La opción es obligatoria'}), 400
    v = supabase.table('vecinos').select('unidad_id, consorcio_id').eq('id', vecino_id).single().execute()
    v_data = v.data or {}
    unidad_id = v_data.get('unidad_id')

    # Sin unidad no hay voto. El control de abajo es "una unidad, un voto", así
    # que con unidad_id en None no aplicaba a nadie: un vecino todavía en
    # "Pendiente" podía votar la misma votación tantas veces como quisiera.
    if not unidad_id:
        return jsonify({'error': 'Necesitás tener una unidad asignada para votar. '
                                 'Pedile al administrador que te asigne la tuya.'}), 403

    votacion_res = supabase.table('votaciones').select('*').eq('id', vid).single().execute()
    if not votacion_res.data:
        return jsonify({'error': 'Votación no encontrada'}), 404
    votacion = votacion_res.data

    # La votación se buscaba sólo por id: con el UUID de la asamblea del edificio
    # de al lado, el voto entraba. Se contesta 404 y no 403 para no confirmar
    # que ese id existe.
    if votacion.get('consorcio_id') != v_data.get('consorcio_id'):
        return jsonify({'error': 'Votación no encontrada'}), 404

    if not _votacion_vigente(votacion):
        limite = votacion.get('fecha_limite')
        if votacion.get('estado') == 'activa' and limite:
            return jsonify({'error': f'La votación venció el {str(limite)[:10]}'}), 400
        return jsonify({'error': 'La votación ya no está activa'}), 400
    opciones_validas = votacion.get('opciones') or ['Si', 'No', 'Abstención']
    if opcion not in opciones_validas:
        return jsonify({'error': f'Opción inválida. Opciones: {opciones_validas}'}), 400
    ya_voto = supabase.table('votos').select('id').eq('votacion_id', vid).eq('unidad_id', unidad_id).execute()
    if ya_voto.data:
        return jsonify({'error': 'Tu unidad ya emitió un voto en esta votación'}), 409
    try:
        res = supabase.table('votos').insert({'votacion_id': vid, 'vecino_id': vecino_id, 'unidad_id': unidad_id, 'opcion': opcion}).execute()
        return jsonify(res.data[0] if res.data else {}), 201
    except Exception:
        return jsonify({'error': 'Ya votaste en esta votación'}), 409


# ══════════════════════════════════════════════════════════════════════════════
# API — ARCHIVOS DEL CONSORCIO
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/archivos')
@require_auth(allowed_roles=['vecino'])
def api_archivos_list():
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify([])
    v = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).single().execute()
    cid = (v.data or {}).get('consorcio_id')
    if not cid:
        return jsonify([])
    q = supabase.table('archivos_consorcio').select('id, categoria, nombre, mime_type, created_at').eq('consorcio_id', cid)
    if request.args.get('categoria'):
        q = q.eq('categoria', request.args['categoria'])
    res = q.order('created_at', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/archivos/<aid>/descargar')
@require_auth(allowed_roles=['vecino'])
def api_archivos_descargar(aid):
    vecino_id = get_vecino_id()
    if not vecino_id:
        return jsonify({'error': 'No autenticado'}), 401
    archivo_res = supabase.table('archivos_consorcio').select('*').eq('id', aid).single().execute()
    if not archivo_res.data:
        return jsonify({'error': 'Archivo no encontrado'}), 404
    archivo = archivo_res.data
    v = supabase.table('vecinos').select('consorcio_id').eq('id', vecino_id).single().execute()
    if (v.data or {}).get('consorcio_id') != archivo['consorcio_id']:
        return jsonify({'error': 'Sin permiso'}), 403
    return enviar_adjunto(archivo['archivo_base64'],
                          archivo.get('nombre', 'archivo'),
                          archivo.get('mime_type'),
                          forzar_descarga=True)


# ══════════════════════════════════════════════════════════════════════════════
# API — ADMIN: COMUNICADOS, VOTACIONES, ARCHIVOS, MEDIOS DE PAGO, RECLAMOS
# ══════════════════════════════════════════════════════════════════════════════

@app.route('/api/admin/comunicados', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_admin_comunicados_list():
    admin_id = get_admin_id()
    q = supabase.table('comunicados').select('*').eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    return jsonify(q.order('created_at', desc=True).execute().data)


@app.route('/api/admin/comunicados', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_admin_comunicados_create():
    admin_id = get_admin_id()
    d = request.json or {}
    payload = {'consorcio_id': d.get('consorcio_id'), 'admin_id': admin_id, 'titulo': (d.get('titulo') or '').strip(), 'cuerpo': (d.get('cuerpo') or '').strip(), 'importante': bool(d.get('importante', False))}
    if not payload['titulo'] or not payload['cuerpo'] or not payload['consorcio_id']:
        return jsonify({'error': 'Faltan campos obligatorios'}), 400
    # El comunicado lo leen los vecinos del consorcio, no el administrador: sin
    # este chequeo se podía publicar un aviso en la cartelera de un edificio
    # ajeno, firmado por su administración.
    consorcio_propio(payload['consorcio_id'], 'id')
    res = supabase.table('comunicados').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/admin/comunicados/<cid_com>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_admin_comunicados_delete(cid_com):
    admin_id = get_admin_id()
    supabase.table('comunicados').delete().eq('id', cid_com).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/admin/votaciones', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_admin_votaciones_create():
    admin_id = get_admin_id()
    d = request.json or {}
    payload = {'consorcio_id': d.get('consorcio_id'), 'admin_id': admin_id, 'titulo': (d.get('titulo') or '').strip(), 'descripcion': d.get('descripcion', ''), 'opciones': d.get('opciones', ['Si', 'No', 'Abstención']), 'fecha_limite': d.get('fecha_limite') or None, 'votos_necesarios': d.get('votos_necesarios') or None, 'estado': 'activa'}
    consorcio_propio(payload['consorcio_id'], 'id')
    res = supabase.table('votaciones').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/admin/votaciones/<vid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_admin_votaciones_update(vid):
    admin_id = get_admin_id()
    d = request.json or {}
    allowed = ('titulo', 'descripcion', 'estado', 'fecha_limite', 'votos_necesarios')
    payload = {k: v for k, v in d.items() if k in allowed}
    res = supabase.table('votaciones').update(payload).eq('id', vid).eq('admin_id', admin_id).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/admin/archivos', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_admin_archivos_create():
    admin_id = get_admin_id()
    d = request.form
    archivo = request.files.get('archivo')
    if not archivo or not archivo.filename:
        return jsonify({'error': 'Se requiere un archivo'}), 400
    consorcio_propio(d.get('consorcio_id'), 'id')
    try:
        file_bytes, mime, nombre = validar_adjunto(archivo)
    except ValueError as e:
        return jsonify({'error': str(e)}), 400
    payload = {'consorcio_id': d.get('consorcio_id'), 'admin_id': admin_id, 'categoria': d.get('categoria', 'otros'), 'nombre': d.get('nombre') or nombre, 'archivo_base64': base64.b64encode(file_bytes).decode('utf-8'), 'mime_type': mime}
    res = supabase.table('archivos_consorcio').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/admin/archivos/<aid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_admin_archivos_delete(aid):
    admin_id = get_admin_id()
    supabase.table('archivos_consorcio').delete().eq('id', aid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/admin/medios-pago', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_admin_medios_pago_list():
    admin_id = get_admin_id()
    q = supabase.table('medios_pago').select('*').eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    return jsonify(q.order('nombre').execute().data)


@app.route('/api/admin/medios-pago', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_admin_medios_pago_create():
    admin_id = get_admin_id()
    d = request.json or {}
    payload = {'consorcio_id': d.get('consorcio_id'), 'admin_id': admin_id, 'nombre': (d.get('nombre') or '').strip(), 'descripcion': d.get('descripcion', ''), 'activo': bool(d.get('activo', True))}
    # Un medio de pago es la cuenta a la que los vecinos giran la expensa. Meter
    # uno en un consorcio ajeno es redirigir su cobranza.
    consorcio_propio(payload['consorcio_id'], 'id')
    res = supabase.table('medios_pago').insert(payload).execute()
    return jsonify(res.data[0] if res.data else {}), 201


@app.route('/api/admin/medios-pago/<mid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_admin_medios_pago_delete(mid):
    admin_id = get_admin_id()
    supabase.table('medios_pago').delete().eq('id', mid).eq('admin_id', admin_id).execute()
    return jsonify({'ok': True})


@app.route('/api/admin/reclamos')
@require_auth(allowed_roles=['admin'])
def api_admin_reclamos_list():
    admin_id = get_admin_id()
    cid = request.args.get('consorcio_id')
    # Columnas explícitas y no `*`: con el asterisco cada fila del listado se
    # traía adjunto_base64 entero, así que abrir la pantalla con veinte reclamos
    # con foto bajaba veinte fotos que nadie iba a mirar. El adjunto se pide de
    # a uno por /api/admin/reclamos/<rid>/adjunto. `consorcios(nombre)` viene
    # para poder mostrar de qué edificio es cada reclamo.
    COLS = ('id, consorcio_id, vecino_id, unidad_id, titulo, descripcion, categoria, '
            'estado, respuesta_admin, adjunto_nombre, created_at, updated_at, '
            'vecinos(nombre, email, unidad), consorcios(nombre)')
    if cid:
        consorcio_propio(cid, 'id')
        q = supabase.table('reclamos').select(COLS).eq('consorcio_id', cid)
    else:
        cons = supabase.table('consorcios').select('id').eq('admin_id', admin_id).execute().data or []
        cids = [c['id'] for c in cons]
        if not cids:
            return jsonify([])
        q = supabase.table('reclamos').select(COLS).in_('consorcio_id', cids)
    if request.args.get('estado'):
        q = q.eq('estado', request.args['estado'])
    res = q.order('created_at', desc=True).execute()
    return jsonify(res.data)


@app.route('/api/admin/reclamos/<rid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_admin_reclamos_update(rid):
    fila_de_consorcio_propio('reclamos', rid)
    d = request.json or {}
    allowed = ('estado', 'respuesta_admin')
    payload = {k: v for k, v in d.items() if k in allowed}
    payload['updated_at'] = now_iso()
    res = supabase.table('reclamos').update(payload).eq('id', rid).execute()
    actualizado = res.data[0] if res.data else {}

    # El aviso va después de guardar y sin poder voltear la respuesta: el
    # reclamo ya se movió, y que Resend esté caído no es motivo para que el
    # administrador vea un error sobre algo que sí quedó hecho.
    if actualizado:
        try:
            _mail_respuesta_reclamo(actualizado, payload.get('estado'),
                                    payload.get('respuesta_admin') or '')
        except Exception:
            app.logger.exception('Falló el aviso del reclamo %s', rid)

    return jsonify(actualizado)


@app.route('/api/admin/reclamos/<rid>/adjunto')
@require_auth(allowed_roles=['admin'])
def api_admin_reclamos_adjunto(rid):
    """La foto que adjuntó el vecino, para el administrador que lo atiende.

    /api/reclamos/<rid>/adjunto es allowed_roles=['vecino'] y compara contra el
    vecino dueño, así que el admin no tenía forma de abrir el adjunto: el
    reclamo con foto le llegaba sin la foto, que suele ser todo el reclamo.

    Sale por `enviar_adjunto`, que vuelve a deducir el mime del contenido: un
    `text/html` guardado antes de que existiera la validación no se abre.
    """
    fila_de_consorcio_propio('reclamos', rid)
    reclamo = supabase.table('reclamos') \
        .select('adjunto_base64, adjunto_nombre, adjunto_mime') \
        .eq('id', rid).single().execute().data or {}
    if not reclamo.get('adjunto_base64'):
        return jsonify({'error': 'Sin adjunto'}), 404
    return enviar_adjunto(reclamo['adjunto_base64'],
                          reclamo.get('adjunto_nombre', 'adjunto'),
                          reclamo.get('adjunto_mime'))


@app.route('/api/admin/avisos-pago')
@require_auth(allowed_roles=['admin'])
def api_admin_avisos_pago_list():
    admin_id = get_admin_id()
    cid = request.args.get('consorcio_id')
    if cid:
        consorcio_propio(cid, 'id')
        q = supabase.table('avisos_pago').select('id, consorcio_id, vecino_id, unidad_id, cobro_id, monto, fecha_pago, medio_pago, observaciones, adjunto_nombre, adjunto_mime, estado, created_at, vecinos(nombre, email, unidad)').eq('consorcio_id', cid)
    else:
        cons = supabase.table('consorcios').select('id').eq('admin_id', admin_id).execute().data or []
        cids = [c['id'] for c in cons]
        if not cids:
            return jsonify([])
        q = supabase.table('avisos_pago').select('id, consorcio_id, vecino_id, unidad_id, cobro_id, monto, fecha_pago, medio_pago, observaciones, adjunto_nombre, adjunto_mime, estado, created_at, vecinos(nombre, email, unidad)').in_('consorcio_id', cids)
    res = q.order('created_at', desc=True).execute()
    return jsonify(res.data)


# El vecino ve este estado en su pantalla, así que dejar pasar cualquier string
# significaba mostrarle una palabra que su UI no sabe pintar.
ESTADOS_AVISO_PAGO = ('pendiente', 'aceptado', 'rechazado')


@app.route('/api/admin/avisos-pago/<aid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_admin_avisos_pago_update(aid):
    fila_de_consorcio_propio('avisos_pago', aid)
    d = request.json or {}
    estado = d.get('estado')
    if estado not in ESTADOS_AVISO_PAGO:
        return jsonify({'error': f'Estado inválido: {", ".join(ESTADOS_AVISO_PAGO)}'}), 400
    res = supabase.table('avisos_pago').update({'estado': estado}).eq('id', aid).execute()
    return jsonify(res.data[0] if res.data else {})



# ══════════════════════════════════════════════════════════════════════════════
# API — LIQUIDACIONES
# ══════════════════════════════════════════════════════════════════════════════

# Mapeo de categorías de gastos a rubros SIPAC simplificados
CATEGORIA_A_RUBRO = {
    'sueldos': (1, 'Remuneraciones y cargas sociales'),
    'aportes': (1, 'Remuneraciones y cargas sociales'),
    'electricidad': (2, 'Servicios públicos'),
    'gas': (2, 'Servicios públicos'),
    'agua': (2, 'Servicios públicos'),
    'internet': (2, 'Servicios públicos'),
    'limpieza': (3, 'Abonos de servicios'),
    'fumigacion': (3, 'Abonos de servicios'),
    'ascensor': (3, 'Abonos de servicios'),
    'mantenimiento': (4, 'Mantenimiento partes comunes'),
    'reparacion': (5, 'Reparaciones en UF'),
    'bancario': (6, 'Gastos bancarios'),
    'honorarios': (8, 'Gastos de administración'),
    'seguro': (9, 'Seguros'),
    'impuesto': (10, 'Otros gastos'),
    'otro': (10, 'Otros gastos'),
}

# Categorías simplificadas para el resumen del vecino
RUBRO_A_CATEGORIA_SIMPLE = {
    1: 'Personal',
    2: 'Servicios',
    3: 'Servicios',
    4: 'Mantenimiento',
    5: 'Mantenimiento',
    6: 'Administración',
    7: 'Administración',
    8: 'Administración',
    9: 'Seguros',
    10: 'Otros',
}


@app.route('/api/liquidaciones', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_liquidaciones_list():
    admin_id = get_admin_id()
    q = supabase.table('liquidaciones').select('*, consorcios(nombre, direccion)').eq('admin_id', admin_id)
    if request.args.get('consorcio_id'):
        q = q.eq('consorcio_id', request.args['consorcio_id'])
    # Se ordena por created_at (no por numero_revision) para que el listado siga
    # funcionando aunque todavía no se haya corrido supabase_schema_v8.sql.
    res = q.order('periodo', desc=True).order('created_at', desc=True).execute()
    return jsonify(res.data)


def _periodo_rango(periodo):
    """'2026-07' → ('2026-07-01', '2026-08-01') para filtrar gastos del mes."""
    year, month = periodo.split('-')[:2]
    desde = f'{year}-{month}-01'
    if int(month) == 12:
        hasta = f'{int(year)+1}-01-01'
    else:
        hasta = f'{year}-{int(month)+1:02d}-01'
    return desde, hasta


def _liquidaciones_del_periodo(consorcio_id, periodo):
    """Liquidaciones existentes de ese consorcio+período (todas sus revisiones)."""
    return supabase.table('liquidaciones').select('*') \
        .eq('consorcio_id', consorcio_id).eq('periodo', periodo).execute().data


def _ultima_revision(liquidaciones):
    """Mayor numero_revision del set. Devuelve 0 si no hay ninguna.

    Se usa .get() con fallback a 1 para que siga funcionando contra una base a la
    que todavía no se le corrió supabase_schema_v8.sql.
    """
    return max([int(l.get('numero_revision') or 1) for l in liquidaciones], default=0)


def _gastos_ya_enviados(consorcio_id, periodo, excluir_liq_id=None):
    """IDs de gastos que ya viajaron en un resumen efectivamente enviado por email.

    Se apoya en `resumen_envios.estado='enviado'` (no en `liquidaciones.estado`) para que
    una liquidación en borrador o cuyo envío falló no marque sus gastos como ya avisados.
    """
    liqs = supabase.table('liquidaciones').select('id') \
        .eq('consorcio_id', consorcio_id).eq('periodo', periodo).execute().data
    liq_ids = [l['id'] for l in liqs if l['id'] != excluir_liq_id]
    if not liq_ids:
        return set()

    envios = supabase.table('resumen_envios').select('liquidacion_id') \
        .in_('liquidacion_id', liq_ids).eq('estado', 'enviado').execute().data
    liq_enviadas = {e['liquidacion_id'] for e in envios}
    if not liq_enviadas:
        return set()

    rubros = supabase.table('liquidacion_rubros').select('id') \
        .in_('liquidacion_id', list(liq_enviadas)).execute().data
    rubro_ids = [r['id'] for r in rubros]
    if not rubro_ids:
        return set()

    items = supabase.table('liquidacion_items').select('gasto_id') \
        .in_('rubro_id', rubro_ids).execute().data
    return {it['gasto_id'] for it in items if it.get('gasto_id')}


@app.route('/api/liquidaciones/gastos-disponibles', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_liquidacion_gastos_disponibles():
    """Gastos del período con la marca de si ya se avisaron por email.

    Alimenta el selector del modal "Nueva liquidación": los `ya_enviado=False`
    vienen tildados por defecto, los `ya_enviado=True` destildados pero elegibles.
    """
    admin_id = get_admin_id()
    consorcio_id = request.args.get('consorcio_id')
    periodo = request.args.get('periodo')
    if not consorcio_id or not periodo:
        return jsonify({'error': 'Faltan consorcio_id y periodo'}), 400
    consorcio_propio(consorcio_id, 'id')

    # El admin puede venir directo a liquidar sin pasar por Gastos; si no
    # generamos acá, el gasto recurrente del mes no existiría todavía y la
    # liquidación saldría sin él.
    _generar_recurrentes_silencioso(admin_id)

    desde, hasta = _periodo_rango(periodo)
    gastos = supabase.table('gastos') \
        .select('id, descripcion, categoria, monto, fecha_gasto, pagado, unidad_id, '
                'tarifa_confirmada, proveedores(nombre), unidades_funcionales(numero)') \
        .eq('consorcio_id', consorcio_id) \
        .eq('admin_id', admin_id) \
        .gte('fecha_gasto', desde) \
        .lt('fecha_gasto', hasta) \
        .order('fecha_gasto').execute().data

    ya = _gastos_ya_enviados(consorcio_id, periodo)
    for g in gastos:
        g['ya_enviado'] = g['id'] in ya

    previas = _liquidaciones_del_periodo(consorcio_id, periodo)
    ultima_rev = _ultima_revision(previas)

    return jsonify({
        'gastos': gastos,
        'liquidaciones_previas': len(previas),
        'proxima_revision': ultima_rev + 1,
        'ya_enviados': len(ya),
    })


@app.route('/api/liquidaciones', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_liquidaciones_create():
    """Crea una liquidación nueva. Opcionalmente auto-genera rubros desde gastos del período.

    Se admite más de una liquidación por consorcio+período (revisiones): la 2da y
    siguientes sirven para liquidar gastos que se cargaron después del primer envío.
    """
    admin_id = get_admin_id()
    d = request.json
    consorcio_id = consorcio_propio(d.get('consorcio_id'), 'id')['id']
    periodo = d['periodo']  # "2026-07"
    gastos_ids = d.get('gastos_ids')  # None = todos los gastos del período

    if gastos_ids is not None and not gastos_ids:
        return jsonify({'error': 'Seleccioná al menos un gasto para liquidar'}), 400

    # Número de revisión: 1 para la primera del período, +1 para cada reliquidación
    previas = _liquidaciones_del_periodo(consorcio_id, periodo)
    numero_revision = _ultima_revision(previas) + 1

    # Crear cabecera
    payload = {
        'consorcio_id': consorcio_id,
        'admin_id': admin_id,
        'periodo': periodo,
        'numero_revision': numero_revision,
        'fecha_vencimiento_1': d.get('fecha_vencimiento_1'),
        'fecha_vencimiento_2': d.get('fecha_vencimiento_2'),
        'interes_2_vto': d.get('interes_2_vto', 0),
        'saldo_inicial': d.get('saldo_inicial', 0),
        'notas': d.get('notas', ''),
        'estado': 'borrador',
    }
    try:
        liq_res = supabase.table('liquidaciones').insert(payload).execute()
    except Exception as e:
        msg = str(e)
        # El texto crudo de Supabase se adjunta siempre: sin él, un fallo acá es
        # indistinguible de otro y no hay forma de diagnosticarlo desde la UI.
        if 'PGRST204' in msg or 'schema cache' in msg:
            return jsonify({'error':
                'PostgREST tiene el esquema viejo en caché y no ve la columna numero_revision. '
                'Corré  NOTIFY pgrst, \'reload schema\';  en el SQL Editor de Supabase '
                f'y reintentá. [{msg}]'
            }), 409
        if 'numero_revision' in msg and ('does not exist' in msg or '42703' in msg):
            return jsonify({'error':
                'Falta correr supabase_schema_v8.sql en Supabase '
                f'(la columna numero_revision no existe). [{msg}]'
            }), 409
        if 'duplicate key' in msg or '23505' in msg:
            return jsonify({'error':
                'Ya existe una liquidación para este consorcio, período y revisión. Si ya corriste '
                'supabase_schema_v8.sql, puede haber quedado la UNIQUE vieja de v7 sin bajar. '
                f'[{msg}]'
            }), 409
        return jsonify({'error': f'Error al crear la liquidación: {msg}'}), 500

    liq = liq_res.data[0] if liq_res.data else {}
    liq_id = liq.get('id')

    if not liq_id:
        return jsonify({'error': 'Error al crear liquidación'}), 500

    # Auto-generar rubros desde gastos del período.
    # Si algo falla acá la cabecera ya está insertada: se borra para no dejar una
    # liquidación vacía que después ocupe un número de revisión y confunda.
    if d.get('auto_generar', True):
        try:
            _generar_rubros_desde_gastos(liq_id, consorcio_id, periodo, admin_id, gastos_ids=gastos_ids)
            _generar_prorrateo(liq_id, consorcio_id, periodo, numero_revision=numero_revision)
            _recalcular_totales(liq_id)
        except Exception as e:
            supabase.table('liquidaciones').delete().eq('id', liq_id).execute()
            if _falta_schema_v9(str(e)):
                return jsonify({'error': f'{ERROR_FALTA_V9} [{e}]'}), 409
            return jsonify({'error':
                f'La liquidación se creó pero falló al generar rubros/prorrateo: {e}'
            }), 500

    # Refetch con datos completos
    liq = supabase.table('liquidaciones').select('*, consorcios(nombre, direccion)').eq('id', liq_id).single().execute().data
    return jsonify(liq), 201


def _generar_rubros_desde_gastos(liq_id, consorcio_id, periodo, admin_id, gastos_ids=None):
    """Agrupa los gastos del consorcio en el período por categoría → rubros.

    `gastos_ids` acota la liquidación a esos gastos (selección manual del admin).
    Con None entran todos los del período, como antes.
    """
    desde, hasta = _periodo_rango(periodo)

    gastos = supabase.table('gastos').select('*, proveedores(nombre, cuit)') \
        .eq('consorcio_id', consorcio_id) \
        .eq('admin_id', admin_id) \
        .gte('fecha_gasto', desde) \
        .lt('fecha_gasto', hasta) \
        .execute().data

    if gastos_ids is not None:
        elegidos = set(gastos_ids)
        gastos = [g for g in gastos if g.get('id') in elegidos]

    # Agrupar por rubro
    rubros_dict = {}  # {numero_rubro: {nombre, items: [{descripcion, monto, gasto_id, unidad_id}]}}
    for g in gastos:
        cat = (g.get('categoria') or 'otro').lower()
        num, nombre = CATEGORIA_A_RUBRO.get(cat, (10, 'Otros gastos'))
        if num not in rubros_dict:
            rubros_dict[num] = {'nombre': nombre, 'items': []}
        prov = g.get('proveedores') or {}
        rubros_dict[num]['items'].append({
            'descripcion': g.get('descripcion', ''),
            'monto': float(g.get('monto', 0)),
            'gasto_id': g.get('id'),
            # El alcance se congela acá: el prorrateo se calcula sobre los ítems,
            # así que editar el gasto después no cambia una liquidación emitida.
            'unidad_id': g.get('unidad_id'),
            # Y por lo mismo se congela el coeficiente: recategorizar un gasto
            # el mes que viene no puede mover una liquidación ya emitida.
            'coeficiente': (g.get('coeficiente') or 'A').upper(),
            # El detalle que la Ley 941 obliga a rendir y que hasta ahora se
            # perdía: el resumen mostraba categoría e importe, nada más.
            'proveedor_nombre': prov.get('nombre') or '',
            'proveedor_cuit': prov.get('cuit') or '',
            'comprobante_tipo': g.get('comprobante_tipo') or '',
            'comprobante_numero': g.get('comprobante_numero') or '',
            'fecha_pago': g.get('fecha_pago'),
        })

    total_general = sum(
        sum(it['monto'] for it in r['items'])
        for r in rubros_dict.values()
    )

    # Insertar rubros e items
    for num in sorted(rubros_dict.keys()):
        r = rubros_dict[num]
        subtotal = sum(it['monto'] for it in r['items'])
        pct = (subtotal / total_general * 100) if total_general > 0 else 0

        rubro_res = supabase.table('liquidacion_rubros').insert({
            'liquidacion_id': liq_id,
            'numero_rubro': num,
            'nombre': r['nombre'],
            'subtotal': subtotal,
            'porcentaje_sobre_total': round(pct, 2),
        }).execute()
        rubro_id = rubro_res.data[0]['id'] if rubro_res.data else None

        if rubro_id:
            items_payload = [{
                'rubro_id': rubro_id,
                'descripcion': it['descripcion'],
                'monto': it['monto'],
                'gasto_id': it['gasto_id'],
                'unidad_id': it['unidad_id'],
                'coeficiente': it['coeficiente'],
                'proveedor_nombre': it['proveedor_nombre'],
                'proveedor_cuit': it['proveedor_cuit'],
                'comprobante_tipo': it['comprobante_tipo'],
                'comprobante_numero': it['comprobante_numero'],
                'fecha_pago': it['fecha_pago'],
            } for it in r['items']]
            if items_payload:
                supabase.table('liquidacion_items').insert(items_payload).execute()


def _egresos_por_alcance(liq_id):
    """Separa los egresos de la liquidación por coeficiente y por UF.

    Devuelve ({coeficiente: total}, {unidad_id: monto}). Lo general se reparte
    según el coeficiente que declara cada gasto; lo particular —el gasto con
    `unidad_id`, que es la reparación de una sola unidad— se le carga entero a
    esa UF sin pasar por ningún reparto.

    El coeficiente es lo que hace que un edificio con cocheras y locales pueda
    liquidar bien: la factura del ascensor va por uno donde la cochera pesa 0,
    y el seguro por otro donde pesa. Los cuatro sistemas del mercado imprimen
    "SUBTOTALES POR COEFICIENTE" justamente por esto.
    """
    rubros = supabase.table('liquidacion_rubros').select('id') \
        .eq('liquidacion_id', liq_id).execute().data
    rubro_ids = [r['id'] for r in rubros]
    if not rubro_ids:
        return {c: 0.0 for c in COEFICIENTES}, {}

    items = supabase.table('liquidacion_items').select('monto, unidad_id, coeficiente') \
        .in_('rubro_id', rubro_ids).execute().data

    por_coeficiente = {c: 0.0 for c in COEFICIENTES}
    particulares = {}
    for it in items:
        monto = float(it.get('monto') or 0)
        unidad_id = it.get('unidad_id')
        if unidad_id:
            particulares[unidad_id] = round(particulares.get(unidad_id, 0.0) + monto, 2)
            continue
        # Un ítem sin coeficiente es de antes de la v18: se reparte por A, que
        # era el único reparto que existía. Así una liquidación vieja sigue
        # dando el mismo número que daba.
        coef = (it.get('coeficiente') or 'A').upper()
        if coef not in por_coeficiente:
            coef = 'A'
        por_coeficiente[coef] += monto

    return ({c: round(v, 2) for c, v in por_coeficiente.items()},
            particulares)


def _repartir(total, pesos):
    """Reparte `total` entre n partes según `pesos`, cerrando los centavos.

    Redondear cada parte por separado casi nunca suma el total (100 entre 3 da
    33.33 tres veces = 99.99). La diferencia se ajusta en la primera parte y se
    devuelve aparte para poder registrarla como `redondeo` en el prorrateo: la
    suma de lo que se le cobra a las UFs tiene que dar exactamente lo gastado.

    Devuelve (partes, ajustes) — ajustes es 0 en todas menos la primera.
    """
    n = len(pesos)
    if n == 0:
        return [], []
    suma_pesos = sum(pesos)
    if suma_pesos <= 0:
        return [0.0] * n, [0.0] * n

    partes = [round(total * p / suma_pesos, 2) for p in pesos]
    ajustes = [0.0] * n
    dif = round(total - sum(partes), 2)
    if dif:
        partes[0] = round(partes[0] + dif, 2)
        ajustes[0] = dif
    return partes, ajustes


def _cargos_de_amenities(consorcio_id, periodo, uf_ids):
    """Lo que cada UF debe por reservar amenities en el período.

    Consorcio Abierto le da columna propia en el prorrateo ("USO AMENITIES") y
    tiene sentido: la reserva del SUM la paga quien la usa, no el edificio.

    Sólo cuenta lo confirmado y con precio cargado. Un amenity sin `precio_uso`
    es gratis, que es como se comportan todos hoy, así que esto no le cobra
    nada a nadie hasta que el administrador ponga un número.
    """
    if not uf_ids:
        return {}
    desde = f'{periodo}-01'
    y, m = periodo.split('-')[:2]
    hasta = f'{int(y)+1}-01-01' if int(m) == 12 else f'{y}-{int(m)+1:02d}-01'

    try:
        amenities = supabase.table('amenities').select('id, precio_uso') \
            .eq('consorcio_id', consorcio_id).execute().data or []
    except Exception:
        # Sin la columna (entorno sin v18) el cargo es 0 y la liquidación sale
        # igual que antes, en vez de romperse entera por un extra.
        return {}
    precios = {a['id']: float(a.get('precio_uso') or 0) for a in amenities}
    cobrables = [aid for aid, p in precios.items() if p > 0]
    if not cobrables:
        return {}

    reservas = supabase.table('reservas_amenities') \
        .select('amenity_id, vecino_id, fecha') \
        .in_('amenity_id', cobrables).eq('estado', 'confirmada') \
        .gte('fecha', desde).lt('fecha', hasta).execute().data or []
    if not reservas:
        return {}

    # La reserva cuelga del vecino; el cargo, de la unidad. El salto es
    # `vecinos.unidad_id`.
    vecino_ids = list({r['vecino_id'] for r in reservas if r.get('vecino_id')})
    if not vecino_ids:
        return {}
    vecinos = supabase.table('vecinos').select('id, unidad_id') \
        .in_('id', vecino_ids).execute().data or []
    unidad_de = {v['id']: v.get('unidad_id') for v in vecinos}

    cargos = {}
    for r in reservas:
        uid = unidad_de.get(r.get('vecino_id'))
        if uid and uid in uf_ids:
            cargos[uid] = round(cargos.get(uid, 0.0) + precios[r['amenity_id']], 2)
    return cargos


def _pesos_del_coeficiente(ufs, coef):
    """Los pesos con que se reparte un coeficiente entre las UFs.

    Si los porcentajes cargados cierran en 100 se usan esos; si están todos en
    0 —que es como vienen— se reparte lineal. La tolerancia de 0,5 absorbe el
    redondeo de cargar 33,333 tres veces.

    Devuelve (pesos, porcentajes_efectivos). Antes se multiplicaba directo por
    el porcentaje, así que sin cargarlo cada unidad recibía $0.
    """
    columna = f'porcentaje_{coef.lower()}'
    cargados = [float(uf.get(columna) or 0) for uf in ufs]
    if abs(sum(cargados) - 100) <= 0.5:
        return cargados, cargados
    n = len(ufs)
    return [1.0] * n, [round(100 / n, 3)] * n


def _generar_prorrateo(liq_id, consorcio_id, periodo, numero_revision=1):
    """Genera la tabla de prorrateo para cada UF del consorcio.

    La fórmula sale de contrastar los cuatro sistemas del mercado, y da igual
    en los cuatro:

        total 1er vto = Σ(expensa de cada coeficiente)
                      + gastos particulares de la unidad
                      + saldo pendiente + intereses
                      + uso de amenities − descuentos
                      + redondeo
        total 2do vto = total 1er vto × (1 + recargo%)

    El redondeo es una columna propia y no un descarte: la suma de lo que se
    le cobra a las UFs tiene que dar exactamente lo gastado.

    El saldo puede quedar negativo —el vecino pagó de más— y se arrastra como
    crédito. Antes se cortaba en 0 y ese pago de más se perdía.

    En una reliquidación del mismo período (numero_revision > 1) NO se vuelve a
    arrastrar el saldo del mes anterior: ya se le cobró en la revisión 1, y
    volver a incluirlo le facturaría dos veces la misma deuda vieja.
    """
    ufs = supabase.table('unidades_funcionales').select('*') \
        .eq('consorcio_id', consorcio_id).order('numero').execute().data
    if not ufs:
        return

    por_coeficiente, particulares = _egresos_por_alcance(liq_id)

    consorcio = supabase.table('consorcios') \
        .select('tasa_interes_mora, dias_gracia_mora, recargo_segundo_vto') \
        .eq('id', consorcio_id).execute().data
    consorcio = consorcio[0] if consorcio else {}
    tasa_interes = float(consorcio.get('tasa_interes_mora') or 0)
    recargo_2do = float(consorcio.get('recargo_segundo_vto') or 0)

    # El recargo del 2do vencimiento puede venir pisado en la liquidación:
    # `interes_2_vto` es por liquidación y el del consorcio es el default.
    liq = supabase.table('liquidaciones').select('interes_2_vto') \
        .eq('id', liq_id).execute().data
    if liq and liq[0].get('interes_2_vto'):
        recargo_2do = float(liq[0]['interes_2_vto'])

    # Un reparto por coeficiente, cada uno con sus propios pesos y su redondeo.
    repartos = {}
    for coef in COEFICIENTES:
        total = por_coeficiente.get(coef, 0.0)
        pesos, pcts = _pesos_del_coeficiente(ufs, coef)
        montos, ajustes = _repartir(total, pesos)
        repartos[coef] = {'montos': montos, 'ajustes': ajustes, 'pcts': pcts}

    # Cobros del período anterior, para el saldo y los intereses.
    year, month = periodo.split('-')[:2]
    if int(month) == 1:
        periodo_ant = f'{int(year)-1}-12'
    else:
        periodo_ant = f'{year}-{int(month)-1:02d}'

    uf_ids = [uf['id'] for uf in ufs]
    cobros_ant_por_uf = {}
    if uf_ids and numero_revision <= 1:
        cobros_ant = supabase.table('cobros') \
            .select('unidad_id, total, estado, fecha_pago') \
            .in_('unidad_id', uf_ids).eq('periodo', periodo_ant).execute().data
        for c in cobros_ant:
            cobros_ant_por_uf.setdefault(c['unidad_id'], c)

    amenities_por_uf = _cargos_de_amenities(consorcio_id, periodo, uf_ids)

    prorrateo_rows = []
    for i, uf in enumerate(ufs):
        particular = particulares.get(uf['id'], 0.0)

        c = cobros_ant_por_uf.get(uf['id'])
        saldo_ant = 0.0
        pago = 0.0
        if c:
            saldo_ant = float(c.get('total', 0))
            if c.get('estado') == 'pagado':
                pago = float(c.get('total', 0))

        # Puede quedar negativo: es un crédito y se arrastra a favor del vecino.
        saldo_pend = round(saldo_ant - pago, 2)

        # Interés sólo sobre deuda real, nunca sobre un saldo a favor.
        interes = round(saldo_pend * tasa_interes / 100, 2) if saldo_pend > 0 and tasa_interes else 0.0

        amenities = amenities_por_uf.get(uf['id'], 0.0)

        expensas = {c2: repartos[c2]['montos'][i] for c2 in COEFICIENTES}
        redondeo = round(sum(repartos[c2]['ajustes'][i] for c2 in COEFICIENTES), 2)

        total_unidad = round(
            sum(expensas.values()) + particular + saldo_pend + interes + amenities, 2)
        total_2do = round(total_unidad * (1 + recargo_2do / 100), 2) if recargo_2do else total_unidad

        prorrateo_rows.append({
            'liquidacion_id': liq_id,
            'unidad_id': uf['id'],
            'saldo_anterior': saldo_ant,
            'pago_realizado': pago,
            'saldo_pendiente': saldo_pend,
            'interes_mora': interes,
            'porcentaje_a': repartos['A']['pcts'][i],
            'expensa_a': expensas['A'],
            'porcentaje_b': repartos['B']['pcts'][i],
            'expensa_b': expensas['B'],
            # `adicional_ordinaria` es el nombre que le puso v7 al reparto del
            # coeficiente C. Se conserva para no romper lo ya emitido.
            'porcentaje_c': repartos['C']['pcts'][i],
            'adicional_ordinaria': expensas['C'],
            'porcentaje_e': repartos['E']['pcts'][i],
            'expensa_e': expensas['E'],
            'gastos_particulares': particular,
            'uso_amenities': amenities,
            'descuentos': 0,
            'extraordinaria': 0,
            'redondeo': redondeo,
            'total_unidad': total_unidad,
            'total_segundo_vto': total_2do,
        })

    if prorrateo_rows:
        supabase.table('liquidacion_prorrateo').insert(prorrateo_rows).execute()


def _recalcular_totales(liq_id):
    """Recalcula total_egresos y saldo_final de la liquidación."""
    rubros = supabase.table('liquidacion_rubros').select('subtotal') \
        .eq('liquidacion_id', liq_id).execute().data
    total_egresos = sum(float(r.get('subtotal', 0)) for r in rubros)

    liq = supabase.table('liquidaciones').select('saldo_inicial, total_ingresos') \
        .eq('id', liq_id).single().execute().data
    saldo_inicial = float(liq.get('saldo_inicial', 0))
    total_ingresos = float(liq.get('total_ingresos', 0))
    saldo_final = saldo_inicial + total_ingresos - total_egresos

    supabase.table('liquidaciones').update({
        'total_egresos': total_egresos,
        'saldo_final': round(saldo_final, 2),
    }).eq('id', liq_id).execute()


@app.route('/api/liquidaciones/<lid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_liquidaciones_update(lid):
    admin_id = get_admin_id()
    d = request.json
    allowed = ('fecha_vencimiento_1', 'fecha_vencimiento_2', 'interes_2_vto',
               'saldo_inicial', 'total_ingresos', 'saldo_bancario', 'saldo_superfondo',
               'saldo_administrador', 'notas', 'estado')
    payload = {k: v for k, v in d.items() if k in allowed}
    res = supabase.table('liquidaciones').update(payload).eq('id', lid).eq('admin_id', admin_id).execute()
    if payload.get('saldo_inicial') is not None or payload.get('total_ingresos') is not None:
        _recalcular_totales(lid)
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/liquidaciones/<lid>', methods=['DELETE'])
@require_auth(allowed_roles=['admin'])
def api_liquidaciones_delete(lid):
    admin_id = get_admin_id()
    res = supabase.table('liquidaciones').delete() \
        .eq('id', lid).eq('admin_id', admin_id).eq('estado', 'borrador').execute()
    if not res.data:
        # Antes devolvía ok igual y la liquidación seguía ahí sin que el admin se enterara.
        return jsonify({'error':
            'Sólo se pueden eliminar liquidaciones en borrador. Esta ya fue publicada: '
            'creá una nueva liquidación del mismo período con los gastos que falten.'
        }), 409
    return jsonify({'ok': True})


def _rubros_con_items(liq_id):
    """Trae los rubros de una liquidación junto con sus items, en 2 queries en vez de 1+N."""
    rubros = supabase.table('liquidacion_rubros').select('*') \
        .eq('liquidacion_id', liq_id).order('numero_rubro').execute().data
    rubro_ids = [r['id'] for r in rubros]
    items_por_rubro = {}
    if rubro_ids:
        todos_items = supabase.table('liquidacion_items') \
            .select('*, unidades_funcionales(numero)') \
            .in_('rubro_id', rubro_ids).execute().data
        for it in todos_items:
            items_por_rubro.setdefault(it['rubro_id'], []).append(it)
    for r in rubros:
        r['items'] = items_por_rubro.get(r['id'], [])
    return rubros


@app.route('/api/liquidaciones/<lid>/rubros', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_liquidacion_rubros(lid):
    liquidacion_propia(lid, 'id')
    return jsonify(_rubros_con_items(lid))


@app.route('/api/liquidaciones/<lid>/prorrateo', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_liquidacion_prorrateo(lid):
    liquidacion_propia(lid, 'id')
    res = supabase.table('liquidacion_prorrateo') \
        .select('*, unidades_funcionales(numero, piso, tipo, vecino_nombre, vecino_email)') \
        .eq('liquidacion_id', lid).order('unidades_funcionales(numero)').execute()
    return jsonify(res.data)


@app.route('/api/liquidaciones/<lid>/prorrateo/<pid>', methods=['PUT'])
@require_auth(allowed_roles=['admin'])
def api_liquidacion_prorrateo_update(lid, pid):
    liquidacion_propia(lid, 'id')
    d = request.json
    allowed = ('saldo_anterior', 'pago_realizado', 'saldo_pendiente', 'interes_mora',
               'porcentaje_a', 'expensa_a', 'porcentaje_c', 'adicional_ordinaria',
               'gastos_particulares', 'extraordinaria', 'redondeo', 'total_unidad')
    payload = {k: v for k, v in d.items() if k in allowed}
    # El `liquidacion_id` también filtra: con la liquidación propia en la URL y
    # un `pid` ajeno en el path, el chequeo de arriba pasaría igual.
    res = supabase.table('liquidacion_prorrateo').update(payload) \
        .eq('id', pid).eq('liquidacion_id', lid).execute()
    return jsonify(res.data[0] if res.data else {})


# ── Resumen personalizado por UF ───────────────────────────────────────────────

def _generar_resumen_html(liq, prorrateo, rubros, consorcio, uf):
    """Genera el HTML del resumen personalizado para una UF.

    Todo lo que viene de la base pasa por `escape()`. Este HTML sale por dos
    puertas y las dos son delicadas: el mail que le llega al vecino, y la
    respuesta de `?format=html`, que el navegador abre como página en el
    dominio de Niddo. La descripción de un gasto la escribe un administrador,
    así que sin escapar alcanzaba con nombrar un gasto `<script>...` para
    ejecutar código en la sesión de todo el edificio.
    """
    periodo_display = liq.get('periodo', '')
    try:
        y, m = periodo_display.split('-')
        meses = ['', 'Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                 'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre']
        periodo_display = f'{meses[int(m)]} {y}'
    except:
        pass

    # Una revisión > 1 es una liquidación complementaria: se emitió después de haber
    # mandado la del mes y factura sólo los gastos que se cargaron más tarde. El mail
    # tiene que decirlo, porque si no el vecino recibe dos veces "Tu expensa este mes"
    # y no puede saber si el segundo reemplaza al primero o se suma.
    es_complementaria = int(liq.get('numero_revision') or 1) > 1

    total_unidad = float(prorrateo.get('total_unidad', 0))
    uf_id = prorrateo.get('unidad_id')

    # Agrupar rubros en categorías simples. Sólo entran los gastos generales del
    # consorcio: los específicos de una UF no se prorratean, así que mezclarlos
    # acá le mostraría al vecino un desglose que no coincide con lo que paga.
    categorias = {}
    particulares_uf = []
    obras_en_curso = []
    for r in rubros:
        cat_simple = RUBRO_A_CATEGORIA_SIMPLE.get(r.get('numero_rubro', 10), 'Otros')

        for it in r.get('items', []):
            monto_item = float(it.get('monto') or 0)
            unidad_item = it.get('unidad_id')
            if unidad_item:
                if unidad_item == uf_id:
                    particulares_uf.append({'desc': it.get('descripcion', ''), 'monto': monto_item})
                continue  # gasto particular de otra UF: no es asunto de ésta
            categorias[cat_simple] = categorias.get(cat_simple, 0) + monto_item

            # Buscar items con cuotas (obras en curso)
            if it.get('es_cuota') and it.get('cuota_actual') and it.get('cuota_total'):
                obras_en_curso.append({
                    'desc': it['descripcion'],
                    'cuota': it['cuota_actual'],
                    'total': it['cuota_total'],
                })

    total_general = sum(categorias.values())
    pct_a = float(prorrateo.get('porcentaje_a', 0))
    gastos_particulares = float(prorrateo.get('gastos_particulares', 0) or 0)
    cat_icons = {'Personal': '👤', 'Servicios': '⚡', 'Mantenimiento': '🔧',
                 'Administración': '📋', 'Seguros': '🛡️', 'Otros': '📦'}

    # Build category rows
    cat_rows = ''
    for cat, monto in sorted(categorias.items(), key=lambda x: -x[1]):
        pct = (monto / total_general * 100) if total_general > 0 else 0
        monto_uf = round(monto * pct_a / 100, 2) if pct_a > 0 else monto
        icon = cat_icons.get(cat, '📦')
        cat_rows += f'''
        <tr>
            <td style="padding:10px 14px;border-bottom:1px solid #f0f0f5;font-size:14px;">{icon} {cat}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #f0f0f5;text-align:right;font-size:14px;font-weight:600;">${monto_uf:,.2f}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #f0f0f5;text-align:right;font-size:13px;color:#888;">{pct:.1f}%</td>
        </tr>'''

    # Gastos que no se prorratean: se le cobran enteros a esta UF y punto. Van en
    # su propio bloque y con el detalle ítem por ítem porque es lo primero que el
    # vecino va a querer discutir si no lo reconoce.
    particulares_html = ''
    if particulares_uf:
        particulares_rows = ''.join(f'''
        <tr>
            <td style="padding:10px 14px;border-bottom:1px solid #f0f0f5;font-size:14px;">{escape(p["desc"])}</td>
            <td style="padding:10px 14px;border-bottom:1px solid #f0f0f5;text-align:right;font-size:14px;font-weight:600;">${p["monto"]:,.2f}</td>
        </tr>''' for p in particulares_uf)
        particulares_html = f'''
<div style="background:#fff;border-radius:12px;margin-top:16px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.05);">
    <h3 style="margin:0 0 6px;font-size:15px;font-weight:700;color:#111;">🔑 Gastos de tu unidad</h3>
    <p style="margin:0 0 12px;font-size:12px;color:#888;">Estos gastos no se reparten entre el edificio: corresponden sólo a tu UF.</p>
    <table style="width:100%;border-collapse:collapse;">
        <tbody>{particulares_rows}</tbody>
        <tfoot>
            <tr>
                <td style="padding:10px 14px;font-size:14px;font-weight:700;">Total</td>
                <td style="padding:10px 14px;text-align:right;font-size:14px;font-weight:700;">${gastos_particulares:,.2f}</td>
            </tr>
        </tfoot>
    </table>
</div>'''

    # Build obras rows
    obras_html = ''
    if obras_en_curso:
        obras_items = ''.join(
            f'<li style="padding:6px 0;font-size:13px;color:#444;">{escape(o["desc"])} — Cuota {escape(o["cuota"])} de {escape(o["total"])}</li>'
            for o in obras_en_curso
        )
        obras_html = f'''
        <div style="margin-top:24px;background:#f8f7ff;border-radius:10px;padding:18px;">
            <h3 style="margin:0 0 10px;font-size:15px;color:#7C3AED;">🏗️ Obras en curso</h3>
            <ul style="margin:0;padding-left:18px;">{obras_items}</ul>
        </div>'''

    saldo_final = float(liq.get('saldo_final', 0))

    aviso_complementaria = ''
    if es_complementaria:
        aviso_complementaria = (
            '<p style="margin:14px 0 0;padding:10px 12px;background:#FEF3C7;border-radius:8px;'
            'font-size:12px;color:#92400E;line-height:1.5;">Este importe <strong>se suma</strong> '
            f'a la expensa de {periodo_display} que ya recibiste: corresponde a gastos del mismo '
            'período que se cargaron después de ese envío. No reemplaza al resumen anterior.</p>'
        )

    # El bloque del fondo se omite en una complementaria: saldo_final se calcula sobre
    # saldo_inicial/total_ingresos propios de esta revisión, que arrancan en cero, así que
    # daría un negativo igual a sus egresos y no el fondo real del consorcio.
    fondo_html = '' if es_complementaria else f'''
<div style="background:#f8f7ff;border-radius:12px;margin-top:16px;padding:18px;text-align:center;">
    <p style="margin:0;font-size:12px;color:#888;text-transform:uppercase;font-weight:600;">Saldo del fondo del consorcio</p>
    <p style="margin:6px 0 0;font-size:22px;font-weight:800;color:#7C3AED;">${saldo_final:,.2f}</p>
</div>'''
    vto1 = liq.get('fecha_vencimiento_1', '—')
    vto2 = liq.get('fecha_vencimiento_2', '—')
    interes_2 = float(liq.get('interes_2_vto', 0))
    banco_cbu = consorcio.get('banco_cbu', '—')
    banco_nombre = consorcio.get('banco_nombre', '—')

    return f'''<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>Resumen de Expensas — {periodo_display}</title>
</head>
<body style="margin:0;padding:0;background:#f5f5fa;font-family:'Segoe UI',Roboto,Helvetica,Arial,sans-serif;">
<div style="max-width:600px;margin:0 auto;padding:20px;">

<!-- Header -->
<div style="background:linear-gradient(135deg,#7C3AED,#10B981);border-radius:14px;padding:28px;color:#fff;text-align:center;">
    <h1 style="margin:0;font-size:22px;font-weight:800;">🏢 {escape(consorcio.get('nombre', ''))}</h1>
    <p style="margin:6px 0 0;font-size:13px;opacity:.85;">{escape(consorcio.get('direccion', ''))}</p>
    <p style="margin:4px 0 0;font-size:13px;opacity:.85;">Período: {periodo_display}</p>
</div>

<!-- Tu expensa -->
<div style="background:#fff;border-radius:12px;margin-top:16px;padding:24px;text-align:center;box-shadow:0 2px 8px rgba(0,0,0,.05);">
    <p style="margin:0;font-size:13px;color:#888;text-transform:uppercase;letter-spacing:.05em;font-weight:600;">{'Expensa complementaria' if es_complementaria else 'Tu expensa este mes'}</p>
    <p style="margin:8px 0 0;font-size:38px;font-weight:800;color:#111;">${total_unidad:,.2f}</p>
    <p style="margin:6px 0 0;font-size:12px;color:#888;">UF {escape(uf.get('numero', ''))} — Piso {escape(uf.get('piso', '—'))} — {escape(uf.get('vecino_nombre', ''))}</p>
    {aviso_complementaria}
</div>

<!-- Desglose por categoría -->
<div style="background:#fff;border-radius:12px;margin-top:16px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.05);">
    <h3 style="margin:0 0 6px;font-size:15px;font-weight:700;color:#111;">📊 Desglose por categoría</h3>
    <p style="margin:0 0 12px;font-size:12px;color:#888;">Los montos son la parte que te toca a vos ({pct_a:.3f}% de los gastos comunes del edificio).</p>
    <table style="width:100%;border-collapse:collapse;">
        <thead>
            <tr style="border-bottom:2px solid #7C3AED;">
                <th style="padding:8px 14px;text-align:left;font-size:11px;color:#888;text-transform:uppercase;">Categoría</th>
                <th style="padding:8px 14px;text-align:right;font-size:11px;color:#888;text-transform:uppercase;">Tu parte</th>
                <th style="padding:8px 14px;text-align:right;font-size:11px;color:#888;text-transform:uppercase;">% del total</th>
            </tr>
        </thead>
        <tbody>{cat_rows}</tbody>
    </table>
</div>
{particulares_html}

<!-- Estado de cuenta -->
<div style="background:#fff;border-radius:12px;margin-top:16px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.05);">
    <h3 style="margin:0 0 14px;font-size:15px;font-weight:700;color:#111;">📒 Tu estado de cuenta</h3>
    <table style="width:100%;font-size:14px;">
        <tr><td style="padding:6px 0;color:#666;">Saldo anterior</td><td style="text-align:right;font-weight:600;">${float(prorrateo.get('saldo_anterior',0)):,.2f}</td></tr>
        <tr><td style="padding:6px 0;color:#666;">Tu pago registrado</td><td style="text-align:right;font-weight:600;color:#10B981;">-${float(prorrateo.get('pago_realizado',0)):,.2f}</td></tr>
        <tr><td style="padding:6px 0;color:#666;">Saldo pendiente</td><td style="text-align:right;font-weight:600;color:#EF4444;">${float(prorrateo.get('saldo_pendiente',0)):,.2f}</td></tr>
        <tr><td style="padding:6px 0;color:#666;">Intereses</td><td style="text-align:right;font-weight:600;">${float(prorrateo.get('interes_mora',0)):,.2f}</td></tr>
        <tr style="border-top:2px solid #eee;">
            <td style="padding:10px 0;font-weight:700;">Expensa ordinaria ({pct_a:.3f}%)</td>
            <td style="text-align:right;font-weight:700;">${float(prorrateo.get('expensa_a',0)):,.2f}</td>
        </tr>
        <tr><td style="padding:6px 0;color:#666;">Adicional ordinaria</td><td style="text-align:right;font-weight:600;">${float(prorrateo.get('adicional_ordinaria',0)):,.2f}</td></tr>
        <tr><td style="padding:6px 0;color:#666;">Gastos de tu unidad</td><td style="text-align:right;font-weight:600;">${gastos_particulares:,.2f}</td></tr>
        <tr style="border-top:2px solid #eee;">
            <td style="padding:10px 0;font-weight:700;">Total a pagar</td>
            <td style="text-align:right;font-weight:800;">${total_unidad:,.2f}</td>
        </tr>
    </table>
</div>

<!-- Datos de pago -->
<div style="background:#fff;border-radius:12px;margin-top:16px;padding:20px;box-shadow:0 2px 8px rgba(0,0,0,.05);">
    <h3 style="margin:0 0 14px;font-size:15px;font-weight:700;color:#111;">💳 Datos de pago</h3>
    <table style="width:100%;font-size:14px;">
        <tr><td style="padding:5px 0;color:#666;">Banco</td><td style="text-align:right;font-weight:500;">{banco_nombre}</td></tr>
        <tr><td style="padding:5px 0;color:#666;">CBU</td><td style="text-align:right;font-weight:600;font-family:monospace;font-size:13px;">{banco_cbu}</td></tr>
        <tr><td style="padding:5px 0;color:#666;">1er vencimiento</td><td style="text-align:right;font-weight:600;">{vto1}</td></tr>
        <tr><td style="padding:5px 0;color:#666;">2do vencimiento</td><td style="text-align:right;font-weight:500;">{vto2} (+{interes_2}%)</td></tr>
    </table>
    <p style="margin:12px 0 0;font-size:12px;color:#888;text-align:center;">📧 Recordá enviar tu comprobante de pago por email o por la plataforma.</p>
</div>

<!-- Fondo del consorcio (se omite en una liquidación complementaria) -->
{fondo_html}

{obras_html}

<!-- Footer -->
<div style="text-align:center;margin-top:24px;padding:16px;">
    <p style="font-size:12px;color:#aaa;">Generado por Niddo — Gestión de consorcios inteligente</p>
    <p style="font-size:11px;color:#ccc;">{escape(liq.get('notas', ''))}</p>
</div>

</div>
</body>
</html>'''


@app.route('/api/liquidaciones/<lid>/resumen/<uid>')
@require_auth(allowed_roles=['admin'])
def api_liquidacion_resumen(lid, uid):
    """Genera y devuelve el resumen HTML personalizado de una UF."""
    liq = liquidacion_propia(lid, '*, consorcios(nombre, direccion, cuit, clave_suterh, '
                                  'tasa_interes_mora, recargo_segundo_vto, banco_nombre, '
                                  'banco_sucursal, banco_cuenta, banco_cbu, banco_cuit_pago, '
                                  'banco_titular, banco_alias)')

    consorcio = liq.get('consorcios', {})

    prorrateo = supabase.table('liquidacion_prorrateo') \
        .select('*, unidades_funcionales(numero, piso, tipo, vecino_nombre, vecino_email)') \
        .eq('liquidacion_id', lid).eq('unidad_id', uid).single().execute().data
    if not prorrateo:
        return jsonify({'error': 'Prorrateo no encontrado para esta UF'}), 404

    uf = prorrateo.get('unidades_funcionales', {})

    rubros = _rubros_con_items(lid)

    html = _generar_resumen_html(liq, prorrateo, rubros, consorcio, uf)

    if request.args.get('format') == 'html':
        return Response(html, mimetype='text/html')
    return jsonify({'html': html, 'uf': uf, 'total': prorrateo.get('total_unidad')})


# ── Envío de resúmenes por email ───────────────────────────────────────────────

def _nombre_pdf_liquidacion(liq, consorcio):
    """Nombre del adjunto. Lo primero que ve el vecino en su bandeja.

    Sin tildes ni ñ a propósito: es un nombre de archivo que viaja por mail y
    se guarda en cualquier sistema de archivos, y "expensas-mío.pdf" llega
    roto a más clientes de los que uno esperaría.
    """
    import unicodedata
    crudo = consorcio.get('nombre') or 'consorcio'
    sin_tildes = ''.join(c for c in unicodedata.normalize('NFKD', crudo)
                         if not unicodedata.combining(c))
    base = re.sub(r'[^a-zA-Z0-9\s-]', '', sin_tildes).strip()
    base = re.sub(r'\s+', '-', base).lower() or 'consorcio'
    rev = int(liq.get('numero_revision') or 1)
    sufijo = f'-rev{rev}' if rev > 1 else ''
    return f"expensas-{base}-{liq.get('periodo', '')}{sufijo}.pdf"


def _pdf_de_liquidacion(liq, consorcio, lid):
    """Arma el PDF completo de la liquidación con lo que hay en la base."""
    from liquidacion_pdf import construir_pdf

    admin = supabase.table('administradores') \
        .select('nombre, nombre_contacto, email, email_contacto, telefono, '
                'cuit, matricula, direccion') \
        .eq('id', liq.get('admin_id')).execute().data
    admin = admin[0] if admin else {}
    datos_admin = {
        'nombre': admin.get('nombre_contacto') or admin.get('nombre') or '',
        'cuit': admin.get('cuit') or '',
        'matricula': admin.get('matricula') or '',
        'direccion': admin.get('direccion') or '',
        'email': admin.get('email_contacto') or admin.get('email') or '',
        'telefono': admin.get('telefono') or '',
    }

    rubros = _rubros_con_items(lid)
    prorrateos = supabase.table('liquidacion_prorrateo') \
        .select('*, unidades_funcionales(numero, piso, vecino_nombre)') \
        .eq('liquidacion_id', lid).execute().data or []
    prorrateos.sort(key=lambda p: str((p.get('unidades_funcionales') or {}).get('numero') or ''))

    return construir_pdf(liq, consorcio, datos_admin, rubros, prorrateos)


def _generar_cobros_de_liquidacion(liq, prorrateos):
    """Convierte el prorrateo en la deuda que el vecino ve en su cuenta.

    Este era el agujero grande del circuito. La liquidación calculaba el total
    por unidad y se lo mandaba por mail, pero `cobros` —que es lo que alimenta
    Mis Expensas, el cupón, la morosidad y el balance— se llenaba desde un modal
    aparte donde el administrador tipeaba UN monto plano igual para todas las
    unidades. Los dos números podían no coincidir y nada los ataba.

    Es idempotente por el índice único (liquidacion_id, unidad_id): emitir dos
    veces no duplica la deuda. Un chequeo en Python no cerraría la ventana entre
    el SELECT y el INSERT con dos pestañas abiertas.

    Devuelve cuántos cobros creó.
    """
    if not prorrateos:
        return 0

    lid = liq.get('id')
    ya = supabase.table('cobros').select('unidad_id') \
        .eq('liquidacion_id', lid).execute().data or []
    existentes = {c['unidad_id'] for c in ya}

    filas = []
    for pr in prorrateos:
        uid = pr.get('unidad_id')
        if not uid or uid in existentes:
            continue
        total = float(pr.get('total_unidad') or 0)
        interes = float(pr.get('interes_mora') or 0)
        filas.append({
            'unidad_id': uid,
            'consorcio_id': liq.get('consorcio_id'),
            'liquidacion_id': lid,
            'periodo': liq.get('periodo'),
            # El monto base es lo del período; el interés viaja aparte para que
            # el cupón pueda mostrarlos separados, como hacen los cuatro.
            'monto_base': round(total - interes, 2),
            'interes_mora': interes,
            'total': total,
            'estado': 'pendiente',
            'fecha_vencimiento': liq.get('fecha_vencimiento_1'),
        })

    if not filas:
        return 0
    try:
        supabase.table('cobros').insert(filas).execute()
    except Exception as e:
        # El índice único rebota los que ya estaban: es el resultado buscado.
        if 'duplicate key' in str(e) or '23505' in str(e):
            return 0
        raise
    return len(filas)


@app.route('/api/liquidaciones/<lid>/pdf')
@require_auth(allowed_roles=['admin'])
def api_liquidacion_pdf(lid):
    """El mismo PDF que se adjunta al mail, para revisarlo antes de enviarlo."""
    liq = liquidacion_propia(lid, '*, consorcios(nombre, direccion, cuit, clave_suterh, '
                                  'tasa_interes_mora, recargo_segundo_vto, banco_nombre, '
                                  'banco_sucursal, banco_cuenta, banco_cbu, banco_cuit_pago, '
                                  'banco_titular, banco_alias)')
    consorcio = liq.get('consorcios') or {}
    buf = io.BytesIO(_pdf_de_liquidacion(liq, consorcio, lid))
    return pdf_response(buf, _nombre_pdf_liquidacion(liq, consorcio))


@app.route('/api/liquidaciones/<lid>/enviar', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_liquidacion_enviar(lid):
    """Envía resúmenes por email a todas las UFs (o las seleccionadas)."""
    d = request.json or {}
    unidades_ids = d.get('unidades_ids')  # None / ausente = todas

    # Una lista vacía es "el admin destildó todo", no "mandale a todo el consorcio".
    if unidades_ids is not None and not unidades_ids:
        return jsonify({'error': 'Seleccioná al menos una unidad funcional para enviar'}), 400

    # El permiso se resuelve antes de tocar nada: es la ruta que manda mails a
    # cuarenta vecinos, y lo único que no se puede deshacer después.
    liq = liquidacion_propia(lid, '*, consorcios(nombre, direccion, cuit, clave_suterh, '
                                  'tasa_interes_mora, recargo_segundo_vto, banco_nombre, '
                                  'banco_sucursal, banco_cuenta, banco_cbu, banco_cuit_pago, '
                                  'banco_titular, banco_alias)')

    import resend
    resend.api_key = os.environ.get('RESEND_API_KEY', '')

    consorcio = liq.get('consorcios', {})

    # Obtener prorrateos
    q = supabase.table('liquidacion_prorrateo') \
        .select('*, unidades_funcionales(id, numero, piso, tipo, vecino_nombre, vecino_email)') \
        .eq('liquidacion_id', lid)
    if unidades_ids:
        q = q.in_('unidad_id', unidades_ids)
    prorrateos = q.execute().data

    # Obtener rubros con items
    rubros = _rubros_con_items(lid)

    # La deuda se crea antes de avisar. Si se hiciera después y fallara, el
    # vecino tendría el mail con un total que su cuenta no muestra.
    cobros_creados = _generar_cobros_de_liquidacion(liq, prorrateos)

    # El PDF se arma UNA vez para todo el consorcio, no uno por unidad: es la
    # liquidación entera y es idéntica para todos, así que generarla 40 veces
    # sería 40 veces el mismo trabajo dentro de una request que ya manda 40
    # mails. Si falla, los mails salen igual sin adjunto: el resumen del cuerpo
    # es el que el vecino lee, y quedarse sin avisar es peor que sin PDF.
    pdf_bytes = None
    try:
        pdf_bytes = _pdf_de_liquidacion(liq, consorcio, lid)
    except Exception:
        app.logger.exception('No se pudo generar el PDF de la liquidación %s', lid)

    enviados = 0
    fallidos = 0
    from_email = os.environ.get('RESEND_FROM_EMAIL', 'Niddo <noreply@niddo.app>')

    for prorrateo in prorrateos:
        uf = prorrateo.get('unidades_funcionales', {})
        email_destino = uf.get('vecino_email', '')

        html = _generar_resumen_html(liq, prorrateo, rubros, consorcio, uf)

        estado = 'enviado'
        error_detalle = None
        fecha_envio = now_iso()

        if email_destino and resend.api_key:
            try:
                periodo_display = liq.get('periodo', '')
                mensaje = {
                    'from': from_email,
                    'to': [email_destino],
                    'subject': (
                        # El asunto distingue la complementaria: es lo primero que ve el
                        # vecino y con el mismo texto los dos mails se confunden entre sí.
                        f'📋 Expensa complementaria — {consorcio.get("nombre", "")} — {periodo_display}'
                        if int(liq.get('numero_revision') or 1) > 1 else
                        f'📋 Resumen de expensas — {consorcio.get("nombre", "")} — {periodo_display}'
                    ),
                    'html': html,
                }
                if pdf_bytes:
                    mensaje['attachments'] = [{
                        'filename': _nombre_pdf_liquidacion(liq, consorcio),
                        'content': list(pdf_bytes),
                    }]
                resend.Emails.send(mensaje)
                enviados += 1
            except Exception as e:
                estado = 'fallido'
                error_detalle = str(e)
                fallidos += 1
        elif not email_destino:
            estado = 'fallido'
            error_detalle = 'Sin email configurado'
            fallidos += 1
        elif not resend.api_key:
            estado = 'fallido'
            error_detalle = 'RESEND_API_KEY no configurada'
            fallidos += 1

        # Registrar envío
        supabase.table('resumen_envios').insert({
            'liquidacion_id': lid,
            'unidad_id': uf.get('id'),
            'canal': 'email',
            'estado': estado,
            'email_destino': email_destino,
            'fecha_envio': fecha_envio if estado == 'enviado' else None,
            'error_detalle': error_detalle,
            'resumen_html': html,
        }).execute()

    # Actualizar estado de liquidación a publicada
    if enviados > 0:
        supabase.table('liquidaciones').update({'estado': 'publicada'}).eq('id', lid).execute()

    return jsonify({'enviados': enviados, 'fallidos': fallidos,
                    'total': len(prorrateos), 'cobros_generados': cobros_creados})


@app.route('/api/liquidaciones/<lid>/envios', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_liquidacion_envios(lid):
    liquidacion_propia(lid, 'id')
    res = supabase.table('resumen_envios') \
        .select('*, unidades_funcionales(numero, vecino_nombre, vecino_email)') \
        .eq('liquidacion_id', lid).order('created_at', desc=True).execute()
    return jsonify(res.data)


# ── Envío programado ───────────────────────────────────────────────────────────

@app.route('/api/envio-programado/<cid>', methods=['GET'])
@require_auth(allowed_roles=['admin'])
def api_envio_programado_get(cid):
    consorcio_propio(cid, 'id')
    res = supabase.table('envio_programado').select('*') \
        .eq('consorcio_id', cid).limit(1).execute()
    return jsonify(res.data[0] if res.data else {})


@app.route('/api/envio-programado/<cid>', methods=['POST'])
@require_auth(allowed_roles=['admin'])
def api_envio_programado_set(cid):
    admin_id = get_admin_id()
    consorcio_propio(cid, 'id')
    d = request.json
    payload = {
        'consorcio_id': cid,
        'admin_id': admin_id,
        'dia_mes': d.get('dia_mes', 1),
        'hora_envio': d.get('hora_envio', '09:00'),
        'canal': d.get('canal', 'email'),
        'activo': d.get('activo', True),
    }
    res = supabase.table('envio_programado').upsert(payload, on_conflict='consorcio_id').execute()
    return jsonify(res.data[0] if res.data else {})


# ── Run ────────────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    print('🏢 Niddo server starting...')
    print('📍 http://localhost:3500')
    app.run(host='127.0.0.1', port=3500, debug=True)

